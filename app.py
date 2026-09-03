"""마케팅 실적 현황 대시보드

위: EP 실적 (트래픽/거래액/구매객수/CR/객단가) — EP실적 데이터
아래: EP 채널 지표 (원부매칭율/최저가율 등) — 기존 EP 데이터, 원부매칭/최저가 필터 적용
"""
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import datetime as _dt
import io

from data_loader import (
    load_data, load_traffic_data, load_category_data, load_product_data, load_brand_names,
    load_coupon_daily, build_coupon_monthly, build_coupon_monthly_detail,
)
from ai_insights import (
    render_metric_insight, generate_ranking_insights, render_ranking_insight_box,
)
from sidebar import render_sidebar, render_sidebar_data_status
from filters import filter_by_combo
from kpi import render_kpi_cards
from charts import main_trend_data
from comparison_table import render_summary_table_html
from utils import (
    COL_DATE, COL_BPU, COL_MATCH, COL_LOWEST, METRIC_COLS, UNIT_CONFIG,
    resample_series, make_period_label, compute_kpi_deltas, week_of_month, raw_cutoff_date,
    effective_month_of_week,
    format_value, format_delta_html, format_delta_text, pct_delta_safe,
)
from utils import _match_mean, _partial_last_period
from styles import CUSTOM_CSS
from insight_card import render_insight_card
from dashboard_helpers import (
    _ref_str, _status_entry, BPU_GROUPS, _reset_date_range,
    render_week_range_filter, aggregate_traffic, FF_BRAND_CODE, exclude_ff_brand,
    exclude_ff_from_traffic, _weekly_of_year, _week_labels_for_year, _correct_partial_week_yoy,
    aggregate_ep, build_weekly_report_excel, build_forecast_excel, render_excel_download,
    _truncate_by_range, compute_bpu_comparison_rows, render_bpu_comparison_table, compute_category_yoy_rows,
    DASHBOARD_EVENTS, render_line_chart, FORECAST_BPU_ROWS, compute_monthly_forecast_series,
    build_forecast_table, _DIGIT_HAS_BATCHIM, _has_batchim, _emphasize,
    _josa_ga, _josa_eun, generate_rule_based_insights, generate_category_page_insights,
    render_monthly_comparison_table, render_insight_panel, render_donut_chart, compute_official_total,
    render_revenue_ranking, render_top_products,
)



# ============================================================
# 인사이트 카드: 좌측 자동요약(API 미사용) + 우측 AI·메모(GitHub 저장)
# ============================================================

st.set_page_config(page_title="마케팅 실적 현황 대시보드", layout="wide", page_icon="📊")
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# --- 사이드바 ---
side = render_sidebar()
if side["refresh"]:
    load_data.clear()
    load_traffic_data.clear()
    load_category_data.clear()
    load_coupon_daily.clear()

# --- 데이터 로드 ---
df_ep = load_data()           # 기존 EP 데이터 (원부매칭율 등)
df_traffic = load_traffic_data()  # EP실적 데이터 (트래픽/거래액 등)
df_category = load_category_data()  # 카테고리/브랜드별 실적 데이터
df_product = load_product_data()  # 상품(SKU)별 거래액 데이터 (카테고리별 상위 상품 랭킹용)
df_coupon_daily = load_coupon_daily()  # 쿠폰명별 일자별 상세 할인 데이터 (BPU: e-영업1~4)
df_coupon = build_coupon_monthly(df_coupon_daily)  # 월별 BPU 집계(Total/자사/입점/e-영업1~4) - daily에서 파생
df_coupon_detail = build_coupon_monthly_detail(df_coupon_daily)  # 월별 쿠폰명별 상세 - daily에서 파생
BRAND_NAMES = load_brand_names()  # 브랜드 코드 -> 한글 브랜드명


def brand_label(code):
    """브랜드 코드를 '코드 (브랜드명)' 형태로 표시. '전체'는 그대로."""
    if code == "전체" or code not in BRAND_NAMES:
        return code
    return f"{code} ({BRAND_NAMES[code]})"


BRAND_LABELS = {code: brand_label(code) for code in BRAND_NAMES}  # 랭킹 차트용 "코드 (이름)" 매핑

# EP채널 업로드
if side["ep_channel_file"] is not None:
    _uf = side["ep_channel_file"]
    df_ep = load_data(uploaded_file=_uf, file_name=getattr(_uf, "name", None))
    st.sidebar.success("EP채널 데이터 업로드 완료")

# EP실적 업로드
if side["ep_traffic_file"] is not None:
    _tf = side["ep_traffic_file"]
    df_traffic = pd.read_csv(_tf)
    df_traffic["날짜"] = pd.to_datetime(df_traffic["날짜"])
    _tr_min = df_traffic["날짜"].min().strftime("%Y-%m-%d")
    _tr_max = df_traffic["날짜"].max().strftime("%Y-%m-%d")
    st.sidebar.success(f"EP실적 데이터 업로드 완료: {_tr_min} ~ {_tr_max}")

# 카테고리 업로드
if side["ep_category_file"] is not None:
    _cf = side["ep_category_file"]
    df_category = pd.read_csv(_cf)
    df_category["날짜"] = pd.to_datetime(df_category["날짜"])
    st.sidebar.success(f"카테고리 데이터 업로드 완료 ({df_category['카테고리'].nunique()}개 카테고리)")

# 상품 데이터 업로드
if side["ep_product_file"] is not None:
    _pf = side["ep_product_file"]
    df_product = pd.read_csv(_pf)
    df_product["날짜"] = pd.to_datetime(df_product["날짜"])
    st.sidebar.success(f"상품 데이터 업로드 완료 ({df_product['상품코드'].nunique()}개 상품)")

# 쿠폰 데이터 업로드 (일자별 원본 하나로 통합 - 월별 집계/상세는 여기서 자동 파생)
if side["ep_coupon_daily_file"] is not None:
    _cpddf = side["ep_coupon_daily_file"]
    df_coupon_daily = pd.read_csv(_cpddf)
    df_coupon_daily["날짜"] = pd.to_datetime(df_coupon_daily["날짜"])
    df_coupon = build_coupon_monthly(df_coupon_daily)
    df_coupon_detail = build_coupon_monthly_detail(df_coupon_daily)
    st.sidebar.success(f"쿠폰 데이터 업로드 완료 ({df_coupon_daily['날짜'].min().strftime('%Y-%m-%d')} ~ {df_coupon_daily['날짜'].max().strftime('%Y-%m-%d')})")

# --- 사이드바 하단: 데이터셋별 반영 현황(기간/일수) ---

_status_items = [
    ("EP채널", *_status_entry(df_ep, COL_DATE)),
    ("EP실적", *_status_entry(df_traffic, "날짜")),
    ("카테고리", *_status_entry(df_category, "날짜")),
    ("상품", *_status_entry(df_product, "날짜")),
    ("쿠폰(일자별)", *_status_entry(df_coupon_daily, "날짜")),
]
render_sidebar_data_status(_status_items)


unit = side["view_unit"]

# 자사/입점 합산용 BPU 그룹 정의




































# 일자별 그래프에 표시할 이벤트 주석 (날짜, 라벨). 필요하면 여기에 추가/수정하면 됨.

































# 데이터 반영 현황
last_date_ep = df_ep[COL_DATE].max()
last_date_tr = df_traffic["날짜"].max()
_weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
st.sidebar.info(
    f"🗓️ EP실적: ~{last_date_tr.strftime('%m/%d')}({_weekday_kr[last_date_tr.weekday()]})\n\n"
    f"EP채널: ~{last_date_ep.strftime('%m/%d')}({_weekday_kr[last_date_ep.weekday()]})"
)

# --- 페이지 헤더 + 매체 필터 + 기준 시점 (스크롤 시 상단 고정) ---
_sticky = st.container(key="sticky_header")
_page_num = side["page"].split(".")[0]

with _sticky:
    # 고정 대상 식별용 마커 (반드시 이 컨테이너의 첫 요소여야 함)
    st.markdown("<div id='sticky-marker-anchor'></div>", unsafe_allow_html=True)

    BPU_OPTIONS = [
        ("전체", "Total"),
        ("e-영업1", "e-영업1"),
        ("e-영업2", "e-영업2"),
        ("e-영업3", "e-영업3"),
        ("e-영업4", "e-영업4"),
        ("자사 (e1+e2)", "자사"),
        ("입점 (e3+e4)", "입점"),
    ]

    _page_titles = {
        "1": "📊 실적 요약", "2": "🗂️ 카테고리 실적 요약", "3": "🧭 종합 요약",
        "4": "📋 누적 데이터", "5": "🏷️ 누적 데이터 (카테고리)",
        "6": "📅 전체 실적 (주차별)", "7": "👤 회원 실적 (주차별)", "8": "✨ 신규 실적 (주차별)",
        "9": "🎟️ 쿠폰 비용 분석", "10": "📈 마감 예상 실적", "11": "📑 주간보고용",
    }

    # ========================================================
    # 페이지 5: 쿠폰 비용 분석 — 매체/쿠폰유형/기준일자(또는 기준시점)를
    # 1·2번 페이지와 동일하게 상단 고정 필터로 올림
    # ========================================================
    if _page_num == "9":
        _cp_bpu_options_top = (
            [b for b in ["Total", "자사", "입점", "e-영업1", "e-영업2", "e-영업3", "e-영업4"] if b in df_coupon["BPU"].unique()]
            if not df_coupon.empty else []
        )
        coupon_unit = "월별" if unit in ("월별", "월마감") else unit

        if not _cp_bpu_options_top:
            st.markdown(
                f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>",
                unsafe_allow_html=True,
            )
        else:
            # 기준 시점 옵션: 매체/쿠폰유형 선택과 무관하게, 존재하는 날짜 전체를 기준으로 산출
            # (페이지1/2도 매체 선택 전에 Total 기준으로 기간 옵션부터 만드는 것과 동일한 방식)
            if coupon_unit == "월별":
                _cp_period_s = df_coupon.groupby("연월")["쿠폰할인"].sum().sort_index()
            else:
                _cp_rule_preview = "D" if coupon_unit == "일별" else "W-SUN"
                _cp_period_s = df_coupon_daily.groupby("날짜")["쿠폰할인"].sum().resample(_cp_rule_preview).sum()
                if coupon_unit == "주별":
                    _cp_period_s.index = _cp_period_s.index - pd.Timedelta(days=6)

            if coupon_unit == "일별":
                _cp_min_d = _cp_period_s.index.min().date()
                _cp_max_d = _cp_period_s.index.max().date()
                _cp_prev_date = st.session_state.get("coupon_ref_date", _cp_max_d)
                try:
                    _cp_ref_preview = pd.Timestamp(_cp_prev_date).strftime("%Y-%m-%d")
                except Exception:
                    _cp_ref_preview = _cp_max_d.strftime("%Y-%m-%d")
            else:
                _cp_period_labels = [
                    (d.strftime("%Y년 %m월") if coupon_unit == "월별" else f"{d.strftime('%Y-%m-%d')} 주")
                    for d in _cp_period_s.index
                ]
                _cp_default_label = _cp_period_labels[-1] if _cp_period_labels else ""
                _cp_prev_label_sel = st.session_state.get("coupon_ref_period", _cp_default_label)
                _cp_ref_preview = _cp_prev_label_sel if _cp_prev_label_sel in _cp_period_labels else _cp_default_label

            st.markdown(
                f"<div style='display:flex;align-items:baseline;gap:10px;flex-wrap:wrap;margin-bottom:6px;'>"
                f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>"
                f"<span style='font-size:0.8rem;color:#6b7280;'>조회 단위: <b>{coupon_unit}</b> · 기준: <b>{_cp_ref_preview}</b></span>"
                f"</div>",
                unsafe_allow_html=True,
            )

            fc1, fc2, fc3, _fc_spacer = st.columns([1, 1, 1, 5])
            with fc1:
                st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>매체</div>", unsafe_allow_html=True)
                coupon_bpu = st.selectbox("매체", _cp_bpu_options_top, index=0, key="coupon_bpu", label_visibility="collapsed")
            with fc2:
                st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>쿠폰 유형</div>", unsafe_allow_html=True)
                coupon_type_sel = st.radio(
                    "쿠폰유형", ["합산", "플러스", "일반"], horizontal=True, key="coupon_type_sel", label_visibility="collapsed",
                )
            with fc3:
                _cp_label3 = "기준 일자" if coupon_unit == "일별" else "기준 시점"
                st.markdown(f"<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>{_cp_label3}</div>", unsafe_allow_html=True)
                if coupon_unit == "일별":
                    _cp_sel_date = st.date_input(
                        "기준 일자", value=_cp_max_d, min_value=_cp_min_d, max_value=_cp_max_d,
                        label_visibility="collapsed", key="coupon_ref_date",
                    )
                    coupon_ref_ts = pd.Timestamp(_cp_sel_date)
                    if coupon_ref_ts not in _cp_period_s.index:
                        _cand = _cp_period_s.index[_cp_period_s.index <= coupon_ref_ts]
                        coupon_ref_ts = _cand[-1] if len(_cand) else _cp_period_s.index[-1]
                else:
                    _cp_sel_label = st.selectbox(
                        "기준 시점", _cp_period_labels, index=len(_cp_period_labels) - 1,
                        label_visibility="collapsed", key="coupon_ref_period",
                    )
                    coupon_ref_ts = _cp_period_s.index[_cp_period_labels.index(_cp_sel_label)]

    # ========================================================
    # 페이지 1 / 2: 매체필터 + 기준시점 (+카테고리/브랜드)
    # ========================================================
    elif _page_num in ("1", "2"):
        # 기준 시점 옵션: 페이지1은 트래픽(EP실적) 데이터, 페이지2는 카테고리 데이터 기준으로 생성
        # (두 데이터의 최신 날짜가 다를 수 있어, 실제 존재하는 기간만 선택지로 제공)
        if _page_num == "2" and not df_category.empty:
            _period_base_df = df_category[(df_category["카테고리"] == "전체") & (df_category["브랜드"] == "전체")]
            if "회원구분" in _period_base_df.columns:
                _period_base_df = _period_base_df[_period_base_df["회원구분"] == "전체"]
            _period_metric_col = "트래픽"
        else:
            _period_base_df = df_traffic[(df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")]
            _period_metric_col = "트래픽"

        _period_s = _period_base_df.set_index("날짜")[_period_metric_col].resample(UNIT_CONFIG[unit]["rule"]).mean().dropna()
        if unit == "주별":
            _period_s.index = _period_s.index - pd.Timedelta(days=6)
        if unit == "월마감":
            _last_base_date = _period_base_df["날짜"].max()
            if not _period_s.empty and _last_base_date < _period_s.index[-1]:
                _period_s = _period_s.iloc[:-1]  # 미완성 달 제외

        if unit == "일별":
            _min_d = _period_s.index.min().date()
            _max_d = _period_s.index.max().date()
        else:
            _period_labels = [make_period_label(d, unit) for d in _period_s.index]

        # 위젯 렌더 전, 세션 상태로 현재 선택값을 미리 파악해 제목 옆에 표시
        if unit == "일별":
            _prev_date = st.session_state.get("period_filter_date", _max_d)
            try:
                _period_label_preview = make_period_label(pd.Timestamp(_prev_date), unit)
            except Exception:
                _period_label_preview = make_period_label(pd.Timestamp(_max_d), unit)
        else:
            _default_label = _period_labels[-1] if _period_labels else ""
            _prev_label_sel = st.session_state.get("period_filter", _default_label)
            _period_label_preview = _prev_label_sel if _prev_label_sel in _period_labels else _default_label

        st.markdown(
            f"<div style='display:flex;align-items:baseline;gap:10px;flex-wrap:wrap;margin-bottom:6px;'>"
            f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>"
            f"<span style='font-size:0.8rem;color:#6b7280;'>조회 단위: <b>{unit}</b> · 기준: <b>{_period_label_preview}</b></span>"
            f"</div>",
            unsafe_allow_html=True,
        )

        _is_cat_page = _page_num == "2"
        _ff_exclude = False  # 기본값 (아래에서 조건에 맞으면 덮어씀)
        if _is_cat_page:
            fc1, fc2, fc3, fc4, fc5, _fc_spacer = st.columns([1, 1, 1, 1, 1, 5])
        else:
            fc1, fc2, fc3, _fc_spacer = st.columns([1, 1, 1, 7])

        with fc1:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>매체 필터</div>", unsafe_allow_html=True)
            _bpu_label_sel = st.selectbox(
                "매체 필터", [l for l, _ in BPU_OPTIONS],
                label_visibility="collapsed", key="bpu_filter",
            )
            bpu = dict(BPU_OPTIONS)[_bpu_label_sel]

        with fc2:
            _label2 = "기준 일자" if unit == "일별" else "기준 시점"
            st.markdown(f"<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>{_label2}</div>", unsafe_allow_html=True)
            if unit == "일별":
                _sel_date = st.date_input(
                    "기준 일자", value=_max_d, min_value=_min_d, max_value=_max_d,
                    label_visibility="collapsed", key="period_filter_date",
                )
                selected_period_date = pd.Timestamp(_sel_date)
                if selected_period_date not in _period_s.index:
                    _cand = _period_s.index[_period_s.index <= selected_period_date]
                    selected_period_date = _cand[-1] if len(_cand) else _period_s.index[-1]
            else:
                _sel_label = st.selectbox(
                    "기준 시점", _period_labels, index=len(_period_labels) - 1,
                    label_visibility="collapsed", key="period_filter",
                )
                selected_period_date = _period_s.index[_period_labels.index(_sel_label)]

        # 1번 페이지(카테고리 필터가 없는 페이지)는 매체필터/기준시점과 같은 줄 fc3에 핏플랍 제외 배치
        # (2번 페이지는 카테고리/브랜드 뒤 fc5에 배치 — 두 페이지 다 '마지막 필터 바로 옆' 위치로 통일)
        if not _is_cat_page:
            if (not df_category.empty) and (df_category["브랜드"] == "FF").any():
                with fc3:
                    st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>&nbsp;</div>", unsafe_allow_html=True)
                    _ff_exclude = st.checkbox(
                        "핏플랍 제외",
                        value=st.session_state.get("cat_ff_exclude", False), key="cat_ff_exclude",
                        help="핏플랍은 2025년 10월에 종료된 브랜드예요. 켜면 카테고리 원본(ep_category.csv)에서 "
                             "FF 실적을 찾아 EP실적(트래픽/거래액/구매객수)에서도 빼고 CR/객단가를 다시 계산해요. "
                             "2번 페이지의 체크박스와 같은 설정을 공유해요.",
                    )
            else:
                _ff_exclude = False

        # 카테고리 페이지일 때만 매체필터 옆에 카테고리/브랜드 필터 노출
        selected_cat, selected_brand = "전체", "전체"
        cat_segment = "전체"
        if _is_cat_page and not df_category.empty:
            # 매체필터(bpu) 기준으로 데이터 범위 결정 (자사/입점은 합산)
            if bpu in BPU_GROUPS:
                _cat_bpu_df_preview = df_category[df_category["BPU"].isin(BPU_GROUPS[bpu])]
            elif bpu == "Total":
                _cat_bpu_df_preview = df_category
            else:
                _cat_bpu_df_preview = df_category[df_category["BPU"] == bpu]

            with fc3:
                st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>카테고리</div>", unsafe_allow_html=True)
                _valid_cats_top = (
                    _cat_bpu_df_preview.dropna(subset=["트래픽", "거래액", "구매객수"], how="all")
                    .loc[lambda d: (d["트래픽"] > 0) | (d["거래액"] > 0) | (d["구매객수"] > 0), "카테고리"]
                    .unique()
                )
                _cat_options_top = ["전체"] + sorted([c for c in _valid_cats_top if c != "전체"])
                selected_cat = st.selectbox("카테고리", _cat_options_top, index=0, label_visibility="collapsed", key="cat_select")

            with fc4:
                st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>브랜드</div>", unsafe_allow_html=True)
                _cat_filtered_top = _cat_bpu_df_preview[_cat_bpu_df_preview["카테고리"] == selected_cat]
                _valid_brands_top = (
                    _cat_filtered_top.dropna(subset=["트래픽", "거래액", "구매객수"], how="all")
                    .loc[lambda d: (d["트래픽"] > 0) | (d["거래액"] > 0) | (d["구매객수"] > 0), "브랜드"]
                    .unique()
                )
                _brand_options_top = ["전체"] + sorted([b for b in _valid_brands_top if b != "전체"])
                selected_brand = st.selectbox("브랜드", _brand_options_top, index=0, format_func=brand_label, label_visibility="collapsed", key="brand_select")

            # 핏플랍(FF) 제외 — 브랜드 필터 옆에 배치, FF 브랜드 데이터가 실제로 있을 때만 노출
            with fc5:
                if (df_category["브랜드"] == "FF").any():
                    st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>&nbsp;</div>", unsafe_allow_html=True)
                    _ff_exclude = st.checkbox(
                        "핏플랍 제외",
                        value=False, key="cat_ff_exclude",
                        help="핏플랍은 2025년 10월에 종료된 브랜드예요. 켜면 슈즈 카테고리·전체 집계에서 "
                             "FF 실적을 빼고 CR/객단가까지 다시 계산해서 보여줘요 (트래픽/구매객수도 같이 빠짐).",
                    )
                else:
                    _ff_exclude = False

            # 세그먼트(고객 구분) — 카테고리 레벨(브랜드=전체)에서만 제공
            _has_segment = "회원구분" in df_category.columns
            if _has_segment and selected_brand == "전체":
                _cat_seg_options = [s for s in ["전체", "회원", "비회원", "신규", "기존"] if s in df_category["회원구분"].unique()]
            else:
                _cat_seg_options = ["전체"]
            if len(_cat_seg_options) > 1:
                st.markdown("<div style='margin-top:6px;'></div>", unsafe_allow_html=True)
                cat_segment = st.radio(
                    "고객 구분", _cat_seg_options, horizontal=True,
                    key="cat_seg_filter", label_visibility="collapsed",
                )
            else:
                cat_segment = "전체"
                if _has_segment and selected_brand != "전체":
                    st.caption("ℹ️ 브랜드별 데이터는 전체 세그먼트만 제공됩니다.")

        # 페이지1(실적요약)일 때는 EP실적용 세그먼트(고객 구분) 필터만 노출 (핏플랍 제외는 위 fc3에서 처리됨)
        if _page_num == "1":
            _seg_options = [s for s in ["전체", "회원", "비회원", "신규", "기존"] if s in df_traffic["회원구분"].unique()]
            st.markdown("<div style='margin-top:6px;'></div>", unsafe_allow_html=True)
            segment = st.radio(
                "고객 구분", _seg_options, horizontal=True,
                key="seg_filter", label_visibility="collapsed",
            )

        period_label = make_period_label(selected_period_date, unit)

    # ========================================================
    # 페이지 6: 주간보고용은 자체 필터(기준 시점)를 페이지 본문에서 처리하므로,
    # 여기서는 간단한 제목만 표시
    # ========================================================
    elif _page_num == "11":
        st.markdown(
            f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>",
            unsafe_allow_html=True,
        )

    # ========================================================
    # 페이지 10: 종합 요약 — 1/2번 페이지와 동일한 스타일로 기준시점(전역 조회단위 기준)
    # + 세그먼트를 상단 고정 영역에 배치 (별도 조회단위 선택은 안 둠 — 사이드바 것과
    # 중복돼서 혼란스러웠던 걸 반영해 제거)
    # ========================================================
    elif _page_num == "3":
        _sum_period_base = df_traffic[(df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")]
        if _sum_period_base.empty:
            _sum_period_base = df_traffic[df_traffic["BPU"] == "Total"]
        _sum_period_s = (
            _sum_period_base.set_index("날짜")["트래픽"].resample(UNIT_CONFIG[unit]["rule"]).mean().dropna()
        )
        if unit == "주별":
            _sum_period_s.index = _sum_period_s.index - pd.Timedelta(days=6)
        if unit == "월마감":
            _sum_last_base_date = _sum_period_base["날짜"].max()
            if not _sum_period_s.empty and _sum_last_base_date < _sum_period_s.index[-1]:
                _sum_period_s = _sum_period_s.iloc[:-1]

        _sum_period_labels = [make_period_label(d, unit) for d in _sum_period_s.index]
        _sum_default_label = _sum_period_labels[-1] if _sum_period_labels else ""
        _sum_prev_label_sel = st.session_state.get("sum_period_filter", _sum_default_label)
        _sum_period_preview = _sum_prev_label_sel if _sum_prev_label_sel in _sum_period_labels else _sum_default_label
        _sum_prev_segment = st.session_state.get("sum_segment_filter", "전체")

        st.markdown(
            f"<div style='display:flex;align-items:baseline;gap:10px;flex-wrap:wrap;margin-bottom:6px;'>"
            f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>"
            f"<span style='font-size:0.8rem;color:#6b7280;'>조회 단위: <b>{unit}</b> · 기준: <b>{_sum_period_preview}</b>"
            f" · 세그먼트: <b>{_sum_prev_segment}</b></span>"
            f"</div>",
            unsafe_allow_html=True,
        )

        _sfc1, _sfc2, _sfc_spacer = st.columns([1.3, 1.3, 6])
        with _sfc1:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>기준시점</div>", unsafe_allow_html=True)
            if _sum_period_labels:
                _sum_sel_label = st.selectbox(
                    "기준시점", _sum_period_labels, index=len(_sum_period_labels) - 1,
                    label_visibility="collapsed", key="sum_period_filter",
                )
                sum_selected_period_date = _sum_period_s.index[_sum_period_labels.index(_sum_sel_label)]
            else:
                sum_selected_period_date = df_traffic["날짜"].max() if not df_traffic.empty else pd.Timestamp.today()
        with _sfc2:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>세그먼트</div>", unsafe_allow_html=True)
            sum_segment = st.radio(
                "세그먼트", ["전체", "회원", "신규"], horizontal=True,
                label_visibility="collapsed", key="sum_segment_filter",
            )

    # ========================================================
    # 페이지 11: 마감 예상 실적 — 연도 하나만 고르면 되는 단순한 필터라 셀렉트박스 하나만
    # ========================================================
    elif _page_num == "10":
        _fc_years = sorted(df_traffic["날짜"].dt.year.unique().tolist(), reverse=True) if not df_traffic.empty else [pd.Timestamp.today().year]
        st.markdown(
            f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>",
            unsafe_allow_html=True,
        )
        _fc_col1, _fc_spacer = st.columns([1, 6])
        with _fc_col1:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>연도</div>", unsafe_allow_html=True)
            forecast_year = st.selectbox("연도", _fc_years, index=0, label_visibility="collapsed", key="forecast_year_filter")

    # ========================================================
    # 페이지 7/8/9: 전체·회원·신규 실적(주차별) — 주차 범위 필터를 상단 고정 영역에 배치
    # (세 페이지가 세그먼트만 다르고 주차 필터 UI는 동일해서 하나로 묶음)
    # ========================================================
    elif _page_num in ("6", "7", "8"):
        st.markdown(
            f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>",
            unsafe_allow_html=True,
        )
        if df_traffic.empty:
            wk_range = (1, 1)
        else:
            # 25년/26년 각각의 실제 데이터로 라벨을 만들고(두 해가 다르게 밀릴 수 있어서),
            # 더 긴 쪽(보통 25년, 최대 53주)을 기준으로 슬라이더를 그린다.
            # _week_labels_for_year가 _weekly_of_year와 정확히 같은 순서로 라벨을 만들어서
            # 인덱스가 어긋나지 않는다.
            _wk7_total = df_traffic[df_traffic["BPU"] == "Total"]
            _wk7_labels_25 = _week_labels_for_year(_wk7_total, "트래픽", 2025)
            _wk7_labels_26 = _week_labels_for_year(_wk7_total, "트래픽", 2026)
            _wk7_labels = _wk7_labels_25 if len(_wk7_labels_25) >= len(_wk7_labels_26) else _wk7_labels_26
            if not _wk7_labels:
                _wk7_labels = ["1주차"]

            # 기본값 = 최근 10주. '최근'은 26년(진행 중인 올해) 기준 마지막 주차로 잡고,
            # 26년 데이터가 없으면 라벨 목록 전체의 마지막 10주로 대체한다.
            _wk7_latest_wk = len(_wk7_labels_26) if _wk7_labels_26 else len(_wk7_labels)
            _wk7_default_end = max(1, min(_wk7_latest_wk, len(_wk7_labels)))
            _wk7_default_start = max(1, _wk7_default_end - 9)

            _wk7_c1, _wk7_spacer = st.columns([2, 3])
            with _wk7_c1:
                st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>주차 범위</div>", unsafe_allow_html=True)
                _wk7_sel_labels = st.select_slider(
                    "주차 범위", options=_wk7_labels,
                    value=(_wk7_labels[_wk7_default_start - 1], _wk7_labels[_wk7_default_end - 1]),
                    label_visibility="collapsed", key="wk_traffic_range",
                )
            wk_range = (_wk7_labels.index(_wk7_sel_labels[0]) + 1, _wk7_labels.index(_wk7_sel_labels[1]) + 1)

    # ========================================================
    # 페이지 3 / 4: 매체필터 + 기간유형 + 시작일/종료일 (+카테고리/브랜드)
    # ========================================================
    else:
        CUM_UNIT_OPTIONS = [("일자별", "일별"), ("주차별", "주별"), ("월별", "월별"), ("월마감", "월마감")]

        _cum_bpu_prev = st.session_state.get("cum_bpu_filter", "전체")
        _cum_unit_prev = st.session_state.get("cum_unit_filter", "일자별")
        _cum_agg_prev = st.session_state.get("cum_agg_mode", "일평균")
        st.markdown(
            f"<div style='display:flex;align-items:baseline;gap:10px;flex-wrap:wrap;margin-bottom:6px;'>"
            f"<span style='font-size:1.15rem;font-weight:700;'>{_page_titles[_page_num]}</span>"
            f"<span style='font-size:0.8rem;color:#6b7280;'>매체: <b>{_cum_bpu_prev}</b> · 기간유형: <b>{_cum_unit_prev}</b> · 집계: <b>{_cum_agg_prev}</b></span>"
            f"</div>",
            unsafe_allow_html=True,
        )

        _is_cum_cat_page = _page_num == "5"
        gc1, gc2, gc3, gc4, gc5, gc6, _gc_spacer = st.columns([1, 1, 1, 1, 1, 1, 2])

        with gc1:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>매체</div>", unsafe_allow_html=True)
            _cum_bpu_label = st.selectbox(
                "매체", [l for l, _ in BPU_OPTIONS],
                label_visibility="collapsed", key="cum_bpu_filter",
            )
            bpu = dict(BPU_OPTIONS)[_cum_bpu_label]

        with gc2:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>기간 유형</div>", unsafe_allow_html=True)
            _cum_unit_label = st.selectbox(
                "기간 유형", [l for l, _ in CUM_UNIT_OPTIONS],
                label_visibility="collapsed", key="cum_unit_filter",
            )
            cum_unit = dict(CUM_UNIT_OPTIONS)[_cum_unit_label]

        with gc3:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>집계 방식</div>", unsafe_allow_html=True)
            if cum_unit == "월마감":
                # 월마감은 '완료된 달의 총 실적'을 보는 기준이라 일평균은 의미가 없어서 누적으로 고정.
                # (일별/주별/월별에서는 그대로 일평균/누적 토글 가능)
                cum_agg_mode = "누적"
                st.markdown(
                    "<div style='padding:0.4rem 0.6rem;background:#f3f4f6;border-radius:6px;"
                    "font-size:0.85rem;color:#374151;'>누적 <span style='color:#9ca3af;font-size:0.72rem;'>"
                    "(월마감은 항상 누적)</span></div>",
                    unsafe_allow_html=True,
                )
            else:
                cum_agg_mode = st.selectbox(
                    "집계 방식", ["일평균", "누적"],
                    label_visibility="collapsed", key="cum_agg_mode",
                )

        # 데이터 있는 전체 날짜 범위 (트래픽 데이터 기준)
        _cum_min_d = df_traffic["날짜"].min().date()
        _cum_max_d = df_traffic["날짜"].max().date()

        with gc4:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>시작일</div>", unsafe_allow_html=True)
            cum_start = st.date_input(
                "시작일", value=_cum_min_d, min_value=_cum_min_d, max_value=_cum_max_d,
                label_visibility="collapsed", key="cum_start_date",
            )
        with gc5:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>종료일</div>", unsafe_allow_html=True)
            cum_end = st.date_input(
                "종료일", value=_cum_max_d, min_value=_cum_min_d, max_value=_cum_max_d,
                label_visibility="collapsed", key="cum_end_date",
            )
        with gc6:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>정렬</div>", unsafe_allow_html=True)
            cum_sort_order = st.selectbox(
                "정렬", ["최신순", "오래된순"],
                label_visibility="collapsed", key="cum_sort_order",
            )

        # 카테고리 페이지(4번)는 둘째 줄에 카테고리/브랜드 필터 추가
        selected_cat, selected_brand = "전체", "전체"
        if _is_cum_cat_page and not df_category.empty:
            if bpu in BPU_GROUPS:
                _cat_bpu_df_preview = df_category[df_category["BPU"].isin(BPU_GROUPS[bpu])]
            elif bpu == "Total":
                _cat_bpu_df_preview = df_category
            else:
                _cat_bpu_df_preview = df_category[df_category["BPU"] == bpu]

            st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
            gc5, gc6, _gc_spacer2 = st.columns([1, 1, 8])

            with gc5:
                st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>카테고리</div>", unsafe_allow_html=True)
                _valid_cats_top = (
                    _cat_bpu_df_preview.dropna(subset=["트래픽", "거래액", "구매객수"], how="all")
                    .loc[lambda d: (d["트래픽"] > 0) | (d["거래액"] > 0) | (d["구매객수"] > 0), "카테고리"]
                    .unique()
                )
                _cat_options_top = ["전체"] + sorted([c for c in _valid_cats_top if c != "전체"])
                selected_cat = st.selectbox("카테고리", _cat_options_top, index=0, label_visibility="collapsed", key="cat_select")

            with gc6:
                st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>브랜드</div>", unsafe_allow_html=True)
                _cat_filtered_top = _cat_bpu_df_preview[_cat_bpu_df_preview["카테고리"] == selected_cat]
                _valid_brands_top = (
                    _cat_filtered_top.dropna(subset=["트래픽", "거래액", "구매객수"], how="all")
                    .loc[lambda d: (d["트래픽"] > 0) | (d["거래액"] > 0) | (d["구매객수"] > 0), "브랜드"]
                    .unique()
                )
                _brand_options_top = ["전체"] + sorted([b for b in _valid_brands_top if b != "전체"])
                selected_brand = st.selectbox("브랜드", _brand_options_top, index=0, format_func=brand_label, label_visibility="collapsed", key="brand_select")


        # 페이지3(누적 데이터)/4(카테고리)는 둘째 줄에 세그먼트(고객 구분) 필터 추가
        cum_segment = "전체"
        if _page_num == "4":
            _cum_seg_options = [s for s in ["전체", "회원", "비회원", "신규", "기존"] if s in df_traffic["회원구분"].unique()]
            if len(_cum_seg_options) > 1:
                st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
                cum_segment = st.radio(
                    "고객 구분", _cum_seg_options, horizontal=True,
                    key="cum_seg_filter", label_visibility="collapsed",
                )
        elif _page_num == "5":
            _has_cum_cat_segment = "회원구분" in df_category.columns
            if _has_cum_cat_segment and selected_brand == "전체":
                _cum_cat_seg_options = [s for s in ["전체", "회원", "비회원", "신규", "기존"] if s in df_category["회원구분"].unique()]
            else:
                _cum_cat_seg_options = ["전체"]
            if len(_cum_cat_seg_options) > 1:
                st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
                cum_segment = st.radio(
                    "고객 구분", _cum_cat_seg_options, horizontal=True,
                    key="cum_cat_seg_filter", label_visibility="collapsed",
                )
            elif _has_cum_cat_segment and selected_brand != "전체":
                st.caption("ℹ️ 브랜드별 데이터는 전체 세그먼트만 제공됩니다.")

# 필터 영역 상단 고정(fixed) CSS — position:fixed는 스크롤 컨테이너 구조와 무관하게 항상 화면에 고정됨
st.markdown(
    """
    <style>
    /* 방법 1: container key 클래스 직접 타겟팅 */
    .st-key-sticky_header {
        position: fixed !important;
        top: 3.7rem !important;
        left: 22rem !important;
        right: 2rem !important;
        z-index: 999 !important;
        background: #f7f8fa !important;
        padding: 6px 16px 8px 16px !important;
        border-bottom: 1px solid #e5e7eb !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.06) !important;
    }
    /* 방법 2: 마커 기반 :has() 셀렉터 (백업) */
    div[data-testid="stVerticalBlock"]:has(> div[data-testid="element-container"] > div#sticky-marker-anchor) {
        position: fixed !important;
        top: 3.7rem !important;
        left: 22rem !important;
        right: 2rem !important;
        z-index: 999 !important;
        background: #f7f8fa !important;
        padding: 6px 16px 8px 16px !important;
        border-bottom: 1px solid #e5e7eb !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.06) !important;
    }
    /* Streamlit 자체 상단 툴바(Share/GitHub 등)는 항상 최상단에 보이도록 z-index 우선 */
    header[data-testid="stHeader"] {
        z-index: 1000 !important;
    }
    /* 고정 영역 내부 위젯(selectbox/date_input) 상하 여백 축소 + 폭 제한(잘림 방지) */
    .st-key-sticky_header div[data-testid="stSelectbox"],
    .st-key-sticky_header div[data-testid="stDateInput"] {
        margin-bottom: -6px !important;
        max-width: 260px !important;
    }
    .st-key-sticky_header div[data-testid="stSelectbox"] > div,
    .st-key-sticky_header div[data-testid="stDateInput"] > div {
        max-width: 260px !important;
    }
    .st-key-sticky_header div[data-testid="element-container"] {
        margin-bottom: 0 !important;
    }
    /* '전년 비교선 표시' 체크박스를 각자 컬럼의 오른쪽 끝에 붙임 (기본은 왼쪽 정렬이라
       [3,1,1] 비율의 마지막 칸에서 체크박스와 라벨 뒤로 빈 공간이 남는 문제 해결).
       내부 label이 컨테이너 폭 전체로 늘어나면 justify-content가 무효화되므로
       label 자체도 콘텐츠 크기로 고정한다. */
    .st-key-tr_yoy, .st-key-ep_yoy_cb, .st-key-cat_yoy {
        width: 100% !important;
        display: flex !important;
        justify-content: flex-end !important;
    }
    .st-key-tr_yoy label, .st-key-ep_yoy_cb label, .st-key-cat_yoy label,
    .st-key-tr_yoy div[data-testid="stCheckbox"], .st-key-ep_yoy_cb div[data-testid="stCheckbox"],
    .st-key-cat_yoy div[data-testid="stCheckbox"] {
        flex: 0 0 auto !important;
        width: auto !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# 고정된 필터 영역이 차지하던 자리만큼, 아래 콘텐츠가 가려지지 않도록 여백 확보
if _page_num == "5":
    _spacer_height = 155
elif _page_num in ("1", "2", "4", "3", "10"):
    _spacer_height = 100
else:
    _spacer_height = 65
st.markdown(
    f"<div style='height:{_spacer_height}px;'></div>"
    "<div id='content-align-marker' style='height:0;'></div>",
    unsafe_allow_html=True,
)

# 사이드바 접힘/펼침/크기조정에 맞춰 고정 영역의 좌우 위치를 아래 콘텐츠와 정확히 일치시킴
components.html(
    """
    <script>
    function adjustStickyPosition() {
        try {
            const doc = window.parent.document;
            const marker = doc.querySelector('#content-align-marker');
            const stickyEls = doc.querySelectorAll('.st-key-sticky_header');
            if (marker && stickyEls.length) {
                const rect = marker.getBoundingClientRect();
                const rightPx = window.parent.innerWidth - rect.right;
                const PAD = 16; // 고정 영역 자체의 좌우 padding(px)과 동일한 값
                stickyEls.forEach(function(el) {
                    el.style.setProperty('left', (rect.left - PAD) + 'px', 'important');
                    el.style.setProperty('right', (rightPx - PAD) + 'px', 'important');
                });
            }
        } catch (e) {}
    }
    adjustStickyPosition();
    try {
        const obs = new MutationObserver(adjustStickyPosition);
        obs.observe(window.parent.document.body, {attributes: true, subtree: true, attributeFilter: ['style', 'class']});
        window.parent.addEventListener('resize', adjustStickyPosition);
    } catch (e) {}
    setInterval(adjustStickyPosition, 400); // 안전망: 주기적 재계산
    </script>
    """,
    height=0,
)



if side["page"].startswith("1."):
    # 핏플랍(FF) 제외 — 2번 페이지와 체크박스를 공유. df_traffic엔 브랜드 정보가 없어서
    # 카테고리 원본(df_category)에서 FF 기여분을 가져와 빼는 방식 (exclude_ff_from_traffic 참고)
    if _ff_exclude:
        df_traffic = exclude_ff_from_traffic(df_traffic, df_category)

    # ============================================================
    # 상단: EP 실적 (트래픽/거래액/구매객수/CR/객단가)
    # ============================================================
    st.markdown("---")
    st.markdown("### 📈 EP 실적")

    # 세그먼트 필터는 상단 고정 영역에서 이미 선택됨 (segment 변수 재사용)

    # 트래픽 데이터 필터 (자사/입점이면 합산)
    if bpu in BPU_GROUPS:
        tr_combo = aggregate_traffic(df_traffic, BPU_GROUPS[bpu], segment)
        tr_member = aggregate_traffic(df_traffic, BPU_GROUPS[bpu], "회원")
    else:
        tr_combo = df_traffic[(df_traffic["BPU"] == bpu) & (df_traffic["회원구분"] == segment)].copy()
        tr_member = df_traffic[(df_traffic["BPU"] == bpu) & (df_traffic["회원구분"] == "회원")].copy()

    if tr_combo.empty:
        st.warning(f"{bpu}의 EP실적 데이터가 없습니다.")
    else:
        # KPI 카드 (트래픽 지표 5개)
        TRAFFIC_METRICS = [
            ("트래픽", "EP UV"),
            ("거래액", "거래액(순결제)"),
            ("구매객수", "구매객수"),
            ("CR", "구매전환율(%)"),
            ("객단가", "객단가"),
        ]
        all_items = TRAFFIC_METRICS

        # 핏플랍(FF) 제외 중일 때, KPI 카드에 참고로 보여줄 FF 기여분 계산용 데이터 준비.
        # df_category에서 FF만 뽑아 지금 선택된 매체(bpu)+세그먼트 범위로 맞춘다.
        # (세그먼트가 카테고리 원본에 없는 비회원/기존이면 FF 기여분을 못 구해서 표기 생략)
        _ff_note_df = None
        if _ff_exclude and not df_category.empty and segment in ("전체", "회원", "신규"):
            _ff_raw = df_category[
                (df_category["브랜드"] == "FF") & (df_category["카테고리"] != "전체") & (df_category["회원구분"] == segment)
            ]
            if bpu in BPU_GROUPS:
                _ff_raw = _ff_raw[_ff_raw["BPU"].isin(BPU_GROUPS[bpu])]
            elif bpu != "Total":
                _ff_raw = _ff_raw[_ff_raw["BPU"] == bpu]
            if not _ff_raw.empty:
                _ff_note_df = _ff_raw.groupby("날짜", as_index=False)[["트래픽", "거래액", "구매객수"]].sum()

        # 1단계: 값/증감 먼저 계산 (AI 인사이트용 payload 구성)
        _kpi_computed = {}
        _ff_computed = {}
        for col_name, display_name in all_items:
            s = tr_combo.set_index("날짜")[col_name].sort_index()
            # 절대값 지표(트래픽/거래액/구매객수)는 월마감이면 합계, 비율 지표(CR/객단가)는 항상 평균
            _agg = "sum" if (unit == "월마감" and col_name in {"트래픽", "거래액", "구매객수"}) else "mean"
            series = s.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
            if unit == "주별":
                series.index = series.index - pd.Timedelta(days=6)
            elif unit == "월마감":
                if not series.empty and s.index.max() < series.index[-1]:
                    series = series.iloc[:-1]
            if not series.empty:
                series = series[series.index <= selected_period_date]
            _s_raw = s[s.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s
            stats = compute_kpi_deltas(series, unit, raw_daily=_s_raw)
            _kpi_computed[display_name] = (col_name, stats)

            # CR/객단가는 계산되는 값이라 FF 기여분을 따로 표기할 필요 없음 — 여기서 건너뜀
            _ff_stats = None
            if _ff_note_df is not None and col_name in ("트래픽", "거래액", "구매객수"):
                ff_s = _ff_note_df.set_index("날짜")[col_name].sort_index()
                # 메인 시리즈(s)와 날짜 범위를 맞춘다 (FF 매출 0인 날엔 원본에 행 자체가 없어서,
                # 안 맞추면 FF 시리즈만 마지막 날짜가 짧아져 '진행 중인 기간' 판정이 어긋나던 버그 있었음)
                ff_s = ff_s.reindex(s.index, fill_value=0)
                ff_series = ff_s.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
                if unit == "주별":
                    ff_series.index = ff_series.index - pd.Timedelta(days=6)
                elif unit == "월마감" and not ff_series.empty and ff_s.index.max() < ff_series.index[-1]:
                    ff_series = ff_series.iloc[:-1]
                ff_series = ff_series[ff_series.index <= selected_period_date] if not ff_series.empty else ff_series
                _ff_s_raw = ff_s[ff_s.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else ff_s
                _ff_stats = compute_kpi_deltas(ff_series, unit, raw_daily=_ff_s_raw)
            _ff_computed[display_name] = _ff_stats

        # 2단계: 인사이트 카드 (좌: 자동요약 / 우: AI·메모)
        _ai_context_ep = f"실적요약 · {bpu} · {segment} · {unit} · 기준 {period_label}"
        _ai_payload_ep = []
        cfg = UNIT_CONFIG[unit]
        for display_name, (col_name, stats) in _kpi_computed.items():
            if stats:
                _is_pct_tmp = col_name == "CR"
                _val_str_tmp = f"{stats['current']:.1f}%" if _is_pct_tmp else f"{stats['current']:,.0f}"
                _ai_payload_ep.append({
                    "name": display_name, "value": _val_str_tmp,
                    "prev_label": cfg["prev_label"], "prev_delta": float(stats["prev_delta"] or 0),
                    "yoy_label": cfg["yoy_label"], "yoy_delta": float(stats["yoy_delta"] or 0),
                })

        # KPI 카드뿐 아니라 하단 '사업부별 실적 비교'표 내용도 요약에 반영 —
        # 거래액 기준으로 e-영업1~4 중 가장 많이 오르고/내린 사업부를 뽑는다.
        # (render_bpu_comparison_table과 완전히 같은 함수를 재사용해서 숫자가 어긋나지 않음)
        _extra_sections_ep = []
        _bpu_rows_all = []
        try:
            _bpu_rows_all, _, _ = compute_bpu_comparison_rows(df_traffic, unit, selected_period_date)
            _bpu_gmv_rows = [
                r for r in _bpu_rows_all
                if r["metric_label"] == "거래액(순결제)" and r["bpu"] in ("e-영업1", "e-영업2", "e-영업3", "e-영업4") and r["stats"]
            ]
            if len(_bpu_gmv_rows) >= 2:
                _key_delta = lambda r: r["stats"].get("yoy_delta") if r["stats"].get("yoy_delta") is not None else r["stats"].get("prev_delta")
                _ranked = sorted([r for r in _bpu_gmv_rows if _key_delta(r) is not None], key=_key_delta, reverse=True)
                if _ranked:
                    _bpu_items = []
                    _top, _bottom = _ranked[0], _ranked[-1]
                    for _r, _tag in ((_top, "최대 상승"), (_bottom, "최대 하락")):
                        if _r is _top and _r is _bottom:
                            continue  # 사업부가 1개뿐이면 하나만
                        _bpu_items.append({
                            "name": f"{_r['bpu']} ({_tag})", "value": f"{_r['stats']['current']:,.0f}",
                            "yoy_label": cfg["yoy_label"], "yoy_delta": _r["stats"].get("yoy_delta"),
                            "prev_label": cfg["prev_label"], "prev_delta": _r["stats"].get("prev_delta"),
                        })
                    if _bpu_items:
                        _extra_sections_ep.append({"header": "사업부(BPU)별 거래액 비교", "items": _bpu_items})
        except Exception:
            pass  # 요약은 보조 기능이라, 계산 중 문제가 있어도 KPI 요약은 그대로 보여준다

        # 규칙 기반 자동 인사이트 — 이미 계산된 _bpu_rows_all/cfg를 그대로 재사용해서
        # 화면 숫자와 100% 일치하는 구조화된 요약을 만든다(LLM 호출 없음). render_insight_card의
        # expander 안에 같이 넣어서 한 번에 접고 펼 수 있게 한다.
        _rb_sections_ep = generate_rule_based_insights(_bpu_rows_all, cfg) if _bpu_rows_all else []

        _memo_key_ep = f"ep_summary::{bpu}::{segment}::{unit}::{period_label}"
        _ai_result_ep = render_insight_card(
            _ai_payload_ep, _ai_context_ep, "ep_summary", _memo_key_ep, period_label,
            extra_sections=_extra_sections_ep, rule_based_sections=_rb_sections_ep,
        )


        # 3단계: KPI 카드 렌더링 (+ 지표별 AI 한줄 인사이트)
        # 일별 이외(주별/월별/월마감)는 "이번 기간에 실제로 어느 날짜까지 들어있는지"가
        # 안 보이면 헷갈릴 수 있어서(예: 월별인데 이번 달이 아직 다 안 지났으면), 금년/전년
        # 정확한 날짜 범위를 캡션으로 보여준다. _partial_last_period는 '부분기간'일 때만
        # 날짜 목록을 주고 완성된 기간이면 None을 줘서(그건 그것대로 맞는 설계 — KPI
        # 계산 자체엔 완성된 기간의 날짜 목록이 필요 없어서), 여기서는 그거에 기대지 않고
        # 기간의 시작/끝을 직접 계산한다 — 완성된 기간이든 진행 중이든 항상 뜨게.
        if unit != "일별" and selected_period_date is not None and _s_raw is not None and not _s_raw.empty:
            _kc_s = (
                selected_period_date - pd.Timedelta(days=selected_period_date.weekday())
                if unit == "주별" else selected_period_date.replace(day=1)
            )
            _kc_e = min(_s_raw.index.max(), raw_cutoff_date(selected_period_date, unit))
            if _kc_s <= _kc_e:
                if unit == "월마감":
                    _kp_s, _kp_e = _kc_s - pd.DateOffset(years=1), _kc_e - pd.DateOffset(years=1)
                else:
                    _kp_s, _kp_e = _kc_s - pd.Timedelta(days=364), _kc_e - pd.Timedelta(days=364)

                def _md(d):
                    return f"{d.month}/{d.day}"

                _kpi_yoy_note = "전년 동월" if unit == "월마감" else "동요일 기준"
                st.markdown(
                    f"<div class='chart-caption'>📅 {_kc_s.year}년: {_md(_kc_s)}-{_md(_kc_e)}"
                    f" &nbsp;vs&nbsp; {_kp_s.year}년({_kpi_yoy_note}): {_md(_kp_s)}-{_md(_kp_e)}</div>",
                    unsafe_allow_html=True,
                )

        kpi_cols = st.columns(5)
        for i, (col_name, display_name) in enumerate(all_items):
            with kpi_cols[i]:
                _, stats = _kpi_computed[display_name]
                if stats:
                    _is_pct = col_name == "CR"
                    if _is_pct:
                        val_str = f"{stats['current']:.1f}%"
                    elif col_name == "객단가":
                        val_str = f"{stats['current']:,.0f}"
                    else:
                        val_str = f"{stats['current']:,.0f}"

                    cfg = UNIT_CONFIG[unit]

                    def _ff_note(v, _is_pct=_is_pct):
                        if v is None or pd.isna(v) or abs(v) < 0.5:
                            return ""
                        _s = f"{v:.1f}%" if _is_pct else f"{v:,.0f}"
                        return f" <span style='color:#ef4444;font-weight:600;'>[FF {_s}]</span>"

                    _ff_stats = _ff_computed.get(display_name)
                    _ff_cur_note = _ff_note(_ff_stats["current"]) if _ff_stats else ""
                    _ff_prev_note = _ff_note(_ff_stats.get("prev_value")) if _ff_stats else ""
                    _ff_avg_note = _ff_note(_ff_stats.get("avg_value")) if _ff_stats else ""
                    _ff_yoy_note = _ff_note(_ff_stats.get("yoy_value")) if _ff_stats else ""

                    st.markdown(
                        f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;min-height:180px;'>"
                        f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>{display_name}</div>"
                        f"<div style='font-size:1.5rem;font-weight:700;color:#111827;'>{val_str}</div>{_ff_cur_note}"
                        f"<div style='font-size:0.78rem;margin-top:6px;'>"
                        f"{cfg['prev_label']} {format_delta_html(stats['prev_delta'])}{_ref_str(stats.get('prev_value'), _is_pct)}{_ff_prev_note}<br/>"
                        f"{cfg['avg_label']} {format_delta_html(stats['avg_delta'])}{_ref_str(stats.get('avg_value'), _is_pct)}{_ff_avg_note}<br/>"
                        f"{cfg['yoy_label']} {format_delta_html(stats['yoy_delta'])}{_ref_str(stats.get('yoy_value'), _is_pct)}{_ff_yoy_note}"
                        f"</div></div>",
                        unsafe_allow_html=True,
                    )
                    render_metric_insight(_ai_result_ep, display_name)

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # 지표 추이 차트 (트래픽 지표)
        st.markdown("**EP 실적 추이**")
        tr_metric_options = ["트래픽", "거래액", "구매객수", "CR", "객단가"]
        tr_metric = st.radio(
            "지표 선택", tr_metric_options, index=0, key="tr_metric",
            horizontal=True, label_visibility="collapsed",
        )

        # 리샘플 (전체 기간 — 전년 비교선용)
        s_raw = tr_combo.set_index("날짜")[tr_metric].sort_index()
        _agg = "sum" if (unit == "월마감" and tr_metric in {"트래픽", "거래액", "구매객수"}) else "mean"
        tr_full = s_raw.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
        if unit == "주별":
            tr_full.index = tr_full.index - pd.Timedelta(days=6)
        elif unit == "월마감" and not tr_full.empty and s_raw.index.max() < tr_full.index[-1]:
            tr_full = tr_full.iloc[:-1]

        # 올해만 추출
        latest_year = int(tr_full.index.max().year)
        tr_series = tr_full[tr_full.index.year == latest_year]
        # 표/KPI 카드는 selected_period_date까지만 자르는데 이 차트는 최신 연도 데이터를
        # 끝까지 다 보여주고 있어서 표랑 차트 마지막 지점이 다른 값을 가리키는 버그가
        # 있었음(2번 페이지에서 먼저 발견됨) — 여기도 동일하게 자른다.
        if selected_period_date is not None and not tr_series.empty:
            tr_series = tr_series[tr_series.index <= selected_period_date]
            # 위에서 tr_full은 '전체 원본'을 먼저 리샘플(평균)한 뒤에 잘랐기 때문에, 원본이
            # 기준시점보다 더 뒤(예: 8/12)까지 있으면 마지막으로 남은 주(예: 8/10 라벨)의
            # 평균값 자체에 기준시점 이후 날짜(8/11,8/12)가 이미 섞여 들어가 있을 수 있다.
            # KPI 카드는 raw_daily를 기준시점까지 먼저 잘라서 이 문제가 없는데 차트만
            # 걸려있었음 — 마지막 지점이 진행 중인(부분) 기간이면 실제 원본을 기준시점까지
            # 자른 값으로 다시 계산해서 덮어쓴다.
            _tr_is_partial, _tr_cur_days, _ = _partial_last_period(
                s_raw[s_raw.index <= raw_cutoff_date(selected_period_date, unit)], unit
            )
            if _tr_is_partial and _tr_cur_days is not None and len(_tr_cur_days) > 0 and not tr_series.empty:
                _corrected = s_raw.loc[_tr_cur_days].mean()
                if pd.notna(_corrected):
                    tr_series.iloc[-1] = _corrected

        # 일별이면 최근 30일 + 기간 조정
        if unit == "일별":
            _default_start = max(tr_series.index.min().date(), tr_series.index.max().date() - _dt.timedelta(days=30))
            _tr_range_key = f"tr_range_{bpu}_{segment}"
            col_d, col_reset, col_y = st.columns([3, 1, 1])
            with col_d:
                dr = st.date_input("기간", value=(_default_start, tr_series.index.max().date()),
                                   min_value=tr_series.index.min().date(), max_value=tr_series.index.max().date(),
                                   key=_tr_range_key)
            with col_reset:
                st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
                st.button(
                    "🔄 최근으로", key=f"tr_range_reset_{bpu}_{segment}", use_container_width=True,
                    on_click=_reset_date_range, args=(_tr_range_key, (_default_start, tr_series.index.max().date())),
                )
            with col_y:
                show_tr_yoy = st.checkbox("전년 비교선 표시", value=True, key="tr_yoy")
            if isinstance(dr, tuple) and len(dr) == 2:
                tr_series = tr_series[(tr_series.index >= pd.Timestamp(dr[0])) & (tr_series.index <= pd.Timestamp(dr[1]))]
        else:
            if unit == "주별":
                col_wk, col_reset, col_y = st.columns([3, 1, 1])
                tr_series = render_week_range_filter(tr_series, f"tr_{bpu}_{segment}", col_wk, col_reset)
                with col_y:
                    st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
                    show_tr_yoy = st.checkbox("전년 비교선 표시", value=True, key="tr_yoy")
            else:
                _sp1, _sp2, col_y = st.columns([3, 1, 1])
                with col_y:
                    show_tr_yoy = st.checkbox("전년 비교선 표시", value=True, key="tr_yoy")

        chart_df = pd.DataFrame({tr_metric: tr_series})

        # 전년 비교선 (동요일 364일 / 월마감은 1년)
        yoy_col_name = None
        _tr_yoy_actual_dates = None
        if show_tr_yoy and not tr_series.empty:
            if unit == "월마감":
                prev_dates = tr_series.index - pd.DateOffset(years=1)
            else:
                prev_dates = tr_series.index - pd.Timedelta(days=364)
            yoy_vals = []
            yoy_actual = []
            # 마지막 지점이 진행 중인(부분) 주/월이면, 표(KPI카드)와 똑같이 '동요일 매칭
            # 평균'으로 계산한다 — 안 그러면 '이번 주 며칠치'를 '작년 그 주 전체 평균'과
            # 비교하는 격이라 차트 비교선이 표랑 다른 값을 보여주는 문제가 있었음.
            _is_partial, _cur_days, _ = _partial_last_period(
                s_raw[s_raw.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s_raw, unit
            )
            for i, pd_date in enumerate(prev_dates):
                if _is_partial and i == len(prev_dates) - 1 and _cur_days is not None:
                    _matched = _match_mean(s_raw, [d - pd.Timedelta(days=364) for d in _cur_days])
                    yoy_vals.append(_matched)
                    yoy_actual.append(pd_date)
                elif pd_date in tr_full.index:
                    yoy_vals.append(tr_full.loc[pd_date])
                    yoy_actual.append(pd_date)
                else:
                    cand = tr_full.index[tr_full.index <= pd_date]
                    yoy_vals.append(tr_full.loc[cand[-1]] if len(cand) else None)
                    yoy_actual.append(cand[-1] if len(cand) else pd_date)
            yoy_label = UNIT_CONFIG[unit]["yoy_label"]
            yoy_col_name = f"{yoy_label}(전년)"
            chart_df[yoy_col_name] = yoy_vals
            _tr_yoy_actual_dates = yoy_actual

        # 금년=진한 파랑, 전년=하늘색
        render_line_chart(chart_df, height=350, unit=unit, yoy_actual_dates=_tr_yoy_actual_dates)

        _tr_start = tr_series.index.min().strftime('%Y-%m-%d')
        _tr_end = tr_series.index.max().strftime('%Y-%m-%d')
        _yoy_note = ""
        if show_tr_yoy and not tr_series.empty:
            _yoy_s = prev_dates[0].strftime('%Y-%m-%d')
            _yoy_e = prev_dates[-1].strftime('%Y-%m-%d')
            _yoy_note = f"<br/>전년 비교: {_yoy_s} - {_yoy_e} (동요일 기준)"
        st.markdown(
            f"<div class='chart-caption'>올해: {_tr_start} - {_tr_end}{_yoy_note}</div>",
            unsafe_allow_html=True,
        )

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # 실적 요약 표 (트래픽 지표)
        st.markdown(f"**EP 실적 요약 표**  ·  <span style='color:#6b7280;font-size:0.85rem'>{bpu}</span>", unsafe_allow_html=True)
        body_rows = []
        prev_label = yoy_label = None
        for col_name, display_name in all_items:
            s = tr_combo.set_index("날짜")[col_name].sort_index()
            _agg = "sum" if (unit == "월마감" and col_name in {"트래픽", "거래액", "구매객수"}) else "mean"
            series = s.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
            if unit == "주별":
                series.index = series.index - pd.Timedelta(days=6)
            elif unit == "월마감" and not series.empty and s.index.max() < series.index[-1]:
                series = series.iloc[:-1]
            if not series.empty:
                series = series[series.index <= selected_period_date]
            _s_raw = s[s.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s
            stats = compute_kpi_deltas(series, unit, raw_daily=_s_raw)
            if stats is None:
                body_rows.append(f"<tr><td>{display_name}</td><td>-</td><td>-</td><td>-</td></tr>")
                continue
            prev_label = stats["prev_label"]
            yoy_label = stats["yoy_label"]
            is_pct = col_name == "CR"
            val = f"{stats['current']:.1f}%" if is_pct else f"{stats['current']:,.0f}"
            body_rows.append(
                f"<tr><td class='m'>{display_name}</td><td class='v'>{val}</td>"
                f"<td class='d'>{format_delta_html(stats['prev_delta'])}{_ref_str(stats.get('prev_value'), is_pct)}</td>"
                f"<td class='d'>{format_delta_html(stats['yoy_delta'])}{_ref_str(stats.get('yoy_value'), is_pct)}</td></tr>"
            )

        # 회원UV 행 추가 (선택된 세그먼트와 무관하게 항상 표시)
        if not tr_member.empty:
            s_mem = tr_member.set_index("날짜")["트래픽"].sort_index()
            _agg_mem = "sum" if unit == "월마감" else "mean"
            series_mem = s_mem.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg_mem)
            if unit == "주별":
                series_mem.index = series_mem.index - pd.Timedelta(days=6)
            elif unit == "월마감" and not series_mem.empty and s_mem.index.max() < series_mem.index[-1]:
                series_mem = series_mem.iloc[:-1]
            if not series_mem.empty:
                series_mem = series_mem[series_mem.index <= selected_period_date]
            _s_mem_raw = s_mem[s_mem.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s_mem
            stats_mem = compute_kpi_deltas(series_mem, unit, raw_daily=_s_mem_raw)
            if stats_mem is None:
                body_rows.append("<tr><td>회원UV</td><td>-</td><td>-</td><td>-</td></tr>")
            else:
                body_rows.append(
                    f"<tr><td class='m'>회원UV</td><td class='v'>{stats_mem['current']:,.0f}</td>"
                    f"<td class='d'>{format_delta_html(stats_mem['prev_delta'])}{_ref_str(stats_mem.get('prev_value'))}</td>"
                    f"<td class='d'>{format_delta_html(stats_mem['yoy_delta'])}{_ref_str(stats_mem.get('yoy_value'))}</td></tr>"
                )

        html = (
            "<table class='summary-table'>"
            f"<thead><tr><th>지표</th><th>값</th><th>{prev_label or '-'}</th><th>{yoy_label or '-'}</th></tr></thead>"
            f"<tbody>{''.join(body_rows)}</tbody></table>"
        )
        st.markdown(html, unsafe_allow_html=True)

        # --- 월마감일 때 누계 표시 ---
        if unit == "월마감" and not tr_combo.empty:
            latest_year = int(tr_combo["날짜"].max().year)
            # 선택 월이 있으면 그 월을 cutoff로, 없으면 마감 완료된 마지막 달
            _cutoff = selected_period_date
            _cutoff_month = _cutoff.month

            # 올해 누계 (1월 ~ 마감 월)
            ytd_cur = tr_combo[
                (tr_combo["날짜"] >= f"{latest_year}-01-01") &
                (tr_combo["날짜"] <= _cutoff)
            ]
            # 전년 동기간
            ytd_prev = tr_combo[
                (tr_combo["날짜"] >= f"{latest_year-1}-01-01") &
                (tr_combo["날짜"] <= f"{latest_year-1}-{_cutoff_month:02d}-{_cutoff.day:02d}")
            ]
            # 회원UV용

            st.markdown(f"<br/>", unsafe_allow_html=True)
            st.markdown(f"**📊 1~{_cutoff_month}월 누계**  ·  <span style='color:#6b7280;font-size:0.85rem'>전년 동기간 비교</span>", unsafe_allow_html=True)

            ytd_items = [
                ("트래픽", "EP UV", False),
                ("거래액", "거래액(순결제)", False),
                ("구매객수", "구매객수", False),
                ("_CR", "구매전환율(%)", True),
                ("_객단가", "객단가", False),
            ]

            ytd_rows = []
            for key, label, is_pct_ytd in ytd_items:
                if key == "_CR":
                    c_val = ytd_cur["구매객수"].sum() / ytd_cur["트래픽"].sum() * 100 if ytd_cur["트래픽"].sum() > 0 else 0
                    p_val = ytd_prev["구매객수"].sum() / ytd_prev["트래픽"].sum() * 100 if not ytd_prev.empty and ytd_prev["트래픽"].sum() > 0 else None
                    c_str = f"{c_val:.1f}%"
                    p_str = f"{p_val:.1f}%" if p_val else "-"
                elif key == "_객단가":
                    c_val = ytd_cur["거래액"].sum() / ytd_cur["구매객수"].sum() if ytd_cur["구매객수"].sum() > 0 else 0
                    p_val = ytd_prev["거래액"].sum() / ytd_prev["구매객수"].sum() if not ytd_prev.empty and ytd_prev["구매객수"].sum() > 0 else None
                    c_str = f"{c_val:,.0f}"
                    p_str = f"{p_val:,.0f}" if p_val else "-"

                else:
                    c_val = ytd_cur[key].sum()
                    p_val = ytd_prev[key].sum() if not ytd_prev.empty else None
                    c_str = f"{c_val:,.0f}"
                    p_str = f"{p_val:,.0f}" if p_val else "-"

                yoy_d = pct_delta_safe(c_val, p_val) if p_val and p_val != 0 else None
                ytd_rows.append(
                    f"<tr><td class='m'>{label}</td><td class='v'>{c_str}</td>"
                    f"<td class='v'>{p_str}</td>"
                    f"<td class='d'>{format_delta_html(yoy_d)}</td></tr>"
                )

            ytd_html = (
                "<table class='summary-table'>"
                f"<thead><tr><th>지표</th><th>{latest_year}년 누계</th>"
                f"<th>{latest_year-1}년 동기간</th><th>YoY</th></tr></thead>"
                f"<tbody>{''.join(ytd_rows)}</tbody></table>"
            )
            st.markdown(ytd_html, unsafe_allow_html=True)



    # --- 쿠폰 비용 요약 (최신월 기준, 상세는 5.쿠폰 비용 분석 페이지) ---
    if not df_coupon.empty:
        st.markdown("---")
        st.markdown("### 🎟️ 쿠폰 비용 요약", unsafe_allow_html=True)
        _cs_sub = df_coupon[df_coupon["BPU"] == "Total"]
        _cs_by_month = _cs_sub.groupby("연월")["쿠폰할인"].sum().sort_index()
        if not _cs_by_month.empty:
            _cs_latest = _cs_by_month.index.max()
            _cs_gmv = df_traffic[(df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")].set_index("날짜")["거래액"].resample("MS").sum()
            _cs_gmv_val = _cs_gmv.get(_cs_latest, None)
            _cs_coupon_val = _cs_by_month.get(_cs_latest, 0)
            _cs_rate = (_cs_coupon_val / _cs_gmv_val * 100) if _cs_gmv_val else None
            _cs_c1, _cs_c2, _cs_c3 = st.columns(3)
            with _cs_c1:
                st.markdown(
                    f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;'>"
                    f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>{_cs_latest.strftime('%Y년 %m월')} 쿠폰할인</div>"
                    f"<div style='font-size:1.4rem;font-weight:700;color:#111827;'>{_cs_coupon_val:,.0f}</div></div>",
                    unsafe_allow_html=True,
                )
            with _cs_c2:
                st.markdown(
                    f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;'>"
                    f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>거래액(Total)</div>"
                    f"<div style='font-size:1.4rem;font-weight:700;color:#111827;'>{_cs_gmv_val:,.0f}</div></div>"
                    if _cs_gmv_val else "<div></div>",
                    unsafe_allow_html=True,
                )
            with _cs_c3:
                st.markdown(
                    f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;'>"
                    f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>비용률</div>"
                    f"<div style='font-size:1.4rem;font-weight:700;color:#7c3aed;'>{_cs_rate:.2f}%</div></div>"
                    if _cs_rate is not None else "<div></div>",
                    unsafe_allow_html=True,
                )
            st.caption("👉 매체별·쿠폰유형별 상세, 쿠폰명 랭킹은 사이드바 '5. 쿠폰 비용 분석' 페이지에서 볼 수 있어요.")

    st.markdown("---")
    st.markdown("### 🏢 사업부별 실적 비교", unsafe_allow_html=True)
    _bpu_cfg = UNIT_CONFIG[unit]
    st.markdown(
        f"<div class='chart-caption'>Total / e-영업1~4 / 자사 / 입점 · "
        f"<b>{unit}</b> 기준 · 기준시점: <b>{period_label}</b> · "
        f"값·{_bpu_cfg['prev_label']}·{_bpu_cfg['avg_label']}·{_bpu_cfg['yoy_label']}</div>",
        unsafe_allow_html=True,
    )
    render_bpu_comparison_table(df_traffic, unit=unit, selected_period_date=selected_period_date)

    # ============================================================
    # 하단: EP 채널 지표 (원부매칭율/최저가율 등)
    # ============================================================
    st.markdown("---")
    st.markdown("### 🏷️ EP 채널 지표")

    # 원부매칭/최저가 필터
    from utils import COL_MATCH, COL_LOWEST
    c1, c2 = st.columns(2)
    match_options = [v for v in ["Total", "매칭"] if v in df_ep[COL_MATCH].unique()]
    lowest_options = [v for v in ["Total", "최저가"] if v in df_ep[COL_LOWEST].unique()]
    match_status = c1.selectbox("원부매칭여부", match_options, index=0, key="ep_match")
    lowest_status = c2.selectbox("최저가여부", lowest_options, index=0, key="ep_lowest")

    if bpu in BPU_GROUPS:
        df_ep_combo = aggregate_ep(df_ep, BPU_GROUPS[bpu], match_status, lowest_status)
    else:
        df_ep_combo = filter_by_combo(df_ep, bpu, match_status, lowest_status)

    if df_ep_combo.empty:
        st.warning("선택한 조합에 데이터가 없습니다.")
    else:
        # EP 채널 지표 KPI
        EP_CHANNEL_METRICS = [
            ("원부매칭율(%)", "원부매칭율(%)"),
            ("최저가율(%)", "최저가율(%)"),
            ("평균 EP 전시 상품수", "전시상품수"),
            ("평균 원부매칭 상품수", "원부매칭상품수"),
            ("평균 최저가 상품수", "최저가상품수"),
        ]

        ep_cols = st.columns(len(EP_CHANNEL_METRICS))
        for i, (metric_key, display_name) in enumerate(EP_CHANNEL_METRICS):
            with ep_cols[i]:
                series = resample_series(df_ep_combo, metric_key, unit).dropna()
                series = series[series.index <= selected_period_date]
                _ep_raw = df_ep_combo.set_index(COL_DATE)[metric_key].sort_index()
                _ep_raw = _ep_raw[_ep_raw.index <= raw_cutoff_date(selected_period_date, unit)]
                stats = compute_kpi_deltas(series, unit, raw_daily=_ep_raw)
                if stats:
                    _is_pct = "%" in metric_key or metric_key == "신규가입율"
                    val_str = f"{stats['current']:.1f}%" if _is_pct else f"{stats['current']:,.0f}"
                    cfg = UNIT_CONFIG[unit]
                    st.markdown(
                        f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;min-height:180px;'>"
                        f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>{display_name}</div>"
                        f"<div style='font-size:1.5rem;font-weight:700;color:#111827;'>{val_str}</div>"
                        f"<div style='font-size:0.78rem;margin-top:6px;'>"
                        f"{cfg['prev_label']} {format_delta_html(stats['prev_delta'])}{_ref_str(stats.get('prev_value'), _is_pct)}<br/>"
                        f"{cfg['avg_label']} {format_delta_html(stats['avg_delta'])}{_ref_str(stats.get('avg_value'), _is_pct)}<br/>"
                        f"{cfg['yoy_label']} {format_delta_html(stats['yoy_delta'])}{_ref_str(stats.get('yoy_value'), _is_pct)}"
                        f"</div></div>",
                        unsafe_allow_html=True,
                    )

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # EP 채널 지표 추이
        st.markdown("**EP 채널 추이**")
        ep_metrics_list = [m for m, _ in EP_CHANNEL_METRICS]
        ep_metric = st.radio(
            "지표", ep_metrics_list, index=0, key="ep_metric",
            horizontal=True, label_visibility="collapsed",
        )

        _ep_latest_year = int(last_date_ep.year)
        if unit == "일별":
            _ep_max_d = last_date_ep.date()
            _ep_min_d = _dt.date(_ep_latest_year, 1, 1)
            _ep_default_start = max(_ep_min_d, _ep_max_d - _dt.timedelta(days=30))
            _ep_range_key = f"ep_range_{bpu}_{match_status}_{lowest_status}"
            col_ed, col_ereset, col_ey = st.columns([3, 1, 1])
            with col_ed:
                ep_dr = st.date_input(
                    "기간", value=(_ep_default_start, _ep_max_d),
                    min_value=_ep_min_d, max_value=_ep_max_d, key=_ep_range_key,
                )
            with col_ereset:
                st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
                st.button(
                    "🔄 최근으로", key=f"ep_range_reset_{bpu}_{match_status}_{lowest_status}", use_container_width=True,
                    on_click=_reset_date_range, args=(_ep_range_key, (_ep_default_start, _ep_max_d)),
                )
            with col_ey:
                show_ep_yoy = st.checkbox("전년 비교선 표시", value=True, key="ep_yoy_cb")
            if isinstance(ep_dr, tuple) and len(ep_dr) == 2:
                _ep_date_start, _ep_date_end = ep_dr
            else:
                _ep_date_start, _ep_date_end = _ep_default_start, _ep_max_d
        else:
            _esp1, _esp2, col_ey = st.columns([3, 1, 1])
            with col_ey:
                show_ep_yoy = st.checkbox("전년 비교선 표시", value=True, key="ep_yoy_cb")
            _ep_date_start = _dt.date(_ep_latest_year, 1, 1)
            # 절대 최신 날짜(last_date_ep)가 아니라 선택된 기준시점까지만 —
            # 안 그러면 표/KPI 카드는 기준시점까지만 보여주는데 이 차트만 그 뒤 데이터까지
            # 더 보여줘서 마지막 지점 값이 서로 달라지는 버그가 있었음(다른 차트 2곳에서
            # 먼저 발견돼서 여기도 같은 패턴인지 점검함).
            _ep_date_end = selected_period_date.date() if selected_period_date is not None else last_date_ep.date()

        ep_trend, ep_yoy = main_trend_data(df_ep_combo, ep_metric, unit, show_yoy=show_ep_yoy,
                                           current_year=_ep_latest_year,
                                           date_start=_ep_date_start,
                                           date_end=_ep_date_end)

        ep_cols = list(ep_trend.columns)
        if len(ep_cols) > 1:
            render_line_chart(ep_trend, height=350, unit=unit)
        else:
            render_line_chart(ep_trend, height=350, unit=unit)

        st.markdown(
            f"<div class='chart-caption'>EP채널 데이터 · {bpu} / {match_status} / {lowest_status} 기준"
            f"{' · 전년 비교선(동요일) 포함' if show_ep_yoy else ''}</div>",
            unsafe_allow_html=True,
        )

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # --- EP 채널 요약 표 (EP실적 요약표와 동일 스타일 · 동일 비교 기준) ---
        st.markdown(f"**EP 채널 요약 표**  ·  <span style='color:#6b7280;font-size:0.85rem'>{bpu} / {match_status} / {lowest_status}</span>", unsafe_allow_html=True)

        ep_body_rows = []
        ep_prev_label = ep_yoy_label = None
        for metric_key, display_name in EP_CHANNEL_METRICS:
            series = resample_series(df_ep_combo, metric_key, unit)
            series = series[series.index <= selected_period_date] if not series.empty else series
            _ep_raw2 = df_ep_combo.set_index(COL_DATE)[metric_key].sort_index()
            _ep_raw2 = _ep_raw2[_ep_raw2.index <= raw_cutoff_date(selected_period_date, unit)]
            stats = compute_kpi_deltas(series, unit, raw_daily=_ep_raw2)
            if stats is None:
                ep_body_rows.append(f"<tr><td>{display_name}</td><td>-</td><td>-</td><td>-</td></tr>")
                continue
            ep_prev_label = stats["prev_label"]
            ep_yoy_label = stats["yoy_label"]
            _is_pct = "%" in metric_key or metric_key == "신규가입율"
            val = f"{stats['current']:.1f}%" if _is_pct else f"{stats['current']:,.0f}"
            ep_body_rows.append(
                f"<tr><td class='m'>{display_name}</td><td class='v'>{val}</td>"
                f"<td class='d'>{format_delta_html(stats['prev_delta'])}{_ref_str(stats.get('prev_value'), _is_pct)}</td>"
                f"<td class='d'>{format_delta_html(stats['yoy_delta'])}{_ref_str(stats.get('yoy_value'), _is_pct)}</td></tr>"
            )
        ep_summary_html = (
            "<table class='summary-table'>"
            f"<thead><tr><th>지표</th><th>값</th><th>{ep_prev_label or '-'}</th><th>{ep_yoy_label or '-'}</th></tr></thead>"
            f"<tbody>{''.join(ep_body_rows)}</tbody></table>"
        )
        st.markdown(ep_summary_html, unsafe_allow_html=True)


if side["page"].startswith("2."):
    # ============================================================
    # 카테고리별 실적 (카테고리 → 브랜드 드릴다운, 전년비교 가능)
    # ============================================================
    st.markdown("---")
    st.markdown("### 🗂️ 카테고리별 실적")

    if df_category.empty:
        st.info("카테고리 데이터가 없습니다. 사이드바에서 ep_category.csv를 업로드해주세요.")
    else:
        # 핏플랍(FF) 제외 여부(_ff_exclude)는 상단 고정 필터에서 이미 정해져서 넘어온다.

        # 매체필터(bpu)에 맞춰 카테고리 데이터 필터링 (자사/입점은 합산)
        if bpu in BPU_GROUPS:
            cat_bpu_df = df_category[df_category["BPU"].isin(BPU_GROUPS[bpu])]
        elif bpu == "Total":
            cat_bpu_df = df_category  # 전체 BPU 합산은 아래에서 groupby로 처리
        else:
            cat_bpu_df = df_category[df_category["BPU"] == bpu]

        # 세그먼트 필터는 상단 고정 영역에서 이미 선택됨 (cat_segment 변수 재사용)
        _has_segment = "회원구분" in df_category.columns

        # 세그먼트 필터 적용 전 원본 보관 (브랜드 랭킹은 세그먼트=전체 데이터만 있으므로 이걸 사용)
        cat_bpu_df_all_seg = cat_bpu_df

        if _has_segment:
            cat_bpu_df = cat_bpu_df[cat_bpu_df["회원구분"] == cat_segment]

        _ff_only_df = None
        if _ff_exclude:
            _ff_only_df = cat_bpu_df[(cat_bpu_df["브랜드"] == "FF") & (cat_bpu_df["카테고리"] != "전체")]
            cat_bpu_df = exclude_ff_brand(cat_bpu_df)
            cat_bpu_df_all_seg = exclude_ff_brand(cat_bpu_df_all_seg)

        # 카테고리별 거래액 요약(전기간비/평균비/전년비) — 아래쪽 '카테고리별 요약'표랑
        # 인사이트 카드(자동요약) 둘 다에서 쓸 거라 여기서 한 번만 계산해둔다.
        # 카테고리 선택 필터와 무관하게 전체 카테고리를 대상으로 함 (개요용이라서).
        _cat_summary_rows = []
        _cat_daily_df_early = cat_bpu_df[(cat_bpu_df["브랜드"] == "전체") & (cat_bpu_df["카테고리"] != "전체")]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _cat_daily_df_early = _cat_daily_df_early.groupby(["날짜", "카테고리"], as_index=False)["거래액"].sum()
        else:
            _cat_daily_df_early = _cat_daily_df_early[["날짜", "카테고리", "거래액"]]
        if not _cat_daily_df_early.empty:
            _cat_cfg_early = UNIT_CONFIG[unit]
            _cat_agg_early = "sum" if unit == "월마감" else "mean"
            for cat_name, g in _cat_daily_df_early.groupby("카테고리"):
                s = g.set_index("날짜")["거래액"].sort_index()
                series = s.resample(_cat_cfg_early["rule"]).agg(_cat_agg_early)
                if unit == "주별":
                    series.index = series.index - pd.Timedelta(days=6)
                elif unit == "월마감" and not series.empty and s.index.max() < series.index[-1]:
                    series = series.iloc[:-1]
                if not series.empty:
                    series = series[series.index <= selected_period_date]
                _s_raw = s[s.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s
                stats = compute_kpi_deltas(series, unit, raw_daily=_s_raw)
                if stats is None or not stats["current"]:
                    continue
                _cat_summary_rows.append({
                    "카테고리": cat_name, "거래액": stats["current"],
                    "prev": stats["prev_delta"], "avg": stats["avg_delta"], "yoy": stats["yoy_delta"],
                    "prev_v": stats.get("prev_value"), "avg_v": stats.get("avg_value"), "yoy_v": stats.get("yoy_value"),
                })
        _cat_summary_rows.sort(key=lambda r: r["거래액"], reverse=True)

        cat_combo = cat_bpu_df[(cat_bpu_df["카테고리"] == selected_cat) & (cat_bpu_df["브랜드"] == selected_brand)]
        if (bpu == "Total" or bpu in BPU_GROUPS) and not cat_combo.empty:
            cat_combo = cat_combo.groupby("날짜", as_index=False).agg({"트래픽": "sum", "거래액": "sum", "구매객수": "sum"})
            cat_combo["CR"] = (cat_combo["구매객수"] / cat_combo["트래픽"] * 100).where(cat_combo["트래픽"] > 0, 0)
            cat_combo["객단가"] = (cat_combo["거래액"] / cat_combo["구매객수"]).where(cat_combo["구매객수"] > 0, 0)

        if cat_combo.empty:
            st.warning(f"{selected_cat} / {brand_label(selected_brand)} 조합에 데이터가 없습니다.")
        else:
            st.markdown(
                f"<div class='chart-caption'>{bpu} · <b>{selected_cat}</b> / <b>{brand_label(selected_brand)}</b> 기준</div>",
                unsafe_allow_html=True,
            )

            # --- KPI 카드 ---
            CAT_METRICS = [
                ("트래픽", "UV"),
                ("거래액", "거래액"),
                ("구매객수", "구매객수"),
                ("CR", "구매전환율(%)"),
                ("객단가", "객단가"),
            ]

            # 핏플랍 제외 중이고, 지금 보는 화면이 실제로 FF의 영향을 받는 범위(카테고리=전체/슈즈,
            # 브랜드=전체)일 때만 'FF 기여분'을 계산해서 KPI 카드에 참고용으로 같이 보여준다.
            _ff_note_df = None
            if _ff_exclude and _ff_only_df is not None and not _ff_only_df.empty and selected_brand == "전체":
                if selected_cat == "전체":
                    _ff_scope = _ff_only_df
                else:
                    _ff_scope = _ff_only_df[_ff_only_df["카테고리"] == selected_cat]
                if not _ff_scope.empty:
                    _ff_note_df = _ff_scope.groupby("날짜", as_index=False).agg(
                        {"트래픽": "sum", "거래액": "sum", "구매객수": "sum"}
                    )
                    _ff_note_df["CR"] = (_ff_note_df["구매객수"] / _ff_note_df["트래픽"] * 100).where(_ff_note_df["트래픽"] > 0)
                    _ff_note_df["객단가"] = (_ff_note_df["거래액"] / _ff_note_df["구매객수"]).where(_ff_note_df["구매객수"] > 0)

            # 1단계: 값/증감 먼저 계산 (AI payload 구성용)
            _cat_computed = {}
            _ff_computed = {}
            for col_name, display_name in CAT_METRICS:
                s = cat_combo.set_index("날짜")[col_name].sort_index()
                _agg = "sum" if (unit == "월마감" and col_name in {"트래픽", "거래액", "구매객수"}) else "mean"
                series = s.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
                if unit == "주별":
                    series.index = series.index - pd.Timedelta(days=6)
                elif unit == "월마감" and not series.empty and s.index.max() < series.index[-1]:
                    series = series.iloc[:-1]
                series = series[series.index <= selected_period_date] if not series.empty else series
                _s_raw = s[s.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s
                _cat_computed[display_name] = (col_name, compute_kpi_deltas(series, unit, raw_daily=_s_raw))

                _ff_stats = None
                # CR/객단가는 트래픽·구매객수·거래액에서 계산되는 값이라, FF 기여분을 따로
                # 표기할 필요 없음(원본 3개 지표만 보면 충분) — 여기서 계산 자체를 건너뜀.
                if _ff_note_df is not None and col_name in ("트래픽", "거래액", "구매객수"):
                    ff_s = _ff_note_df.set_index("날짜")[col_name].sort_index()
                    # 메인 시리즈(s)와 날짜 범위를 맞춘다. FF는 매출 0인 날엔 원본에 행 자체가
                    # 없을 수 있어서, 그대로 두면 FF 시리즈의 마지막 날짜가 메인보다 짧아져
                    # '진행 중인 달' 판정(cur_days)이 메인과 어긋나는 문제가 있었음.
                    ff_s = ff_s.reindex(s.index, fill_value=0)
                    ff_series = ff_s.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
                    if unit == "주별":
                        ff_series.index = ff_series.index - pd.Timedelta(days=6)
                    elif unit == "월마감" and not ff_series.empty and ff_s.index.max() < ff_series.index[-1]:
                        ff_series = ff_series.iloc[:-1]
                    ff_series = ff_series[ff_series.index <= selected_period_date] if not ff_series.empty else ff_series
                    _ff_s_raw = ff_s[ff_s.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else ff_s
                    # compute_kpi_deltas를 그대로 재사용해서, 메인 지표와 정확히 같은 기준 시점
                    # (전기간/전년동기)의 FF 값을 뽑는다. 전년비교 기준연도(2025)에 FF가 활발했으므로,
                    # 지금(2026) 시점만 보면 0에 가깝지만 전년동기 쪽엔 값이 커질 수 있음.
                    _ff_stats = compute_kpi_deltas(ff_series, unit, raw_daily=_ff_s_raw)
                _ff_computed[display_name] = _ff_stats

            # 2단계: 인사이트 카드 (좌: 자동요약 / 우: AI·메모)
            _cfg_cat = UNIT_CONFIG[unit]
            _ai_payload_cat = []
            for display_name, (col_name, stats) in _cat_computed.items():
                if stats:
                    _is_pct_c = col_name == "CR"
                    _ai_payload_cat.append({
                        "name": display_name,
                        "value": f"{stats['current']:.1f}%" if _is_pct_c else f"{stats['current']:,.0f}",
                        "prev_label": _cfg_cat["prev_label"], "prev_delta": float(stats["prev_delta"] or 0),
                        "yoy_label": _cfg_cat["yoy_label"], "yoy_delta": float(stats["yoy_delta"] or 0),
                    })
            _ai_context_cat = (
                f"카테고리 실적 · {bpu} · {selected_cat}/{brand_label(selected_brand)} · {cat_segment} · "
                f"{unit} · 기준 {period_label}" + (" · 핏플랍제외" if _ff_exclude else "")
            )
            _memo_key_cat = (
                f"cat_summary::{bpu}::{selected_cat}::{selected_brand}::{cat_segment}::{unit}::{period_label}"
                + ("::ff제외" if _ff_exclude else "")
            )

            # KPI 카드뿐 아니라 하단 '카테고리별 요약'표 내용도 요약에 반영 —
            # 거래액 전년비 기준으로 가장 많이 오르고/내린 카테고리를 뽑는다.
            # (_cat_summary_rows는 위에서 이미 계산해둔 걸 그대로 씀 — 표와 숫자가 어긋나지 않음)
            _extra_sections_cat = []
            try:
                _cat_movers = [r for r in _cat_summary_rows if r.get("yoy") is not None]
                if len(_cat_movers) >= 2:
                    _ranked_cat = sorted(_cat_movers, key=lambda r: r["yoy"], reverse=True)
                    _top_c, _bottom_c = _ranked_cat[0], _ranked_cat[-1]
                    _cat_items = []
                    for _r, _tag in ((_top_c, "최대 상승"), (_bottom_c, "최대 하락")):
                        _cat_items.append({
                            "name": f"{_r['카테고리']} ({_tag})", "value": f"{_r['거래액']:,.0f}",
                            "yoy_label": _cfg_cat["yoy_label"], "yoy_delta": _r["yoy"],
                        })
                    if _cat_items:
                        _extra_sections_cat.append({"header": "카테고리별 거래액 톱무버", "items": _cat_items})
            except Exception:
                pass  # 요약은 보조 기능이라, 계산 중 문제가 있어도 KPI 요약은 그대로 보여준다

            _rb_sections_cat = generate_category_page_insights(_ai_payload_cat, _cfg_cat, _cat_summary_rows)

            _ai_result_cat = render_insight_card(
                _ai_payload_cat, _ai_context_cat, "cat_summary", _memo_key_cat, period_label,
                extra_sections=_extra_sections_cat, rule_based_sections=_rb_sections_cat,
            )

            # 3단계: KPI 카드 렌더링 (+ 지표별 AI 한줄 인사이트)
            cat_cols = st.columns(5)
            for i, (col_name, display_name) in enumerate(CAT_METRICS):
                with cat_cols[i]:
                    _, stats = _cat_computed[display_name]
                    if stats:
                        _is_pct = col_name == "CR"
                        val_str = f"{stats['current']:.1f}%" if _is_pct else f"{stats['current']:,.0f}"
                        cfg = UNIT_CONFIG[unit]

                        def _ff_note(v, _is_pct=_is_pct):
                            # 값이 거의 0이면(그 시점엔 FF 영향이 없으면) 표시 안 함 —
                            # 지금이 2026년이면 '현재'/전기간 쪽은 대부분 0, 전년비교 쪽(2025)에
                            # 큰 값이 나오는 게 정상 (FF는 2025-10월에 종료됐으므로)
                            if v is None or pd.isna(v) or abs(v) < 0.5:
                                return ""
                            _s = f"{v:.1f}%" if _is_pct else f"{v:,.0f}"
                            return f" <span style='color:#ef4444;font-weight:600;'>[FF {_s}]</span>"

                        _ff_stats = _ff_computed.get(display_name)
                        _ff_cur_note = _ff_note(_ff_stats["current"]) if _ff_stats else ""
                        _ff_prev_note = _ff_note(_ff_stats.get("prev_value")) if _ff_stats else ""
                        _ff_yoy_note = _ff_note(_ff_stats.get("yoy_value")) if _ff_stats else ""

                        st.markdown(
                            f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;min-height:180px;'>"
                            f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>{display_name}</div>"
                            f"<div style='font-size:1.4rem;font-weight:700;color:#111827;'>{val_str}</div>{_ff_cur_note}"
                            f"<div style='font-size:0.76rem;margin-top:6px;'>"
                            f"{cfg['prev_label']} {format_delta_html(stats['prev_delta'])}{_ref_str(stats.get('prev_value'), _is_pct)}{_ff_prev_note}<br/>"
                            f"{cfg['yoy_label']} {format_delta_html(stats['yoy_delta'])}{_ref_str(stats.get('yoy_value'), _is_pct)}{_ff_yoy_note}"
                            f"</div></div>",
                            unsafe_allow_html=True,
                        )
                        render_metric_insight(_ai_result_cat, display_name)
                    else:
                        st.markdown(
                            f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;min-height:160px;'>"
                            f"<div style='color:#6b7280;font-size:0.8rem;'>{display_name}</div>"
                            f"<div style='font-size:1.2rem;color:#9ca3af;'>-</div></div>",
                            unsafe_allow_html=True,
                        )

            st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

            # --- 추이 차트 (전년 비교선 포함) ---
            st.markdown("**카테고리 실적 추이**")
            cat_metric = st.radio(
                "지표", ["트래픽", "거래액", "구매객수", "CR", "객단가"],
                index=1, key="cat_metric", horizontal=True, label_visibility="collapsed",
            )

            s_raw = cat_combo.set_index("날짜")[cat_metric].sort_index()
            _agg = "sum" if (unit == "월마감" and cat_metric in {"트래픽", "거래액", "구매객수"}) else "mean"
            cat_full = s_raw.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
            if unit == "주별":
                cat_full.index = cat_full.index - pd.Timedelta(days=6)
            elif unit == "월마감" and not cat_full.empty and s_raw.index.max() < cat_full.index[-1]:
                cat_full = cat_full.iloc[:-1]

            latest_year_cat = int(cat_full.index.max().year) if not cat_full.empty else None
            cat_series = cat_full[cat_full.index.year == latest_year_cat] if latest_year_cat else cat_full
            # 표(카테고리 실적 요약 표)/KPI 카드는 selected_period_date까지만 자르는데,
            # 이 차트는 그걸 안 하고 최신 연도 데이터를 끝까지 다 보여주고 있어서 표랑 차트
            # 마지막 지점이 다른 값을 가리키는 버그가 있었음 — 여기서도 동일하게 자른다.
            if selected_period_date is not None and not cat_series.empty:
                cat_series = cat_series[cat_series.index <= selected_period_date]
                # 마지막 지점이 진행 중인(부분) 주/월이면, 원본이 기준시점보다 더 뒤까지
                # 있을 때 그 평균값에 기준시점 이후 날짜가 섞여 들어가 있을 수 있다(리샘플을
                # 먼저 하고 나중에 자르는 순서라서). KPI 카드처럼 실제 원본을 기준시점까지
                # 자른 값으로 마지막 지점을 다시 계산해서 덮어쓴다.
                _cat_is_partial2, _cat_cur_days2, _ = _partial_last_period(
                    s_raw[s_raw.index <= raw_cutoff_date(selected_period_date, unit)], unit
                )
                if _cat_is_partial2 and _cat_cur_days2 is not None and len(_cat_cur_days2) > 0 and not cat_series.empty:
                    _corrected2 = s_raw.loc[_cat_cur_days2].mean()
                    if pd.notna(_corrected2):
                        cat_series.iloc[-1] = _corrected2

            if unit == "일별" and not cat_series.empty:
                _cat_max_d = cat_series.index.max().date()
                _cat_min_d = cat_series.index.min().date()
                _cat_default_start = max(_cat_min_d, _cat_max_d - _dt.timedelta(days=30))
                # 카테고리/브랜드/매체 조합마다 별도 키를 써서, 다른 조합에서 수동으로 바꾼
                # 기간이 이어서 남지 않도록 함(예전엔 키가 고정돼 있어 다른 카테고리 선택시에도
                # 이전에 봤던 기간이 그대로 남아있는 문제가 있었음)
                _cat_range_key = f"cat_range_{bpu}_{selected_cat}_{selected_brand}"
                col_cd, col_reset, col_cy = st.columns([3, 1, 1])
                with col_cd:
                    cat_dr = st.date_input(
                        "기간", value=(_cat_default_start, _cat_max_d),
                        min_value=_cat_min_d, max_value=_cat_max_d, key=_cat_range_key,
                    )
                with col_reset:
                    st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
                    st.button(
                        "🔄 최근으로", key=f"cat_range_reset_{bpu}_{selected_cat}_{selected_brand}", use_container_width=True,
                        on_click=_reset_date_range, args=(_cat_range_key, (_cat_default_start, _cat_max_d)),
                    )
                with col_cy:
                    show_cat_yoy = st.checkbox("전년 비교선 표시", value=True, key="cat_yoy")
                if isinstance(cat_dr, tuple) and len(cat_dr) == 2:
                    cat_series = cat_series[(cat_series.index >= pd.Timestamp(cat_dr[0])) & (cat_series.index <= pd.Timestamp(cat_dr[1]))]
            else:
                if unit == "주별":
                    col_wk, col_reset, col_cy = st.columns([3, 1, 1])
                    cat_series = render_week_range_filter(
                        cat_series, f"cat_{bpu}_{selected_cat}_{selected_brand}", col_wk, col_reset
                    )
                    with col_cy:
                        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
                        show_cat_yoy = st.checkbox("전년 비교선 표시", value=True, key="cat_yoy")
                else:
                    _csp1, _csp2, col_cy = st.columns([3, 1, 1])
                    with col_cy:
                        show_cat_yoy = st.checkbox("전년 비교선 표시", value=True, key="cat_yoy")

            cat_chart_df = pd.DataFrame({cat_metric: cat_series})

            _cat_yoy_actual_dates = None
            if show_cat_yoy and not cat_series.empty:
                if unit == "월마감":
                    prev_dates = cat_series.index - pd.DateOffset(years=1)
                else:
                    prev_dates = cat_series.index - pd.Timedelta(days=364)
                yoy_vals = []
                yoy_actual = []
                # 마지막 지점이 진행 중인(부분) 주/월이면 표(카테고리 실적 요약 표)와 동일하게
                # '동요일 매칭 평균'으로 계산 (이유는 페이지1의 동일 로직 주석 참고)
                _cat_is_partial, _cat_cur_days, _ = _partial_last_period(
                    s_raw[s_raw.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s_raw, unit
                )
                for i, pd_date in enumerate(prev_dates):
                    if _cat_is_partial and i == len(prev_dates) - 1 and _cat_cur_days is not None:
                        _matched = _match_mean(s_raw, [d - pd.Timedelta(days=364) for d in _cat_cur_days])
                        yoy_vals.append(_matched)
                        yoy_actual.append(pd_date)
                    elif pd_date in cat_full.index:
                        yoy_vals.append(cat_full.loc[pd_date])
                        yoy_actual.append(pd_date)
                    else:
                        cand = cat_full.index[cat_full.index <= pd_date]
                        yoy_vals.append(cat_full.loc[cand[-1]] if len(cand) else None)
                        yoy_actual.append(cand[-1] if len(cand) else pd_date)
                yoy_label_cat = UNIT_CONFIG[unit]["yoy_label"]
                cat_chart_df[f"{yoy_label_cat}(전년)"] = yoy_vals
                _cat_yoy_actual_dates = yoy_actual
            render_line_chart(cat_chart_df, height=350, unit=unit, yoy_actual_dates=_cat_yoy_actual_dates)

            st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

            # --- 카테고리 실적 요약 표 (EP실적/EP채널과 동일 스타일·기준) ---
            st.markdown(
                f"**카테고리 실적 요약 표**  ·  <span style='color:#6b7280;font-size:0.85rem'>{bpu} · {selected_cat} / {brand_label(selected_brand)}</span>",
                unsafe_allow_html=True,
            )
            cat_summary_rows = []
            cat_prev_label = cat_yoy_label = None
            for col_name, display_name in CAT_METRICS:
                s2 = cat_combo.set_index("날짜")[col_name].sort_index()
                _agg2 = "sum" if (unit == "월마감" and col_name in {"트래픽", "거래액", "구매객수"}) else "mean"
                series2 = s2.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg2)
                if unit == "주별":
                    series2.index = series2.index - pd.Timedelta(days=6)
                elif unit == "월마감" and not series2.empty and s2.index.max() < series2.index[-1]:
                    series2 = series2.iloc[:-1]
                series2 = series2[series2.index <= selected_period_date] if not series2.empty else series2
                _s2_raw = s2[s2.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s2
                stats2 = compute_kpi_deltas(series2, unit, raw_daily=_s2_raw)
                if stats2 is None:
                    cat_summary_rows.append(f"<tr><td>{display_name}</td><td>-</td><td>-</td><td>-</td></tr>")
                    continue
                cat_prev_label = stats2["prev_label"]
                cat_yoy_label = stats2["yoy_label"]
                _is_pct2 = col_name == "CR"
                val2 = f"{stats2['current']:.1f}%" if _is_pct2 else f"{stats2['current']:,.0f}"
                cat_summary_rows.append(
                    f"<tr><td class='m'>{display_name}</td><td class='v'>{val2}</td>"
                    f"<td class='d'>{format_delta_html(stats2['prev_delta'])}{_ref_str(stats2.get('prev_value'), _is_pct2)}</td>"
                    f"<td class='d'>{format_delta_html(stats2['yoy_delta'])}{_ref_str(stats2.get('yoy_value'), _is_pct2)}</td></tr>"
                )
            cat_summary_html = (
                "<table class='summary-table'>"
                f"<thead><tr><th>지표</th><th>값</th><th>{cat_prev_label or '-'}</th><th>{cat_yoy_label or '-'}</th></tr></thead>"
                f"<tbody>{''.join(cat_summary_rows)}</tbody></table>"
            )
            st.markdown(cat_summary_html, unsafe_allow_html=True)

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # --- 카테고리별 거래액 비중 (브랜드=전체 기준, 선택 시점까지) ---
        _share_df = cat_bpu_df[(cat_bpu_df["브랜드"] == "전체") & (cat_bpu_df["카테고리"] != "전체")]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _share_df = _share_df.groupby(["날짜", "카테고리"], as_index=False)["거래액"].sum()

        # 진짜 전체값(카테고리=전체&브랜드=전체) — 개별 카테고리 합산이 아니라 이걸로 중앙에 표시
        _official_all_df = cat_bpu_df[(cat_bpu_df["카테고리"] == "전체") & (cat_bpu_df["브랜드"] == "전체")]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _official_all_df = _official_all_df.groupby("날짜", as_index=False)["거래액"].sum()
        _official_cat_total = compute_official_total(_official_all_df, unit, selected_period_date)

        render_revenue_ranking(_share_df, "카테고리", unit, selected_period_date, "카테고리별 거래액 비중", f"{bpu} 기준",
                               donut=True, official_total=_official_cat_total,
                               ai_key="cat_share", ai_context=f"카테고리별 거래액 비중 · {bpu} · {cat_segment} · {unit} · 기준 {period_label}" + (" · 핏플랍제외" if _ff_exclude else ""))

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # --- 브랜드별 거래액 랭킹 ---
        # 카테고리='전체'면 전체 브랜드 랭킹, 특정 카테고리 선택시 그 카테고리 안의 브랜드만
        # (브랜드 레벨 데이터는 세그먼트=전체만 존재하므로 cat_bpu_df_all_seg 사용)
        if selected_cat == "전체":
            _brand_share_df = cat_bpu_df_all_seg[(cat_bpu_df_all_seg["카테고리"] == "전체") & (cat_bpu_df_all_seg["브랜드"] != "전체")]
            _brand_subtitle = f"{bpu} · 전체 카테고리 기준"
        else:
            _brand_share_df = cat_bpu_df_all_seg[(cat_bpu_df_all_seg["카테고리"] == selected_cat) & (cat_bpu_df_all_seg["브랜드"] != "전체")]
            _brand_subtitle = f"{bpu} · {selected_cat} 카테고리 기준"
        if _has_segment:
            _brand_share_df = _brand_share_df[_brand_share_df["회원구분"] == "전체"]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _brand_share_df = _brand_share_df.groupby(["날짜", "브랜드"], as_index=False)["거래액"].sum()

        # 진짜 전체값(선택한 카테고리 범위 기준, 브랜드=전체) — 브랜드 합산이 아니라 이걸로 중앙에 표시
        _official_brand_scope_df = cat_bpu_df_all_seg[(cat_bpu_df_all_seg["카테고리"] == selected_cat) & (cat_bpu_df_all_seg["브랜드"] == "전체")]
        if _has_segment:
            _official_brand_scope_df = _official_brand_scope_df[_official_brand_scope_df["회원구분"] == "전체"]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _official_brand_scope_df = _official_brand_scope_df.groupby("날짜", as_index=False)["거래액"].sum()
        _official_brand_total = compute_official_total(_official_brand_scope_df, unit, selected_period_date)

        render_revenue_ranking(_brand_share_df, "브랜드", unit, selected_period_date, "브랜드별 거래액 랭킹", _brand_subtitle,
                               label_map=BRAND_LABELS, hide_zero=True,
                               donut=True, official_total=_official_brand_total,
                               ai_key="brand_rank", ai_context=f"브랜드별 거래액 랭킹 · {_brand_subtitle} · {unit} · 기준 {period_label}" + (" · 핏플랍제외" if _ff_exclude else ""))

        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

        # --- 카테고리별 상위 상품 (거래액) ---
        # 상품(SKU) 단위 데이터가 업로드된 경우에만 표시. BPU는 다른 섹션과 동일하게
        # BPU_GROUPS(자사/입점)로 합산하고, 카테고리는 상단에서 선택한 selected_cat을 그대로 씀
        # ('전체'면 전 카테고리 통합 랭킹, 특정 카테고리면 그 안에서의 상품 랭킹).
        if df_product.empty:
            st.caption("ℹ️ 상품 단위 데이터가 없어서 '카테고리별 상위 상품'은 건너뛰었어요 — 사이드바에서 ep_product.csv를 업로드하면 볼 수 있어요.")
        else:
            _prod_df = df_product
            if bpu in BPU_GROUPS:
                _prod_df = _prod_df[_prod_df["BPU"].isin(BPU_GROUPS[bpu])]
            elif bpu != "Total":
                _prod_df = _prod_df[_prod_df["BPU"] == bpu]
            if _ff_exclude:
                _prod_df = _prod_df[_prod_df["브랜드"] != "FF"]
            if selected_cat != "전체":
                _prod_df = _prod_df[_prod_df["카테고리"] == selected_cat]

            _prod_title = "카테고리별 상위 상품 (거래액)" if selected_cat == "전체" else f"{selected_cat} 상위 상품 (거래액)"
            render_top_products(
                _prod_df, unit, selected_period_date, _prod_title, _brand_subtitle,
                brand_label=brand_label,
            )

        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

        # --- 카테고리별 트래픽 비중 (거래액 랭킹과 동일 로직, 지표만 트래픽) ---
        # 거래액 섹션(카테고리+브랜드) 다음에 트래픽 섹션(카테고리+브랜드)이 오도록,
        # 지표 단위로 묶어서 배치한다 (레벨 단위로 번갈아 나오면 헷갈린다는 피드백 반영).
        _share_df_traffic = cat_bpu_df[(cat_bpu_df["브랜드"] == "전체") & (cat_bpu_df["카테고리"] != "전체")]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _share_df_traffic = _share_df_traffic.groupby(["날짜", "카테고리"], as_index=False)["트래픽"].sum()
        _official_all_df_traffic = cat_bpu_df[(cat_bpu_df["카테고리"] == "전체") & (cat_bpu_df["브랜드"] == "전체")]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _official_all_df_traffic = _official_all_df_traffic.groupby("날짜", as_index=False)["트래픽"].sum()
        _official_cat_total_traffic = compute_official_total(_official_all_df_traffic, unit, selected_period_date, metric_col="트래픽")

        render_revenue_ranking(_share_df_traffic, "카테고리", unit, selected_period_date, "카테고리별 트래픽 비중", f"{bpu} 기준",
                               donut=True, official_total=_official_cat_total_traffic,
                               metric_col="트래픽", metric_label="트래픽",
                               bar_color_cur="#ea580c", bar_color_prev="#fdba74",
                               donut_colors=["#ea580c", "#f97316", "#fb923c", "#fdba74", "#fed7aa",
                                             "#c2410c", "#9a3412", "#7c2d12", "#fef3c7", "#fde68a", "#e2e8f0"],
                               ai_key="cat_share_traffic", ai_context=f"카테고리별 트래픽 비중 · {bpu} · {cat_segment} · {unit} · 기준 {period_label}" + (" · 핏플랍제외" if _ff_exclude else ""))

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # --- 브랜드별 트래픽 랭킹 (거래액 랭킹과 동일 로직, 지표만 트래픽) ---
        if selected_cat == "전체":
            _brand_share_df_traffic = cat_bpu_df_all_seg[(cat_bpu_df_all_seg["카테고리"] == "전체") & (cat_bpu_df_all_seg["브랜드"] != "전체")]
        else:
            _brand_share_df_traffic = cat_bpu_df_all_seg[(cat_bpu_df_all_seg["카테고리"] == selected_cat) & (cat_bpu_df_all_seg["브랜드"] != "전체")]
        if _has_segment:
            _brand_share_df_traffic = _brand_share_df_traffic[_brand_share_df_traffic["회원구분"] == "전체"]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _brand_share_df_traffic = _brand_share_df_traffic.groupby(["날짜", "브랜드"], as_index=False)["트래픽"].sum()

        _official_brand_scope_df_traffic = cat_bpu_df_all_seg[(cat_bpu_df_all_seg["카테고리"] == selected_cat) & (cat_bpu_df_all_seg["브랜드"] == "전체")]
        if _has_segment:
            _official_brand_scope_df_traffic = _official_brand_scope_df_traffic[_official_brand_scope_df_traffic["회원구분"] == "전체"]
        if bpu == "Total" or bpu in BPU_GROUPS:
            _official_brand_scope_df_traffic = _official_brand_scope_df_traffic.groupby("날짜", as_index=False)["트래픽"].sum()
        _official_brand_total_traffic = compute_official_total(_official_brand_scope_df_traffic, unit, selected_period_date, metric_col="트래픽")

        render_revenue_ranking(_brand_share_df_traffic, "브랜드", unit, selected_period_date, "브랜드별 트래픽 랭킹", _brand_subtitle,
                               label_map=BRAND_LABELS, hide_zero=True,
                               donut=True, official_total=_official_brand_total_traffic,
                               metric_col="트래픽", metric_label="트래픽",
                               bar_color_cur="#ea580c", bar_color_prev="#fdba74",
                               donut_colors=["#ea580c", "#f97316", "#fb923c", "#fdba74", "#fed7aa",
                                             "#c2410c", "#9a3412", "#7c2d12", "#fef3c7", "#fde68a", "#e2e8f0"],
                               ai_key="brand_rank_traffic", ai_context=f"브랜드별 트래픽 랭킹 · {_brand_subtitle} · {unit} · 기준 {period_label}" + (" · 핏플랍제외" if _ff_exclude else ""))

        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

        # --- 카테고리별 요약표 (조회 단위에 맞춰 전기간비/평균비/전년비) ---
        # 사이드바 조회단위(일별/주별/월별/월마감)를 그대로 따라간다 — KPI 카드와 같은
        # compute_kpi_deltas를 재사용해서 라벨(전일비/전주비/전월비 등)도 자동으로 맞춰짐.
        # 카테고리 선택 필터와 무관하게 전체 카테고리를 대상으로 함 (개요용 표라서).
        # _cat_summary_rows는 위(cat_bpu_df 확정 직후)에서 이미 계산해둔 걸 그대로 씀
        # (인사이트 카드 자동요약에서도 같은 값을 써서 숫자가 어긋나지 않게 하기 위함).
        st.markdown(f"**카테고리별 요약**  ·  <span style='color:#6b7280;font-size:0.85rem'>{bpu} · 거래액 기준 · {unit}</span>", unsafe_allow_html=True)

        if not _cat_summary_rows:
            st.info("거래액이 있는 카테고리가 없습니다.")
        else:
            if True:
                _cat_cfg = UNIT_CONFIG[unit]
                _cat_summary_body = "".join(
                    f"<tr><td class='m'>{r['카테고리']}</td>"
                    f"<td class='v' style='text-align:right;'>{r['거래액']:,.0f}</td>"
                    f"<td style='text-align:right;'>{format_delta_html(r['prev'])}{_ref_str(r['prev_v'])}</td>"
                    f"<td style='text-align:right;'>{format_delta_html(r['avg'])}{_ref_str(r['avg_v'])}</td>"
                    f"<td style='text-align:right;'>{format_delta_html(r['yoy'])}{_ref_str(r['yoy_v'])}</td></tr>"
                    for r in _cat_summary_rows
                )
                st.markdown(
                    "<div style='overflow-x:auto;'><table class='summary-table'><thead><tr>"
                    "<th>카테고리</th><th style='text-align:right;'>거래액</th>"
                    f"<th style='text-align:right;'>{_cat_cfg['prev_label']}</th>"
                    f"<th style='text-align:right;'>{_cat_cfg['avg_label']}</th>"
                    f"<th style='text-align:right;'>{_cat_cfg['yoy_label']}</th>"
                    "</tr></thead><tbody>" + _cat_summary_body + "</tbody></table></div>",
                    unsafe_allow_html=True,
                )
                st.caption(f"ℹ️ 기준: {period_label} · 거래액이 있는 카테고리만 표시돼요.")

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # ── 월별 실적 비교 (26년 | 전년비 | 25년) — 지금 선택된 매체/카테고리/브랜드/
        # 세그먼트 필터 그대로 반영. 3번(종합요약) 페이지와 같은 함수 재사용. ──
        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        st.markdown("---")
        st.markdown(
            f"#### 📆 월별 실적 비교 (26년 vs 25년 · {bpu} · {selected_cat}/{brand_label(selected_brand)} · {cat_segment})"
        )
        if cat_combo.empty:
            st.info("선택한 조합에 데이터가 없습니다.")
        else:
            _p2_base = cat_combo.groupby("날짜", as_index=False)[["거래액", "트래픽", "구매객수"]].sum()
            render_monthly_comparison_table(
                _p2_base, f"{bpu} · {selected_cat}/{brand_label(selected_brand)}",
                caption_extra="매체·카테고리·브랜드 필터를 바꾸면 이 표도 같이 바뀌어요.",
            )



# ============================================================
# 페이지 3: 누적 데이터 (EP실적 + EP채널 합쳐서 기간별 표)
# ============================================================
if side["page"].startswith("4."):
    st.markdown("---")
    st.markdown(
        f"<div class='chart-caption'>매체: <b>{bpu}</b> · 기간유형: <b>{cum_unit}</b> · 집계: <b>{cum_agg_mode}</b> · "
        f"고객구분: <b>{cum_segment}</b> · {cum_start} ~ {cum_end}"
        f"{' (EP채널 지표는 고객구분 구분 없이 전체 기준)' if cum_segment != '전체' else ''}</div>",
        unsafe_allow_html=True,
    )
    _agg_func = "sum" if cum_agg_mode == "누적" else "mean"

    # --- EP실적 부분 (트래픽/거래액/구매객수는 합계 가능, CR/객단가는 합계 기반 재계산) ---
    if bpu in BPU_GROUPS:
        _cum_tr = aggregate_traffic(df_traffic, BPU_GROUPS[bpu], cum_segment)
    else:
        _cum_tr = df_traffic[(df_traffic["BPU"] == bpu) & (df_traffic["회원구분"] == cum_segment)]

    tr_table_rows = {}
    if not _cum_tr.empty:
        base_series = {}
        for base_metric in ["트래픽", "거래액", "구매객수"]:
            s = _cum_tr.set_index("날짜")[base_metric].sort_index()
            series = s.resample(UNIT_CONFIG[cum_unit]["rule"]).agg(_agg_func)
            if cum_unit == "주별":
                series.index = series.index - pd.Timedelta(days=6)
            elif cum_unit == "월마감" and not series.empty and s.index.max() < series.index[-1]:
                series = series.iloc[:-1]
            series = _truncate_by_range(series, cum_start, cum_end, cum_unit)
            base_series[base_metric] = series
            tr_table_rows[base_metric] = series

        # 비율 지표(CR/객단가)는 일평균/누적 모드와 무관하게 항상 기간 합계 기반으로 재계산
        # (일별 비율값을 단순 평균하면 부정확 — 분자·분모를 각각 합산한 뒤 비율을 구해야 정확함)
        _sum_series = {}
        for base_metric in ["트래픽", "거래액", "구매객수"]:
            s = _cum_tr.set_index("날짜")[base_metric].sort_index()
            series_sum = s.resample(UNIT_CONFIG[cum_unit]["rule"]).sum()
            if cum_unit == "주별":
                series_sum.index = series_sum.index - pd.Timedelta(days=6)
            elif cum_unit == "월마감" and not series_sum.empty and s.index.max() < series_sum.index[-1]:
                series_sum = series_sum.iloc[:-1]
            series_sum = _truncate_by_range(series_sum, cum_start, cum_end, cum_unit)
            _sum_series[base_metric] = series_sum
        tr_table_rows["CR"] = (_sum_series["구매객수"] / _sum_series["트래픽"] * 100).replace([float("inf")], None)
        tr_table_rows["객단가"] = (_sum_series["거래액"] / _sum_series["구매객수"]).replace([float("inf")], None)

    # --- EP채널 부분 (전시상품수 등은 모드에 따라 평균/합계, 원부매칭율/최저가율은 항상 합계기반 재계산) ---
    if bpu in BPU_GROUPS:
        _cum_ep = aggregate_ep(df_ep, BPU_GROUPS[bpu], "Total", "Total")
    elif bpu == "Total":
        _cum_ep = df_ep[(df_ep[COL_BPU] == "Total") & (df_ep[COL_MATCH] == "Total") & (df_ep[COL_LOWEST] == "Total")]
    else:
        _cum_ep = df_ep[(df_ep[COL_BPU] == bpu) & (df_ep[COL_MATCH] == "Total") & (df_ep[COL_LOWEST] == "Total")]

    ep_table_rows = {}
    if not _cum_ep.empty:
        # 전시상품수/원부매칭상품수/최저가상품수는 집계 모드(일평균/누적)에 따라 평균 또는 합계
        for base_metric in ["평균 EP 전시 상품수", "평균 원부매칭 상품수", "평균 최저가 상품수"]:
            s = _cum_ep.set_index(COL_DATE)[base_metric].sort_index()
            series = s.resample(UNIT_CONFIG[cum_unit]["rule"]).agg(_agg_func)
            if cum_unit == "주별":
                series.index = series.index - pd.Timedelta(days=6)
            elif cum_unit == "월마감" and not series.empty and s.index.max() < series.index[-1]:
                series = series.iloc[:-1]
            series = _truncate_by_range(series, cum_start, cum_end, cum_unit)
            ep_table_rows[base_metric] = series

        # 원부매칭율/최저가율은 일평균/누적 모드와 무관하게 항상 기간 합계 기반으로 재계산
        ep_base_sum = {}
        for base_metric in ["평균 EP 전시 상품수", "평균 원부매칭 상품수", "평균 최저가 상품수"]:
            s = _cum_ep.set_index(COL_DATE)[base_metric].sort_index()
            series_sum = s.resample(UNIT_CONFIG[cum_unit]["rule"]).sum()
            if cum_unit == "주별":
                series_sum.index = series_sum.index - pd.Timedelta(days=6)
            elif cum_unit == "월마감" and not series_sum.empty and s.index.max() < series_sum.index[-1]:
                series_sum = series_sum.iloc[:-1]
            series_sum = _truncate_by_range(series_sum, cum_start, cum_end, cum_unit)
            ep_base_sum[base_metric] = series_sum
        ep_table_rows["원부매칭율(%)"] = (ep_base_sum["평균 원부매칭 상품수"] / ep_base_sum["평균 EP 전시 상품수"] * 100).replace([float("inf")], None)
        ep_table_rows["최저가율(%)"] = (ep_base_sum["평균 최저가 상품수"] / ep_base_sum["평균 EP 전시 상품수"] * 100).replace([float("inf")], None)

    # --- 병합해서 표 만들기 (정렬 순서는 상단 필터로 선택) ---
    all_dates = sorted(set().union(
        *[s.index for s in tr_table_rows.values()],
        *[s.index for s in ep_table_rows.values()],
    ), reverse=(cum_sort_order == "최신순"))

    if not all_dates:
        st.info("선택한 조건에 데이터가 없습니다.")
    else:
        COLS = [
            ("트래픽", "UV", False), ("거래액", "거래액(순결제)", False), ("구매객수", "구매객수", False),
            ("CR", "구매전환율(%)", True), ("객단가", "객단가", False),
            ("원부매칭율(%)", "원부매칭율(%)", True), ("최저가율(%)", "최저가율(%)", True),
            ("평균 EP 전시 상품수", "전시상품수", False), ("평균 원부매칭 상품수", "원부매칭상품수", False),
            ("평균 최저가 상품수", "최저가상품수", False),
        ]
        header_html = "<th>구분</th>" + "".join(f"<th>{label}</th>" for _, label, _ in COLS)
        body_rows = []
        export_rows = []
        for d in all_dates:
            row_label = make_period_label(d, cum_unit)
            cells = []
            export_row = {"구분": row_label}
            for key, label, is_pct in COLS:
                src = tr_table_rows.get(key, ep_table_rows.get(key))
                val = src.get(d) if src is not None and d in src.index else None
                if val is None or pd.isna(val):
                    cells.append("<td>-</td>")
                    export_row[label] = None
                elif is_pct:
                    cells.append(f"<td>{val:.1f}%</td>")
                    export_row[label] = float(val)  # 엑셀에는 반올림 없이 원본 값 그대로
                else:
                    cells.append(f"<td>{val:,.0f}</td>")
                    export_row[label] = float(val)  # 엑셀에는 반올림 없이 원본 값 그대로
            body_rows.append(f"<tr><td class='m'>{row_label}</td>{''.join(cells)}</tr>")
            export_rows.append(export_row)

        # --- 합계 행: 절대값 지표는 선택 기간 전체 합산, 비율 지표(CR/객단가/원부매칭율/
        # 최저가율)는 분자·분모를 각각 합산한 뒤 재계산한다(단순 평균 금지 원칙 그대로).
        # (엑셀 다운로드에도 포함되도록, export_rows 조립보다 먼저 계산해서 앞에 끼워넣는다.) ---
        _tot_cur_range = (
            _cum_tr[(_cum_tr["날짜"] >= pd.Timestamp(cum_start)) & (_cum_tr["날짜"] <= pd.Timestamp(cum_end))]
            if not _cum_tr.empty else pd.DataFrame()
        )
        _tot_ep_range = (
            _cum_ep[(_cum_ep[COL_DATE] >= pd.Timestamp(cum_start)) & (_cum_ep[COL_DATE] <= pd.Timestamp(cum_end))]
            if not _cum_ep.empty else pd.DataFrame()
        )
        _tot_vals = {}
        for _m in ["트래픽", "거래액", "구매객수"]:
            _tot_vals[_m] = _tot_cur_range[_m].sum() if not _tot_cur_range.empty else None
        _tot_vals["CR"] = (
            (_tot_vals["구매객수"] / _tot_vals["트래픽"] * 100)
            if _tot_vals["트래픽"] else None
        )
        _tot_vals["객단가"] = (
            (_tot_vals["거래액"] / _tot_vals["구매객수"])
            if _tot_vals["구매객수"] else None
        )
        for _m in ["평균 EP 전시 상품수", "평균 원부매칭 상품수", "평균 최저가 상품수"]:
            _tot_vals[_m] = _tot_ep_range[_m].sum() if not _tot_ep_range.empty else None
        _tot_vals["원부매칭율(%)"] = (
            (_tot_vals["평균 원부매칭 상품수"] / _tot_vals["평균 EP 전시 상품수"] * 100)
            if _tot_vals["평균 EP 전시 상품수"] else None
        )
        _tot_vals["최저가율(%)"] = (
            (_tot_vals["평균 최저가 상품수"] / _tot_vals["평균 EP 전시 상품수"] * 100)
            if _tot_vals["평균 EP 전시 상품수"] else None
        )
        _tot_cells = []
        _tot_export_row = {"구분": "합계"}
        for key, label, is_pct in COLS:
            v = _tot_vals.get(key)
            if v is None or pd.isna(v):
                _tot_cells.append("<td>-</td>")
                _tot_export_row[label] = None
            elif is_pct:
                _tot_cells.append(f"<td>{v:.1f}%</td>")
                _tot_export_row[label] = float(v)
            else:
                _tot_cells.append(f"<td>{v:,.0f}</td>")
                _tot_export_row[label] = float(v)
        _total_row_html = f"<tr style='background:#f3f4f6;font-weight:700;'><td class='m'>합계</td>{''.join(_tot_cells)}</tr>"
        export_rows.insert(0, _tot_export_row)

        _tc1, _tc2 = st.columns([4, 1])
        with _tc1:
            st.markdown(f"**누적 데이터**  ·  <span style='color:#6b7280;font-size:0.85rem'>{len(all_dates)}개 기간</span>", unsafe_allow_html=True)
        with _tc2:
            _fname = f"누적데이터_{bpu}_{cum_unit}_{cum_start}_{cum_end}.xlsx".replace(" ", "")
            render_excel_download(pd.DataFrame(export_rows), _fname)

        table_html = (
            "<div style='overflow-x:auto;'><table class='summary-table'>"
            f"<thead><tr>{header_html}</tr></thead>"
            f"<tbody>{_total_row_html}{''.join(body_rows)}</tbody></table></div>"
        )
        st.markdown(table_html, unsafe_allow_html=True)

        # 규칙 기반 자동 인사이트 — 이 페이지는 전기간비/전년비 개념이 없는 단순 누적표라
        # (표 자체에 이미 '합계' 행이 있음) 짧게 한 줄만: 선택 기간의 핵심 합계값 요약.
        if _tot_vals.get("거래액") is not None:
            _cum_body = (
                f"{cum_start} ~ {cum_end} 누적 — 트래픽 {_tot_vals.get('트래픽', 0):,.0f} · "
                f"거래액 {_tot_vals['거래액']:,.0f} · 구매객수 {_tot_vals.get('구매객수', 0):,.0f}"
            )
            if _tot_vals.get("CR") is not None:
                _cum_body += f" · CR {_tot_vals['CR']:.1f}%"
            if _tot_vals.get("객단가") is not None:
                _cum_body += f" · 객단가 {_tot_vals['객단가']:,.0f}"
            render_insight_panel([{"title": "① 이 기간 누적 실적은?", "body": _cum_body}])


# ============================================================
# 페이지 4: 누적 데이터 (카테고리)
# ============================================================
if side["page"].startswith("5."):
    st.markdown("---")
    st.markdown(
        f"<div class='chart-caption'>매체: <b>{bpu}</b> · 기간유형: <b>{cum_unit}</b> · 집계: <b>{cum_agg_mode}</b> · "
        f"고객구분: <b>{cum_segment}</b> · {cum_start} ~ {cum_end} · <b>{selected_cat}</b> / <b>{brand_label(selected_brand)}</b></div>",
        unsafe_allow_html=True,
    )
    _agg_func = "sum" if cum_agg_mode == "누적" else "mean"

    if df_category.empty:
        st.info("카테고리 데이터가 없습니다.")
    else:
        if bpu in BPU_GROUPS:
            _cum_cat_df = df_category[df_category["BPU"].isin(BPU_GROUPS[bpu])]
        elif bpu == "Total":
            _cum_cat_df = df_category
        else:
            _cum_cat_df = df_category[df_category["BPU"] == bpu]

        if "회원구분" in _cum_cat_df.columns:
            _cum_cat_df = _cum_cat_df[_cum_cat_df["회원구분"] == cum_segment]

        _cum_cat_combo = _cum_cat_df[(_cum_cat_df["카테고리"] == selected_cat) & (_cum_cat_df["브랜드"] == selected_brand)]
        if (bpu == "Total" or bpu in BPU_GROUPS) and not _cum_cat_combo.empty:
            _cum_cat_combo = _cum_cat_combo.groupby("날짜", as_index=False).agg({"트래픽": "sum", "거래액": "sum", "구매객수": "sum"})
            _cum_cat_combo["CR"] = (_cum_cat_combo["구매객수"] / _cum_cat_combo["트래픽"] * 100).where(_cum_cat_combo["트래픽"] > 0, 0)
            _cum_cat_combo["객단가"] = (_cum_cat_combo["거래액"] / _cum_cat_combo["구매객수"]).where(_cum_cat_combo["구매객수"] > 0, 0)

        if _cum_cat_combo.empty:
            st.warning(f"{selected_cat} / {brand_label(selected_brand)} 조합에 데이터가 없습니다.")
        else:
            CAT_CUM_COLS = [
                ("트래픽", "UV", False), ("거래액", "거래액", False), ("구매객수", "구매객수", False),
                ("CR", "구매전환율(%)", True), ("객단가", "객단가", False),
            ]
            cat_table_rows = {}
            # UV/거래액/구매객수는 집계 모드(일평균/누적)에 따라 평균 또는 합계
            for base_metric in ["트래픽", "거래액", "구매객수"]:
                s = _cum_cat_combo.set_index("날짜")[base_metric].sort_index()
                series = s.resample(UNIT_CONFIG[cum_unit]["rule"]).agg(_agg_func)
                if cum_unit == "주별":
                    series.index = series.index - pd.Timedelta(days=6)
                elif cum_unit == "월마감" and not series.empty and s.index.max() < series.index[-1]:
                    series = series.iloc[:-1]
                series = _truncate_by_range(series, cum_start, cum_end, cum_unit)
                cat_table_rows[base_metric] = series

            # CR/객단가는 일평균/누적 모드와 무관하게 항상 기간 합계 기반으로 재계산
            cat_base_sum = {}
            for base_metric in ["트래픽", "거래액", "구매객수"]:
                s = _cum_cat_combo.set_index("날짜")[base_metric].sort_index()
                series_sum = s.resample(UNIT_CONFIG[cum_unit]["rule"]).sum()
                if cum_unit == "주별":
                    series_sum.index = series_sum.index - pd.Timedelta(days=6)
                elif cum_unit == "월마감" and not series_sum.empty and s.index.max() < series_sum.index[-1]:
                    series_sum = series_sum.iloc[:-1]
                series_sum = _truncate_by_range(series_sum, cum_start, cum_end, cum_unit)
                cat_base_sum[base_metric] = series_sum
            cat_table_rows["CR"] = (cat_base_sum["구매객수"] / cat_base_sum["트래픽"] * 100).replace([float("inf")], None)
            cat_table_rows["객단가"] = (cat_base_sum["거래액"] / cat_base_sum["구매객수"]).replace([float("inf")], None)

            all_cat_dates = sorted(set().union(*[s.index for s in cat_table_rows.values()]), reverse=(cum_sort_order == "최신순"))
            if not all_cat_dates:
                st.info("선택한 조건에 데이터가 없습니다.")
            else:
                header_html2 = "<th>구분</th>" + "".join(f"<th>{label}</th>" for _, label, _ in CAT_CUM_COLS)
                body_rows2 = []
                export_rows2 = []
                for d in all_cat_dates:
                    row_label = make_period_label(d, cum_unit)
                    cells = []
                    export_row2 = {"구분": row_label}
                    for key, label, _is_pct in CAT_CUM_COLS:
                        val = cat_table_rows[key].get(d)
                        if val is None or pd.isna(val):
                            cells.append("<td>-</td>")
                            export_row2[label] = None
                        elif _is_pct:
                            cells.append(f"<td>{val:.1f}%</td>")
                            export_row2[label] = float(val)  # 엑셀에는 반올림 없이 원본 값 그대로
                        else:
                            cells.append(f"<td>{val:,.0f}</td>")
                            export_row2[label] = float(val)  # 엑셀에는 반올림 없이 원본 값 그대로
                    body_rows2.append(f"<tr><td class='m'>{row_label}</td>{''.join(cells)}</tr>")
                    export_rows2.append(export_row2)

                _tc3, _tc4 = st.columns([4, 1])
                with _tc3:
                    st.markdown(f"**카테고리 누적 데이터**  ·  <span style='color:#6b7280;font-size:0.85rem'>{len(all_cat_dates)}개 기간</span>", unsafe_allow_html=True)
                with _tc4:
                    _fname2 = f"카테고리누적데이터_{bpu}_{selected_cat}_{brand_label(selected_brand)}_{cum_unit}.xlsx".replace(" ", "")
                    render_excel_download(pd.DataFrame(export_rows2), _fname2)
                table_html2 = (
                    "<div style='overflow-x:auto;'><table class='summary-table'>"
                    f"<thead><tr>{header_html2}</tr></thead>"
                    f"<tbody>{''.join(body_rows2)}</tbody></table></div>"
                )
                st.markdown(table_html2, unsafe_allow_html=True)

                # 규칙 기반 자동 인사이트 — 4번 페이지와 동일하게 짧은 누적 합계 요약.
                try:
                    _cat_cum_gmv = cat_base_sum.get("거래액")
                    _cat_cum_uv = cat_base_sum.get("트래픽")
                    if _cat_cum_gmv is not None and not _cat_cum_gmv.empty:
                        _cat_body = f"{selected_cat}/{brand_label(selected_brand)} 누적 — 거래액 {_cat_cum_gmv.sum():,.0f}"
                        if _cat_cum_uv is not None and not _cat_cum_uv.empty:
                            _cat_body += f" · 트래픽 {_cat_cum_uv.sum():,.0f}"
                        render_insight_panel([{"title": "① 이 조합의 누적 실적은?", "body": _cat_body}])
                except Exception:
                    pass


# ============================================================
# 페이지 5: 쿠폰 비용 분석
# ============================================================
if side["page"].startswith("9."):
    st.markdown("---")

    _has_daily = not df_coupon_daily.empty

    if df_coupon.empty:
        st.info("쿠폰 데이터가 없습니다. 사이드바에서 ep_coupon_daily.csv를 업로드해주세요.")
    else:
        # 매체(coupon_bpu)·쿠폰유형(coupon_type_sel)·조회단위(coupon_unit)·기준시점(coupon_ref_ts)은
        # 모두 상단 고정 필터(사이드바 바로 아래)에서 이미 정해져서 넘어온다.

        def _bpu_gmv_source(bpu_val):
            if bpu_val in BPU_GROUPS:
                return df_traffic[(df_traffic["BPU"].isin(BPU_GROUPS[bpu_val])) & (df_traffic["회원구분"] == "전체")]
            elif bpu_val == "Total":
                return df_traffic[(df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")]
            else:
                return df_traffic[(df_traffic["BPU"] == bpu_val) & (df_traffic["회원구분"] == "전체")]

        # ============================================================
        # 전체 시계열(_combined_full) 생성 — 아래 상단 기간필터 적용 전 원본.
        # 전기간비/전년비 조회는 항상 이 전체 시계열 기준으로 해서, 선택 구간을
        # 좁혀도 비교값을 못 찾는 일이 없게 한다.
        # ============================================================
        if coupon_unit == "월별":
            _cp_sub = df_coupon[df_coupon["BPU"] == coupon_bpu].copy()
            if coupon_type_sel != "합산":
                _cp_sub = _cp_sub[_cp_sub["쿠폰유형"] == coupon_type_sel]

            _coupon_by_month = _cp_sub.groupby("연월", as_index=False)["쿠폰할인"].sum().set_index("연월")["쿠폰할인"].sort_index()
            _gmv_by_month = _bpu_gmv_source(coupon_bpu).set_index("날짜")["거래액"].resample("MS").sum()
            _coupon_by_month.index = _coupon_by_month.index.to_period("M").to_timestamp()

            _combined_full = pd.DataFrame({"쿠폰할인": _coupon_by_month, "거래액": _gmv_by_month}).dropna(how="all")
            _combined_full["거래액"] = _combined_full["거래액"].fillna(0)
            _combined_full["쿠폰할인"] = _combined_full["쿠폰할인"].fillna(0)
            _combined_full["비용률"] = (_combined_full["쿠폰할인"] / _combined_full["거래액"] * 100).where(_combined_full["거래액"] > 0)
            _combined_full = _combined_full.sort_index()
            _period_fmt = lambda d: d.strftime("%Y년 %m월")
            _prev_label, _has_yoy = "전월비", True
            _step = pd.DateOffset(months=1)

        # ============================================================
        # 일별/주별 조회: 일자별 상세(ep_coupon_daily.csv) 사용
        # 참고: 이제 원본에 25년치도 있지만, 일/주별 전년비교 로직은 아직 미구현 상태(항상 _has_yoy=False)
        # ============================================================
        else:
            if coupon_bpu in BPU_GROUPS:
                _cpd_sub = df_coupon_daily[df_coupon_daily["BPU"].isin(BPU_GROUPS[coupon_bpu])].copy()
            elif coupon_bpu == "Total":
                _cpd_sub = df_coupon_daily.copy()
            else:
                _cpd_sub = df_coupon_daily[df_coupon_daily["BPU"] == coupon_bpu].copy()
            if coupon_type_sel != "합산":
                _cpd_sub = _cpd_sub[_cpd_sub["쿠폰유형"] == coupon_type_sel]

            _rule = "D" if coupon_unit == "일별" else "W-SUN"
            _coupon_series = _cpd_sub.groupby("날짜")["쿠폰할인"].sum().resample(_rule).sum()
            _gmv_series = _bpu_gmv_source(coupon_bpu).set_index("날짜")["거래액"].resample(_rule).sum()
            if coupon_unit == "주별":
                _coupon_series.index = _coupon_series.index - pd.Timedelta(days=6)
                _gmv_series.index = _gmv_series.index - pd.Timedelta(days=6)

            _combined_full = pd.DataFrame({"쿠폰할인": _coupon_series, "거래액": _gmv_series}).dropna(how="all")
            _combined_full["거래액"] = _combined_full["거래액"].fillna(0)
            _combined_full["쿠폰할인"] = _combined_full["쿠폰할인"].fillna(0)
            _combined_full["비용률"] = (_combined_full["쿠폰할인"] / _combined_full["거래액"] * 100).where(_combined_full["거래액"] > 0)
            _combined_full = _combined_full.sort_index()
            _prev_label_txt = "전일비" if coupon_unit == "일별" else "전주비"
            _period_fmt = (lambda d: d.strftime("%Y-%m-%d")) if coupon_unit == "일별" else (lambda d: f"{d.strftime('%Y-%m-%d')} 주")
            _prev_label, _has_yoy = _prev_label_txt, False
            _step = {"일별": pd.DateOffset(days=1), "주별": pd.DateOffset(weeks=1)}[coupon_unit]

        if _combined_full.empty:
            st.warning("해당 조건에 데이터가 없습니다.")
        else:
            # ============================================================
            # 상단 고정 필터의 기준일자/기준시점(coupon_ref_ts)까지만 사용.
            # KPI 카드·쿠폰명 랭킹·전년비 비교표가 모두 "그 시점 기준"으로 표시됨.
            # 전기간비/전년비는 그 이전 데이터를 찾아야 하므로 _combined_full을 그대로 쓴다.
            # ============================================================
            _latest = coupon_ref_ts
            if _latest not in _combined_full.index:
                _cp_cand = _combined_full.index[_combined_full.index <= _latest]
                _latest = _cp_cand[-1] if len(_cp_cand) else _combined_full.index[-1]
            _combined = _combined_full[_combined_full.index <= _latest]

            if _combined.empty:
                st.warning("선택한 기준일자에 데이터가 없습니다.")
            else:
                _prev_period = _latest - _step
                _prev_year = _latest - pd.DateOffset(years=1)

                _cur_coupon = _combined.loc[_latest, "쿠폰할인"]
                _cur_gmv = _combined.loc[_latest, "거래액"]
                _cur_rate = _combined.loc[_latest, "비용률"]

                def _delta_pct(cur, ref):
                    return pct_delta_safe(cur, ref)

                # 전기간비/전년비는 선택 구간 밖 값도 찾을 수 있도록 _combined_full 기준으로 조회
                _prev_coupon = _combined_full.loc[_prev_period, "쿠폰할인"] if _prev_period in _combined_full.index else None
                _prev_gmv = _combined_full.loc[_prev_period, "거래액"] if _prev_period in _combined_full.index else None
                _prev_rate = _combined_full.loc[_prev_period, "비용률"] if _prev_period in _combined_full.index else None
                if _has_yoy:
                    _yoy_coupon = _combined_full.loc[_prev_year, "쿠폰할인"] if _prev_year in _combined_full.index else None
                    _yoy_gmv = _combined_full.loc[_prev_year, "거래액"] if _prev_year in _combined_full.index else None
                    _yoy_rate = _combined_full.loc[_prev_year, "비용률"] if _prev_year in _combined_full.index else None
                else:
                    _yoy_coupon = _yoy_gmv = _yoy_rate = None

                _yoy_note = "" if _has_yoy else " · <span style='color:#9ca3af'>일별/주별 전년비교는 아직 지원 예정 기능이에요</span>"
                st.markdown(
                    f"<div class='chart-caption'>{coupon_bpu} · {coupon_type_sel} · {coupon_unit} · 기준: {_period_fmt(_latest)}{_yoy_note}</div>",
                    unsafe_allow_html=True,
                )
                kc1, kc2, kc3 = st.columns(3)
                with kc1:
                    _yoy_line = (
                        f"전년동기비 {format_delta_html(_delta_pct(_cur_coupon, _yoy_coupon))}{_ref_str(_yoy_coupon)}<br/>"
                        if _has_yoy else ""
                    )
                    st.markdown(
                        f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;min-height:150px;'>"
                        f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>쿠폰할인</div>"
                        f"<div style='font-size:1.4rem;font-weight:700;color:#111827;'>{_cur_coupon:,.0f}</div>"
                        f"<div style='font-size:0.76rem;margin-top:6px;'>"
                        f"{_prev_label} {format_delta_html(_delta_pct(_cur_coupon, _prev_coupon))}{_ref_str(_prev_coupon)}<br/>"
                        f"{_yoy_line}"
                        f"</div></div>", unsafe_allow_html=True,
                    )
                with kc2:
                    _yoy_line2 = (
                        f"전년동기비 {format_delta_html(_delta_pct(_cur_gmv, _yoy_gmv))}{_ref_str(_yoy_gmv)}<br/>"
                        if _has_yoy else ""
                    )
                    st.markdown(
                        f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;min-height:150px;'>"
                        f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>거래액(기존 대시보드 기준)</div>"
                        f"<div style='font-size:1.4rem;font-weight:700;color:#111827;'>{_cur_gmv:,.0f}</div>"
                        f"<div style='font-size:0.76rem;margin-top:6px;'>"
                        f"{_prev_label} {format_delta_html(_delta_pct(_cur_gmv, _prev_gmv))}{_ref_str(_prev_gmv)}<br/>"
                        f"{_yoy_line2}"
                        f"</div></div>", unsafe_allow_html=True,
                    )
                with kc3:
                    _rate_prev_delta = (_cur_rate - _prev_rate) if (_prev_rate is not None and pd.notna(_prev_rate) and pd.notna(_cur_rate)) else None
                    _rate_yoy_delta = (_cur_rate - _yoy_rate) if (_has_yoy and _yoy_rate is not None and pd.notna(_yoy_rate) and pd.notna(_cur_rate)) else None
                    _yoy_line3 = (
                        f"전년동기비 {format_delta_html(_rate_yoy_delta) if _rate_yoy_delta is not None else '-'}%p{_ref_str(_yoy_rate, True)}<br/>"
                        if _has_yoy else ""
                    )
                    st.markdown(
                        f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;min-height:150px;'>"
                        f"<div style='color:#6b7280;font-size:0.8rem;margin-bottom:4px;'>비용률 (쿠폰할인/거래액)</div>"
                        f"<div style='font-size:1.4rem;font-weight:700;color:#111827;'>{_cur_rate:.2f}%</div>"
                        f"<div style='font-size:0.76rem;margin-top:6px;'>"
                        f"{_prev_label} {format_delta_html(_rate_prev_delta) if _rate_prev_delta is not None else '-'}%p{_ref_str(_prev_rate, True)}<br/>"
                        f"{_yoy_line3}"
                        f"</div></div>", unsafe_allow_html=True,
                    )

                st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

                # 쿠폰 유형별 비용률 (합산 선택시)
                if coupon_type_sel == "합산":
                    st.markdown("**쿠폰 유형별 비용률 (같은 매체·기간 기준)**")
                    _type_rows = []
                    for _t in ["플러스", "일반"]:
                        if coupon_unit == "월별":
                            _t_src = df_coupon[(df_coupon["BPU"] == coupon_bpu) & (df_coupon["쿠폰유형"] == _t)]
                            if _t_src.empty:
                                continue
                            _t_by_period = _t_src.groupby("연월")["쿠폰할인"].sum()
                            _t_by_period.index = _t_by_period.index.to_period("M").to_timestamp()
                        else:
                            if coupon_bpu in BPU_GROUPS:
                                _t_src = df_coupon_daily[(df_coupon_daily["BPU"].isin(BPU_GROUPS[coupon_bpu])) & (df_coupon_daily["쿠폰유형"] == _t)]
                            elif coupon_bpu == "Total":
                                _t_src = df_coupon_daily[df_coupon_daily["쿠폰유형"] == _t]
                            else:
                                _t_src = df_coupon_daily[(df_coupon_daily["BPU"] == coupon_bpu) & (df_coupon_daily["쿠폰유형"] == _t)]
                            if _t_src.empty:
                                continue
                            _t_by_period = _t_src.groupby("날짜")["쿠폰할인"].sum().resample(_rule).sum()
                            if coupon_unit == "주별":
                                _t_by_period.index = _t_by_period.index - pd.Timedelta(days=6)
                        _t_cur = _t_by_period.get(_latest, 0)
                        _t_rate = (_t_cur / _cur_gmv * 100) if _cur_gmv > 0 else None
                        _type_rows.append({"쿠폰유형": _t, "쿠폰할인": _t_cur, "비용률": _t_rate})
                    _type_df = pd.DataFrame(_type_rows)
                    _tcols = st.columns(len(_type_df)) if len(_type_df) else []
                    for _i, _r in enumerate(_type_df.itertuples()):
                        with _tcols[_i]:
                            _rate_html = f"비용률 {_r.비용률:.2f}%" if _r.비용률 is not None else "비용률 -"
                            st.markdown(
                                f"<div style='background:#f8fafc;border:1px solid #e5e7eb;border-radius:8px;padding:10px 14px;'>"
                                f"<div style='font-size:0.78rem;color:#6b7280;'>{_r.쿠폰유형}</div>"
                                f"<div style='font-size:1.1rem;font-weight:700;'>{_r.쿠폰할인:,.0f}</div>"
                                f"<div style='font-size:0.85rem;color:#7c3aed;'>{_rate_html}</div>"
                                f"</div>", unsafe_allow_html=True,
                            )
                    st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)

                # ============================================================
                # 추이 차트: 상단 기간필터가 적용된 _combined을 기본으로 사용.
                # 일별 조회일 땐 점이 너무 많아지므로, 그래프 전용 기간(줌) 컨트롤을
                # 추가로 둔다 (기본 최근 30일 + "최근으로" 리셋 버튼).
                # ============================================================
                st.markdown(f"**{coupon_unit} 쿠폰할인 · 비용률 추이**")

                _chart_series = _combined
                if coupon_unit == "일별" and len(_combined) > 1:
                    _cp_chart_default_start = max(
                        _combined.index.min().date(), _combined.index.max().date() - _dt.timedelta(days=30)
                    )
                    _cp_chart_key = f"coupon_chart_range_{coupon_bpu}_{coupon_type_sel}"
                    _cc1, _cc2 = st.columns([2, 1])
                    with _cc1:
                        st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>그래프 기간</div>", unsafe_allow_html=True)
                        _cp_chart_dr = st.date_input(
                            "그래프 기간",
                            value=(_cp_chart_default_start, _combined.index.max().date()),
                            min_value=_combined.index.min().date(), max_value=_combined.index.max().date(),
                            label_visibility="collapsed", key=_cp_chart_key,
                        )
                    with _cc2:
                        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
                        st.button(
                            "🔄 최근으로", key=f"{_cp_chart_key}_reset", use_container_width=True,
                            on_click=_reset_date_range, args=(_cp_chart_key, (_cp_chart_default_start, _combined.index.max().date())),
                        )
                    if isinstance(_cp_chart_dr, tuple) and len(_cp_chart_dr) == 2:
                        _chart_series = _combined[
                            (_combined.index >= pd.Timestamp(_cp_chart_dr[0])) & (_combined.index <= pd.Timestamp(_cp_chart_dr[1]))
                        ]

                import altair as alt
                _trend_df = _chart_series.copy()
                _trend_df.index.name = "연월"
                _trend_df = _trend_df.reset_index()
                _trend_df["연월_label"] = _trend_df["연월"].apply(_period_fmt)
                _month_order = _trend_df["연월_label"].tolist()
                _base = alt.Chart(_trend_df).encode(
                    x=alt.X("연월_label:O", title=None, sort=_month_order, axis=alt.Axis(labelAngle=-40))
                )
                # 막대를 클릭하면 그 날짜(구간)의 쿠폰명별 랭킹이 아래 표시되도록, 클릭
                # 선택(click selection)을 막대에 건다. 선택된 막대는 진하게, 나머지는
                # 흐리게 표시해서 "지금 뭘 클릭했는지" 눈에 보이게 한다.
                _click_sel = alt.selection_point(fields=["연월_label"], on="click", empty=True, name="clicked")
                _bar = _base.mark_bar(color="#93c5fd", size=18).encode(
                    y=alt.Y("쿠폰할인:Q", title="쿠폰할인", axis=alt.Axis(format="~s")),
                    opacity=alt.condition(_click_sel, alt.value(1.0), alt.value(0.55)),
                    tooltip=[alt.Tooltip("연월_label:N", title="기간"), alt.Tooltip("쿠폰할인:Q", title="쿠폰할인", format=",.0f")],
                ).add_params(_click_sel)
                _line = _base.mark_line(color="#dc2626", strokeWidth=2, point=alt.OverlayMarkDef(size=45, filled=True)).encode(
                    y=alt.Y("비용률:Q", title="비용률(%)", axis=alt.Axis(format=".1f")),
                    tooltip=[alt.Tooltip("연월_label:N", title="기간"), alt.Tooltip("비용률:Q", title="비용률", format=".2f")],
                )
                _chart = alt.layer(_bar, _line).resolve_scale(y="independent").properties(height=350)
                _chart_click_key = f"coupon_chart_click_{coupon_bpu}_{coupon_type_sel}_{coupon_unit}"
                _cp_event = st.altair_chart(
                    _chart, use_container_width=True, on_select="rerun", key=_chart_click_key,
                )
                st.caption("💡 막대를 클릭하면 그 날짜(구간)의 쿠폰명별 할인액 랭킹을 아래에서 볼 수 있어요.")

                # 클릭된 막대가 있으면 그 날짜를, 없으면 기존처럼 기준시점(_latest)을 랭킹 기준으로 쓴다.
                _rank_date = _latest
                _rank_date_is_clicked = False
                try:
                    _clicked_pts = _cp_event.selection.get("clicked", []) if _cp_event else []
                except AttributeError:
                    _clicked_pts = (_cp_event or {}).get("selection", {}).get("clicked", [])
                if _clicked_pts:
                    _clicked_label = _clicked_pts[0].get("연월_label")
                    _label_to_period = dict(zip(_trend_df["연월_label"], _trend_df["연월"]))
                    if _clicked_label in _label_to_period:
                        _rank_date = _label_to_period[_clicked_label]
                        _rank_date_is_clicked = True

                _export_df = _trend_df.copy()
                _export_df["연월"] = _export_df["연월_label"]
                _export_df = _export_df.drop(columns=["연월_label"])
                render_excel_download(_export_df, f"쿠폰비용_{coupon_bpu}_{coupon_type_sel}_{coupon_unit}.xlsx")

                st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

                # 쿠폰명별 랭킹 (클릭한 날짜가 있으면 그 날짜, 없으면 기준시점 기준)
                _rank_title_suffix = " (클릭한 날짜)" if _rank_date_is_clicked else ""
                st.markdown(f"**쿠폰명별 할인액 랭킹 · {_period_fmt(_rank_date)} 기준{_rank_title_suffix}**")
                if coupon_unit == "월별":
                    if df_coupon_detail.empty:
                        st.info("쿠폰 상세 데이터가 없습니다. 사이드바에서 ep_coupon_daily.csv를 업로드해주세요.")
                        _detail_sub = None
                    else:
                        _detail_sub = df_coupon_detail[df_coupon_detail["연월"] == _rank_date]
                else:
                    if not _has_daily:
                        _detail_sub = None
                    else:
                        if coupon_unit == "일별":
                            _detail_sub = df_coupon_daily[df_coupon_daily["날짜"] == _rank_date]
                        else:  # 주별: 해당 주(월~일) 범위 합산
                            _week_start, _week_end = _rank_date, _rank_date + pd.Timedelta(days=6)
                            _detail_sub = df_coupon_daily[(df_coupon_daily["날짜"] >= _week_start) & (df_coupon_daily["날짜"] <= _week_end)]

                if _detail_sub is None:
                    pass
                else:
                    if coupon_bpu in BPU_GROUPS:
                        _detail_sub = _detail_sub[_detail_sub["BPU"].isin(BPU_GROUPS[coupon_bpu])]
                    elif coupon_bpu != "Total":
                        _detail_sub = _detail_sub[_detail_sub["BPU"] == coupon_bpu]
                    if coupon_type_sel != "합산":
                        _detail_sub = _detail_sub[_detail_sub["쿠폰유형"] == coupon_type_sel]

                    if _detail_sub.empty:
                        st.info("해당 조건의 쿠폰 상세 데이터가 없습니다.")
                    else:
                        _rank = _detail_sub.groupby(["쿠폰ID", "쿠폰명", "쿠폰유형"], as_index=False)["쿠폰할인"].sum()
                        _rank = _rank.sort_values("쿠폰할인", ascending=False).head(15).reset_index(drop=True)
                        _rank_html = "".join(
                            f"<tr><td>{i+1}</td><td>{r['쿠폰ID']}</td><td>{r['쿠폰명']}</td><td>{r['쿠폰유형']}</td>"
                            f"<td style='text-align:right;'>{r['쿠폰할인']:,.0f}</td></tr>"
                            for i, r in _rank.iterrows()
                        )
                        st.markdown(
                            "<table class='summary-table'><thead><tr>"
                            "<th>#</th><th>쿠폰번호</th><th>쿠폰명</th><th>유형</th><th style='text-align:right;'>할인액</th>"
                            "</tr></thead><tbody>" + _rank_html + "</tbody></table>",
                            unsafe_allow_html=True,
                        )

                st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

                # 전년비 비교표는 월별 조회에서만 (일/주별 전년비교는 아직 미구현)
                # 선택한 기간(_combined) 범위만 표에 나오되, 전년 값 조회는 _combined_full에서 한다.
                if _has_yoy:
                    st.markdown("**월별 전년비 비교**")
                    _yoy_rows = []
                    for _m in sorted(_combined.index, reverse=True):
                        _py = _m - pd.DateOffset(years=1)
                        _cur_gmv_m = _combined.loc[_m, "거래액"]
                        _cur_coupon_m = _combined.loc[_m, "쿠폰할인"]
                        _cur_rate_m = _combined.loc[_m, "비용률"]
                        if _py in _combined_full.index:
                            _py_gmv = _combined_full.loc[_py, "거래액"]
                            _py_coupon = _combined_full.loc[_py, "쿠폰할인"]
                            _py_rate = _combined_full.loc[_py, "비용률"]
                        else:
                            _py_gmv = _py_coupon = _py_rate = None
                        _gmv_yoy = pct_delta_safe(_cur_gmv_m, _py_gmv) if _py_gmv else None
                        _coupon_yoy = pct_delta_safe(_cur_coupon_m, _py_coupon) if _py_coupon else None
                        _rate_yoy_pt = (_cur_rate_m - _py_rate) if (_py_rate is not None and pd.notna(_py_rate) and pd.notna(_cur_rate_m)) else None
                        _yoy_rows.append({
                            "연월": _m, "거래액": _cur_gmv_m, "작년거래액": _py_gmv, "거래액증감": _gmv_yoy,
                            "쿠폰할인": _cur_coupon_m, "작년쿠폰할인": _py_coupon, "쿠폰할인증감": _coupon_yoy,
                            "비용률": _cur_rate_m, "작년비용률": _py_rate, "비용률증감": _rate_yoy_pt,
                        })

                    def _fmt_num(v):
                        return f"{v:,.0f}" if v is not None and pd.notna(v) else "-"

                    def _fmt_pct(v):
                        return f"{v:.2f}%" if v is not None and pd.notna(v) else "-"

                    _yoy_body = ""
                    for r in _yoy_rows:
                        _yoy_body += (
                            f"<tr><td class='m'>{r['연월'].strftime('%Y년 %m월')}</td>"
                            f"<td style='text-align:right;'>{_fmt_num(r['거래액'])}</td>"
                            f"<td style='text-align:right;color:#9ca3af;'>{_fmt_num(r['작년거래액'])}</td>"
                            f"<td style='text-align:right;'>{format_delta_html(r['거래액증감'])}</td>"
                            f"<td style='text-align:right;'>{_fmt_num(r['쿠폰할인'])}</td>"
                            f"<td style='text-align:right;color:#9ca3af;'>{_fmt_num(r['작년쿠폰할인'])}</td>"
                            f"<td style='text-align:right;'>{format_delta_html(r['쿠폰할인증감'])}</td>"
                            f"<td style='text-align:right;'>{_fmt_pct(r['비용률'])}</td>"
                            f"<td style='text-align:right;color:#9ca3af;'>{_fmt_pct(r['작년비용률'])}</td>"
                            f"<td style='text-align:right;'>{format_delta_html(r['비용률증감'])}</td></tr>"
                        )
                    st.markdown(
                        "<div style='overflow-x:auto;'><table class='summary-table'>"
                        "<thead><tr><th>연월</th>"
                        "<th colspan='3' style='text-align:center;background:#eef2ff;'>거래액</th>"
                        "<th colspan='3' style='text-align:center;background:#fef3c7;'>쿠폰할인</th>"
                        "<th colspan='3' style='text-align:center;background:#fce7f3;'>비용률</th>"
                        "</tr><tr><th></th>"
                        "<th style='text-align:right;'>올해</th><th style='text-align:right;'>작년</th><th style='text-align:right;'>증감</th>"
                        "<th style='text-align:right;'>올해</th><th style='text-align:right;'>작년</th><th style='text-align:right;'>증감</th>"
                        "<th style='text-align:right;'>올해</th><th style='text-align:right;'>작년</th><th style='text-align:right;'>증감</th>"
                        "</tr></thead>"
                        f"<tbody>{_yoy_body}</tbody></table></div>",
                        unsafe_allow_html=True,
                    )

                    _yoy_export = pd.DataFrame(_yoy_rows)
                    _yoy_export["연월"] = _yoy_export["연월"].dt.strftime("%Y-%m")
                    render_excel_download(_yoy_export, f"쿠폰비용_전년비교_{coupon_bpu}_{coupon_type_sel}.xlsx")
                else:
                    st.caption("ℹ️ 전년비 비교표는 월별 조회에서만 제공돼요 (일별/주별 전년비교는 아직 지원 예정 기능이에요).")

        # --- 구분별(자사/정상/이월/입점) 조회단위+기준시점으로 선택한 '그 기간'의 비용·비용률
        # — 상단 매체 필터(coupon_bpu)와 무관하게 전체 구분을 한번에 보여준다. ---
        st.markdown("---")
        _cost_abs_last = df_traffic["날짜"].max() if not df_traffic.empty else df_coupon_daily["날짜"].max()
        # coupon_ref_ts(선택된 기준시점)+coupon_unit으로 '그 기간'의 시작~끝을 직접 계산한다.
        # coupon_ref_ts는 이미 그 기간의 시작(일별=그날, 주별=그 주 월요일, 월별=그 달 1일)
        # 라벨이라, raw_cutoff_date로 끝만 구하면 된다. 이전엔 _chart_series(일별 전용 좁힘
        # 변수)를 재사용했는데, 주별/월별에선 안 좁혀져서 늘 '전체 누적'으로 보이는
        # 문제가 있었음 — 이제 선택한 기준시점 그 기간만 정확히 반영된다.
        _cost_ytd_start = coupon_ref_ts
        _cost_abs_last = min(raw_cutoff_date(coupon_ref_ts, coupon_unit), _cost_abs_last)
        st.markdown(
            f"**구분별 비용 현황**  ·  <span style='color:#6b7280;font-size:0.85rem'>"
            f"{_cost_ytd_start.strftime('%Y-%m-%d')} ~ {_cost_abs_last.strftime('%Y-%m-%d')} "
            f"({coupon_unit} 기준시점, 실적/예상 아님)</span>",
            unsafe_allow_html=True,
        )
        _cost_rows = []
        for _row_name, _bpu_list in FORECAST_BPU_ROWS.items():
            if _bpu_list is None:
                _cost_sub = (
                    df_coupon_daily[df_coupon_daily["BPU"] == "Total"]
                    if "Total" in df_coupon_daily["BPU"].unique().tolist()
                    else df_coupon_daily[df_coupon_daily["BPU"].isin(["e-영업1", "e-영업2", "e-영업3", "e-영업4"])]
                )
                _gmv_sub = df_traffic[(df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")]
            else:
                _cost_sub = df_coupon_daily[df_coupon_daily["BPU"].isin(_bpu_list)]
                _gmv_sub = df_traffic[(df_traffic["BPU"].isin(_bpu_list)) & (df_traffic["회원구분"] == "전체")]
            _cost_sub = _cost_sub[(_cost_sub["날짜"] >= _cost_ytd_start) & (_cost_sub["날짜"] <= _cost_abs_last)]
            _gmv_sub = _gmv_sub[(_gmv_sub["날짜"] >= _cost_ytd_start) & (_gmv_sub["날짜"] <= _cost_abs_last)]
            _cost_v = _cost_sub["쿠폰할인"].sum() if not _cost_sub.empty else 0.0
            _gmv_v = _gmv_sub["거래액"].sum() if not _gmv_sub.empty else 0.0
            _rate_v = (_cost_v / _gmv_v * 100) if _gmv_v else None
            _cost_rows.append({"구분": _row_name, "비용(쿠폰할인)": _cost_v, "거래액": _gmv_v, "비용률": _rate_v})
        _cost_df = pd.DataFrame(_cost_rows).set_index("구분")
        _cost_styled = _cost_df.style.format({
            "비용(쿠폰할인)": lambda v: f"{v:,.0f}",
            "거래액": lambda v: f"{v:,.0f}",
            "비용률": lambda v: "-" if v is None or pd.isna(v) else f"{v:.2f}%",
        })
        st.dataframe(_cost_styled, use_container_width=True)

        # 규칙 기반 자동 인사이트 — Total 행 비용률을 재사용(bpu_rows 없이 coupon_stats만
        # 넘겨서 ⑤번 섹션만 나오게 함).
        try:
            _total_row = next((r for r in _cost_rows if r["구분"] == "Total"), None)
            if _total_row and _total_row.get("비용률") is not None:
                _coupon_rb = generate_rule_based_insights([], {}, coupon_stats={"비용률": _total_row["비용률"]})
                render_insight_panel(_coupon_rb)
        except Exception:
            pass


# ============================================================
# 페이지 6: 주간보고용 (전년동요일 요약 엑셀 다운로드)
# ============================================================
if side["page"].startswith("11."):
    st.markdown("---")
    st.markdown("### 📑 주간보고용 · 전년동요일 요약")

    if df_traffic.empty and df_category.empty:
        st.info("데이터가 없습니다. 사이드바에서 EP실적/카테고리 CSV를 업로드해주세요.")
    else:
        # 이 페이지는 항상 '월별' 기준으로 만든다 (사이드바 조회단위와 무관 — 예전 스크린샷과
        # 동일하게 "진행 중인 달 1~N일" vs "작년 같은 날짜들"을 비교하는 게 목적이라서).
        _wk_unit = "월별"

        if not df_traffic.empty:
            _wk_base = df_traffic[(df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")] \
                if "회원구분" in df_traffic.columns else df_traffic[df_traffic["BPU"] == "Total"]
            _all_dates = pd.DatetimeIndex(sorted(_wk_base["날짜"].unique()))
        else:
            _all_dates = pd.DatetimeIndex(sorted(df_category["날짜"].unique()))

        # 카테고리 세그먼트/핏플랍 제외는 2번 페이지에서 지금 설정해둔 값을 그대로 가져온다
        # (2번 페이지를 안 들렀으면 기본값: 전체 세그먼트, 핏플랍 포함)
        _wk_cat_segment = st.session_state.get("cat_seg_filter", "전체")
        _wk_ff_exclude = st.session_state.get("cat_ff_exclude", False)

        _rule = UNIT_CONFIG[_wk_unit]["rule"]
        _wk_s = pd.Series(1, index=_all_dates).resample(_rule).sum()
        _wk_periods = list(_wk_s.index)
        _wk_labels = [make_period_label(d, _wk_unit) for d in _wk_periods]
        _wk_c1, _wk_c2 = st.columns([1, 3])
        with _wk_c1:
            st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>기준 시점(월)</div>", unsafe_allow_html=True)
            _wk_sel_label = st.selectbox(
                "기준 시점", _wk_labels, index=len(_wk_labels) - 1,
                label_visibility="collapsed", key="wk_ref_period",
            )
        _wk_ref = _wk_periods[_wk_labels.index(_wk_sel_label)]
        st.caption(
            f"BPU별 상세는 **1. 실적 요약**과 동일한 로직, 카테고리별 거래액은 **2. 카테고리 실적 요약**과 "
            f"동일한 로직(세그먼트: {_wk_cat_segment}"
            + (", 핏플랍 제외 적용" if _wk_ff_exclude else "")
            + ")을 그대로 재사용해서 만들어요."
        )
        if _wk_ff_exclude:
            st.caption(
                "ℹ️ 핏플랍 제외는 **카테고리별 시트에만** 반영돼요 — EP실적 원본(ep_traffic.csv)엔 "
                "브랜드 정보가 없어서 BPU별 시트에서는 애초에 제외할 수가 없어요. 그래서 이 상태로는 "
                "두 시트의 '전체' 합계가 핏플랍 매출만큼 차이 날 수 있어요."
            )

        try:
            _wk_xlsx, _pv_left, _pv_right, _pv_right_traffic = build_weekly_report_excel(
                _wk_unit, _wk_ref, df_traffic, df_category, _wk_cat_segment, _wk_ff_exclude
            )

            # 실제 계산(build_weekly_report_excel)이 selected_period_date(=_wk_ref)를 그대로
            # 쓰니까, 화면 표시용 날짜 범위도 그거 기준으로 그대로 보여준다 (전엔 두 원본 파일
            # 중 더 짧은 쪽에 맞추는 로직이 있었는데, 실제로 필요했던 적이 없고 오히려 1·2번
            # 페이지 숫자랑 안 맞는 원인이 돼서 제거함 — build_weekly_report_excel 쪽 주석 참고).
            _wk_month_start = pd.Timestamp(_wk_ref).replace(day=1)
            _wk_cur_days = _all_dates[(_all_dates >= _wk_month_start) & (_all_dates <= pd.Timestamp(_wk_ref))]
            _wk_prev_days = pd.DatetimeIndex([d - pd.Timedelta(days=364) for d in _wk_cur_days])
            _wk_cur_rng = f"{_wk_cur_days.min().strftime('%y.%m.%d')} - {_wk_cur_days.max().strftime('%m.%d')}" if len(_wk_cur_days) else "-"
            _wk_prev_rng = f"{_wk_prev_days.min().strftime('%y.%m.%d')} - {_wk_prev_days.max().strftime('%m.%d')}" if len(_wk_prev_days) else "-"

            _wk_tr_max = df_traffic["날짜"].max() if not df_traffic.empty else None
            _wk_cat_max = df_category["날짜"].max() if not df_category.empty else None
            if _wk_tr_max is not None and _wk_cat_max is not None and _wk_tr_max.normalize() != _wk_cat_max.normalize():
                st.caption(
                    f"ℹ️ 참고: EP실적 원본은 {_wk_tr_max.strftime('%Y-%m-%d')}까지, 카테고리 원본은 "
                    f"{_wk_cat_max.strftime('%Y-%m-%d')}까지 있어요. BPU별/카테고리별 시트는 각각 원본 "
                    "그대로(1·2번 페이지와 동일 기준) 계산되니, 두 시트의 마지막 날짜가 다를 수 있어요."
                )

            st.markdown(
                f"<div style='background:#f8fafc;border:1px solid #e5e7eb;border-radius:8px;padding:10px 14px;margin:8px 0;'>"
                f"<div style='font-size:0.85rem;color:#374151;'>📅 <b>올해</b> {_wk_cur_rng}"
                f" &nbsp;vs&nbsp; <b>전년 동요일</b> {_wk_prev_rng}</div>"
                f"<div style='font-size:0.76rem;color:#6b7280;margin-top:3px;'>"
                f"트래픽·거래액·구매객수는 일평균(1·2번 페이지 KPI 카드와 동일 기준), CR·객단가는 그 평균에서 재계산 · 거래액 없는 카테고리 제외</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

            st.download_button(
                "⬇️ 전년동요일 요약 엑셀 다운로드",
                data=_wk_xlsx,
                file_name=f"EP_주간보고_전년동요일_{pd.Timestamp(_wk_ref).strftime('%Y%m')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
            )

            st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

            def _style_yoy_delta(df):
                """전년비(%) 컬럼: 증가=초록, 감소=빨강+△ (엑셀 서식과 동일 규칙)."""
                def _fmt(v):
                    if pd.isna(v):
                        return ""
                    return f"{v:.1f}%" if v >= 0 else f"△{abs(v):.1f}%"
                def _color(v):
                    if pd.isna(v):
                        return ""
                    return "color:#16a34a;" if v >= 0 else "color:#dc2626;"
                sty = df.style
                if "전년비(%)" in df.columns:
                    sty = sty.format({"전년비(%)": _fmt}).map(_color, subset=["전년비(%)"])
                return sty

            _pc1, _pc2, _pc3 = st.columns(3)
            with _pc1:
                st.markdown("**BPU별**")
                # 화면 미리보기용으로만 CR(구매전환율) 행을 "4.8%"같은 보기 좋은 문자열로
                # 바꾼다 — 엑셀 파일 쪽은 진짜 퍼센트 셀서식을 쓰니 그대로(소수) 둬야
                # 정상 작동하고, 여기 화면 표시본만 따로 포맷한다.
                _pv_left_disp = _pv_left.copy()
                _cr_mask = _pv_left_disp["지표"] == "구매전환율(%)"
                _disp_cur_year = pd.Timestamp(_wk_ref).year
                _disp_col_prev, _disp_col_cur = f"{_disp_cur_year - 1}년", f"{_disp_cur_year}년"
                for _c in [_disp_col_prev, _disp_col_cur]:
                    if _c in _pv_left_disp.columns:
                        # float 컬럼에 문자열("6.0%")을 바로 대입하면 최신 pandas에서
                        # dtype 에러가 나서, 먼저 object로 바꾼 뒤 대입한다.
                        _pv_left_disp[_c] = _pv_left_disp[_c].astype(object)
                        _pv_left_disp.loc[_cr_mask, _c] = _pv_left_disp.loc[_cr_mask, _c].apply(
                            lambda v: f"{v*100:.1f}%" if pd.notna(v) else v
                        )
                st.dataframe(_style_yoy_delta(_pv_left_disp), use_container_width=True, hide_index=True, height=360)
            with _pc2:
                st.markdown("**카테고리별 (거래액)**")
                st.dataframe(_style_yoy_delta(_pv_right), use_container_width=True, hide_index=True, height=360)
            with _pc3:
                st.markdown("**카테고리별 (트래픽)**")
                st.dataframe(_style_yoy_delta(_pv_right_traffic), use_container_width=True, hide_index=True, height=360)
        except Exception as _e:
            st.error(f"요약 엑셀 생성 중 문제가 발생했어요: {_e}")

        # --- 최근 4주 일평균 표 (Total/자사/정상/이월/입점, 전주비/전년비) ---
        st.markdown("---")
        st.markdown("**최근 4주 일평균**")
        _wk4_abs_last = df_traffic["날짜"].max()
        _wk4_ref_monday = _wk4_abs_last - pd.Timedelta(days=_wk4_abs_last.weekday())
        _wk4_starts = [_wk4_ref_monday - pd.Timedelta(weeks=w) for w in range(3, -1, -1)]  # 오래된 주 -> 최신 주 순
        _wk4_labels = [f"{effective_month_of_week(d).month}월 {week_of_month(d)}주차" for d in _wk4_starts]

        _wk4_metric_defs = [
            ("거래액", "거래액", None, False),
            ("트래픽", "트래픽", None, False),
            ("구매객수", "구매객수", None, False),
            ("CR", "구매객수", "트래픽", True),
            ("객단가", "거래액", "구매객수", True),
        ]

        def _wk4_series_for(bpu_list, num_col, den_col):
            sub = df_traffic if bpu_list is None else df_traffic[df_traffic["BPU"].isin(bpu_list)]
            sub = sub[(sub["BPU"] == "Total") & (sub["회원구분"] == "전체")] if bpu_list is None else sub[sub["회원구분"] == "전체"]
            # bpu_list가 여러 개(자사=e1+e2, 입점=e3+e4)면 날짜별로 먼저 합산해야 한다 —
            # 그냥 이어붙이기만 하면 나중에 .mean()을 취할 때 'e1값과 e2값의 평균'이 돼버려서
            # (실제로 원하는 e1+e2 합산의 절반으로) 반토막 나는 버그가 있었음.
            _cols = [num_col] + ([den_col] if den_col else [])
            if bpu_list is not None and len(bpu_list) > 1:
                sub = sub.groupby("날짜", as_index=False)[_cols].sum()
            s_num = sub.set_index("날짜")[num_col].sort_index()
            s_den = sub.set_index("날짜")[den_col].sort_index() if den_col else None
            return s_num, s_den

        _wk4_all_rows = []  # HTML/엑셀 공통으로 쓸 구조화된 결과
        for _label, _num_col, _den_col, _is_ratio in _wk4_metric_defs:
            for _row_name, _bpu_list in FORECAST_BPU_ROWS.items():
                _s_num, _s_den = _wk4_series_for(_bpu_list, _num_col, _den_col)
                _wk_vals = []
                for _ws in _wk4_starts:
                    _we = _ws + pd.Timedelta(days=6)
                    _num_wk = _s_num[(_s_num.index >= _ws) & (_s_num.index <= _we)]
                    if _is_ratio:
                        _den_wk = _s_den[(_s_den.index >= _ws) & (_s_den.index <= _we)]
                        _v = (_num_wk.sum() / _den_wk.sum() * (100 if _label == "CR" else 1)) if _den_wk.sum() else None
                    else:
                        _v = _num_wk.mean() if not _num_wk.empty else None
                    _wk_vals.append(_v)
                _latest_v, _prev_v = _wk_vals[-1], _wk_vals[-2]
                _wow = pct_delta_safe(_latest_v, _prev_v) if (_latest_v is not None and _prev_v) else None
                # 전년비: 최신 주에 '실제로 존재하는 날짜'만 골라서 그 날짜들의 동요일(364일 전)과
                # 매칭한다 — 최신 주가 아직 다 안 지난(부분) 주면, 작년도 7일 전체가 아니라
                # 딱 그만큼의 날짜만 비교해야 공정하다(1번 페이지 compute_kpi_deltas와 동일 원칙).
                # 그냥 7일 전체로 비교하면 부분주(예: 3일치)를 작년 완성된 7일과 비교하게 돼서
                # 부당하게 낮게(또는 높게) 나옴 — 실제로 이 버그를 사용자가 발견함.
                _latest_ws = _wk4_starts[-1]
                _latest_we = _latest_ws + pd.Timedelta(days=6)
                _cur_actual_days = _s_num[(_s_num.index >= _latest_ws) & (_s_num.index <= _latest_we)].index
                _matched_dates = [d - pd.Timedelta(days=364) for d in _cur_actual_days]
                _num_yoy = _s_num.reindex(_matched_dates).dropna()
                if _is_ratio:
                    _den_yoy = _s_den.reindex(_matched_dates).dropna()
                    _yoy_v = (_num_yoy.sum() / _den_yoy.sum() * (100 if _label == "CR" else 1)) if _den_yoy.sum() else None
                else:
                    _yoy_v = _num_yoy.mean() if not _num_yoy.empty else None
                _yoy = pct_delta_safe(_latest_v, _yoy_v) if (_latest_v is not None and _yoy_v) else None
                _wk4_all_rows.append({
                    "지표": _label, "구분": _row_name, "값": _wk_vals,
                    "전주비": _wow, "전년비": _yoy, "작년값": _yoy_v, "is_pct": _label == "CR",
                })

        # --- HTML 표 ---
        _wk4_sections_html = ""
        _cur_metric = None
        for _r in _wk4_all_rows:
            if _r["지표"] != _cur_metric:
                _cur_metric = _r["지표"]
                _wk4_sections_html += f"<tr><td colspan='8' style='background:#eef2ff;font-weight:700;'>{_cur_metric}</td></tr>"

            def _fmt_wk(v, is_pct):
                if v is None or pd.isna(v):
                    return "-"
                return f"{v:.1f}%" if is_pct else f"{v:,.0f}"

            _cells = "".join(f"<td style='text-align:right;'>{_fmt_wk(v, _r['is_pct'])}</td>" for v in _r["값"])
            _wk4_sections_html += (
                f"<tr><td class='m'>{_r['구분']}</td>{_cells}"
                f"<td style='text-align:right;'>{format_delta_html(_r['전주비'])}</td>"
                f"<td style='text-align:right;'>{format_delta_html(_r['전년비'])}</td>"
                f"<td style='text-align:right;color:#9ca3af;'>{_fmt_wk(_r['작년값'], _r['is_pct'])}</td></tr>"
            )
        _wk4_header = "".join(
            f"<th>{l}<br><span style='font-weight:400;font-size:0.72rem;color:#9ca3af;'>"
            f"{_ws.month}/{_ws.day}~{(_ws + pd.Timedelta(days=6)).month}/{(_ws + pd.Timedelta(days=6)).day}</span></th>"
            for l, _ws in zip(_wk4_labels, _wk4_starts)
        )
        st.markdown(
            "<div style='overflow-x:auto;'><table class='summary-table'>"
            f"<thead><tr><th>구분</th>{_wk4_header}<th>전주비</th><th>전년비</th><th>작년(동요일)</th></tr></thead>"
            f"<tbody>{_wk4_sections_html}</tbody></table></div>",
            unsafe_allow_html=True,
        )

        # --- 엑셀 다운로드 (CR은 진짜 %서식으로) ---
        from openpyxl.styles import Font, PatternFill
        import openpyxl
        _wk4_wb = openpyxl.Workbook()
        _wk4_ws = _wk4_wb.active
        _wk4_ws.title = "최근4주"
        _wk4_header_fill = PatternFill("solid", fgColor="D9D9D9")
        _wk4_section_font = Font(bold=True)
        _wk4_up_font = Font(color="16A34A")
        _wk4_down_font = Font(color="DC2626")

        _wk4_col_headers = ["구분"] + _wk4_labels + ["전주비", "전년비", "작년(동요일)"]
        _wk4_ws.append(_wk4_col_headers)
        for _c in range(1, len(_wk4_col_headers) + 1):
            _cell = _wk4_ws.cell(row=1, column=_c)
            _cell.fill = _wk4_header_fill
            _cell.font = Font(bold=True)

        _cur_metric = None
        for _r in _wk4_all_rows:
            if _r["지표"] != _cur_metric:
                _cur_metric = _r["지표"]
                _wk4_ws.append([_cur_metric])
                _wk4_ws.cell(row=_wk4_ws.max_row, column=1).font = _wk4_section_font
            _row_vals = [_r["구분"]]
            for v in _r["값"]:
                _row_vals.append(None if v is None or pd.isna(v) else (round(v / 100, 4) if _r["is_pct"] else round(v)))
            _row_vals.append(None if _r["전주비"] is None else round(_r["전주비"] / 100, 4))
            _row_vals.append(None if _r["전년비"] is None else round(_r["전년비"] / 100, 4))
            _yoy_v = _r.get("작년값")
            _row_vals.append(None if _yoy_v is None or pd.isna(_yoy_v) else (round(_yoy_v / 100, 4) if _r["is_pct"] else round(_yoy_v)))
            _wk4_ws.append(_row_vals)
            _rr = _wk4_ws.max_row
            if _r["is_pct"]:
                for _c in range(2, 2 + len(_wk4_labels)):
                    _wk4_ws.cell(row=_rr, column=_c).number_format = "0.0%"
                _wk4_ws.cell(row=_rr, column=4 + len(_wk4_labels)).number_format = "0.0%"
            for _c, _val in [(2 + len(_wk4_labels), _r["전주비"]), (3 + len(_wk4_labels), _r["전년비"])]:
                _cell = _wk4_ws.cell(row=_rr, column=_c)
                _cell.number_format = '+0.0%;-0.0%'
                if _val is not None:
                    _cell.font = _wk4_up_font if _val >= 0 else _wk4_down_font

        for _c in range(1, len(_wk4_col_headers) + 1):
            _wk4_ws.column_dimensions[openpyxl.utils.get_column_letter(_c)].width = 13
        _wk4_ws.freeze_panes = "B2"

        # --- 카테고리별 최근 4주 일평균 — BPU(Total/자사/정상/이월/입점) x 지표(거래액/
        # 트래픽/구매객수)로 세분화. 5x3=15개 조합이라 페이지가 너무 길어지지 않게
        # expander로 감싼다(기본 접힘). 부분주 보정 등 계산 원칙은 위 BPU별 표와 동일. ---
        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("**카테고리별 최근 4주 일평균 (BPU·지표별 세분화)**")

        _wk4_cat_metric_defs = [
            ("거래액", "거래액", None, False, False),
            ("트래픽", "트래픽", None, False, False),
            ("구매객수", "구매객수", None, False, False),
            ("CR", "구매객수", "트래픽", True, True),
            ("객단가", "거래액", "구매객수", True, False),
        ]

        _wk4_cat_base_all = pd.DataFrame()
        if not df_category.empty:
            _wk4_cat_base_all = df_category[(df_category["카테고리"] != "전체") & (df_category["브랜드"] == "전체")]
            if "회원구분" in _wk4_cat_base_all.columns:
                _wk4_cat_base_all = _wk4_cat_base_all[_wk4_cat_base_all["회원구분"] == _wk_cat_segment]
            if _wk_ff_exclude:
                _wk4_cat_base_all = exclude_ff_brand(_wk4_cat_base_all)

        def _fmt_wk4(v, is_pct=False):
            if v is None or pd.isna(v):
                return "-"
            return f"{v:.1f}%" if is_pct else f"{v:,.0f}"

        def _compute_cat_rows(base_df, num_col, den_col, is_ratio, is_pct, category_list):
            """base_df에서 category_list에 있는 카테고리 전부(값이 없으면 None)에 대해
            4주 흐름을 계산한다. category_list를 고정해서 넘기기 때문에, 어떤 지표를
            계산하든 항상 같은 카테고리 집합·순서가 나온다(지표마다 카테고리가 들쭉날쭉하게
            빠지는 문제 방지 — 예: 거래액엔 있는데 트래픽엔 0이라 안 보이던 카테고리)."""
            if base_df.empty or num_col not in base_df.columns or (den_col and den_col not in base_df.columns):
                return [{"카테고리": c, "값": [None] * 4, "전주비": None, "전년비": None, "작년값": None} for c in category_list]
            _cols = [num_col] + ([den_col] if den_col else [])
            daily = base_df.groupby(["날짜", "카테고리"], as_index=False)[_cols].sum()
            rows = []
            for _cat_name in category_list:
                _g = daily[daily["카테고리"] == _cat_name]
                if _g.empty:
                    rows.append({"카테고리": _cat_name, "값": [None] * 4, "전주비": None, "전년비": None, "작년값": None})
                    continue
                _s_num = _g.set_index("날짜")[num_col].sort_index()
                _s_den = _g.set_index("날짜")[den_col].sort_index() if den_col else None
                _wk_vals_cat = []
                for _ws in _wk4_starts:
                    _we = _ws + pd.Timedelta(days=6)
                    _num_wk = _s_num[(_s_num.index >= _ws) & (_s_num.index <= _we)]
                    if is_ratio:
                        _den_wk = _s_den[(_s_den.index >= _ws) & (_s_den.index <= _we)]
                        _v = (_num_wk.sum() / _den_wk.sum() * (100 if is_pct else 1)) if _den_wk.sum() else None
                    else:
                        _v = _num_wk.mean() if not _num_wk.empty else None
                    _wk_vals_cat.append(_v)
                _latest_v, _prev_v = _wk_vals_cat[-1], _wk_vals_cat[-2]
                _wow = pct_delta_safe(_latest_v, _prev_v) if (_latest_v is not None and _prev_v) else None
                _cur_days = _s_num[(_s_num.index >= _wk4_starts[-1]) & (_s_num.index <= _wk4_starts[-1] + pd.Timedelta(days=6))].index
                _matched = [d - pd.Timedelta(days=364) for d in _cur_days]
                _num_yoy = _s_num.reindex(_matched).dropna()
                if is_ratio:
                    _den_yoy = _s_den.reindex(_matched).dropna()
                    _yoy_v = (_num_yoy.sum() / _den_yoy.sum() * (100 if is_pct else 1)) if _den_yoy.sum() else None
                else:
                    _yoy_v = _num_yoy.mean() if not _num_yoy.empty else None
                _yoy = pct_delta_safe(_latest_v, _yoy_v) if (_latest_v is not None and _yoy_v) else None
                rows.append({"카테고리": _cat_name, "값": _wk_vals_cat, "전주비": _wow, "전년비": _yoy, "작년값": _yoy_v})
            return rows

        # 카테고리 순서를 '거래액' 기준으로 BPU별로 미리 한 번만 계산해서 고정한다 —
        # 그래야 트래픽/CR/객단가 표에서도 거래액 표와 항상 같은 카테고리 순서로 나온다
        # (지표마다 따로 정렬하면 표마다 카테고리 순서가 들쭉날쭉해지는 문제가 있었음).
        _cat_order_by_bpu = {}
        for _ob_label, _ob_list in FORECAST_BPU_ROWS.items():
            _ob_base = _wk4_cat_base_all if _ob_list is None else _wk4_cat_base_all[_wk4_cat_base_all["BPU"].isin(_ob_list)]
            _ob_all_cats = sorted(_ob_base["카테고리"].dropna().unique().tolist()) if not _ob_base.empty else []
            _ob_rows = _compute_cat_rows(_ob_base, "거래액", None, False, False, _ob_all_cats)
            _ob_rows.sort(key=lambda r: r["값"][-1] or 0, reverse=True)
            _cat_order_by_bpu[_ob_label] = [r["카테고리"] for r in _ob_rows]

        _wk4_cat_excel_ws = None
        if not _wk4_cat_base_all.empty:
            with st.expander("펼쳐서 보기 (지표 5개 x 구분 5개 = 25개 표)", expanded=False):
                for _m_label, _num_col, _den_col, _is_ratio, _is_pct in _wk4_cat_metric_defs:
                    st.markdown(f"##### {_m_label}")
                    for _bpu_label, _bpu_list in FORECAST_BPU_ROWS.items():
                        _base_bpu = (
                            _wk4_cat_base_all if _bpu_list is None
                            else _wk4_cat_base_all[_wk4_cat_base_all["BPU"].isin(_bpu_list)]
                        )
                        _order = _cat_order_by_bpu.get(_bpu_label, [])
                        if not _order:
                            continue
                        # 거래액 기준으로 확정한 카테고리 목록(_order)을 그대로 넘긴다 —
                        # 이 지표에서 값이 0/None이어도 행 자체는 그대로 유지되고 '-'로
                        # 표시되므로, 모든 지표 표가 항상 같은 카테고리 집합·순서를 갖는다.
                        _rows = _compute_cat_rows(_base_bpu, _num_col, _den_col, _is_ratio, _is_pct, _order)
                        st.markdown(f"**{_bpu_label} {_m_label}**")

                        # TOTAL 행 — 카테고리 전체 합산. 이미 위 BPU별 표에서 정확히 이 조합
                        # (지표=_m_label, 구분=_bpu_label)을 계산해둔 _wk4_all_rows를 그대로
                        # 재사용한다 — 새로 계산하면 반올림 등으로 미세하게 어긋날 수 있는데,
                        # 재사용하면 위 표와 100% 같은 숫자가 보장된다.
                        _total_match = next(
                            (r for r in _wk4_all_rows if r["지표"] == _m_label and r["구분"] == _bpu_label), None
                        )
                        _total_row_html = ""
                        if _total_match:
                            _total_row_html = (
                                "<tr style='background:#f3f4f6;font-weight:700;'><td class='m'>TOTAL</td>"
                                + "".join(f"<td style='text-align:right;'>{_fmt_wk4(v, _is_pct)}</td>" for v in _total_match["값"])
                                + f"<td style='text-align:right;'>{format_delta_html(_total_match['전주비'])}</td>"
                                f"<td style='text-align:right;'>{format_delta_html(_total_match['전년비'])}</td>"
                                f"<td style='text-align:right;color:#9ca3af;'>{_fmt_wk4(_total_match.get('작년값'), _is_pct)}</td></tr>"
                            )

                        _body = "".join(
                            f"<tr><td class='m'>{r['카테고리']}</td>"
                            + "".join(f"<td style='text-align:right;'>{_fmt_wk4(v, _is_pct)}</td>" for v in r["값"])
                            + f"<td style='text-align:right;'>{format_delta_html(r['전주비'])}</td>"
                            f"<td style='text-align:right;'>{format_delta_html(r['전년비'])}</td>"
                            f"<td style='text-align:right;color:#9ca3af;'>{_fmt_wk4(r['작년값'], _is_pct)}</td></tr>"
                            for r in _rows
                        )
                        st.markdown(
                            "<div style='overflow-x:auto;margin-bottom:14px;'><table class='summary-table'>"
                            f"<thead><tr><th>카테고리</th>{_wk4_header}<th>전주비</th><th>전년비</th><th>작년(동요일)</th></tr></thead>"
                            f"<tbody>{_total_row_html}{_body}</tbody></table></div>",
                            unsafe_allow_html=True,
                        )

                        # 엑셀 시트(지표 x BPU 조합 하나당 섹션으로 누적)
                        if _wk4_cat_excel_ws is None:
                            _wk4_cat_excel_ws = _wk4_wb.create_sheet("카테고리별")
                        _wk4_cat_excel_ws.append([f"{_bpu_label} · {_m_label}"])
                        _wk4_cat_excel_ws.cell(row=_wk4_cat_excel_ws.max_row, column=1).font = Font(bold=True)
                        _hdr_row = ["카테고리"] + _wk4_labels + ["전주비", "전년비", "작년(동요일)"]
                        _wk4_cat_excel_ws.append(_hdr_row)
                        for _c in range(1, len(_hdr_row) + 1):
                            _cell = _wk4_cat_excel_ws.cell(row=_wk4_cat_excel_ws.max_row, column=_c)
                            _cell.fill = _wk4_header_fill
                            _cell.font = Font(bold=True)
                        if _total_match:
                            _tot_vals = ["TOTAL"] + [
                                None if v is None or pd.isna(v) else (round(v / 100, 4) if _is_pct else round(v))
                                for v in _total_match["값"]
                            ]
                            _tot_vals.append(None if _total_match["전주비"] is None else round(_total_match["전주비"] / 100, 4))
                            _tot_vals.append(None if _total_match["전년비"] is None else round(_total_match["전년비"] / 100, 4))
                            _tw = _total_match.get("작년값")
                            _tot_vals.append(None if _tw is None or pd.isna(_tw) else (round(_tw / 100, 4) if _is_pct else round(_tw)))
                            _wk4_cat_excel_ws.append(_tot_vals)
                            _rr = _wk4_cat_excel_ws.max_row
                            for _c in range(1, len(_hdr_row) + 1):
                                _wk4_cat_excel_ws.cell(row=_rr, column=_c).font = Font(bold=True)
                            if _is_pct:
                                for _c in range(2, 2 + len(_wk4_labels)):
                                    _wk4_cat_excel_ws.cell(row=_rr, column=_c).number_format = "0.0%"
                                _wk4_cat_excel_ws.cell(row=_rr, column=len(_hdr_row)).number_format = "0.0%"
                            for _c, _val in [(len(_hdr_row) - 2, _total_match["전주비"]), (len(_hdr_row) - 1, _total_match["전년비"])]:
                                _cell = _wk4_cat_excel_ws.cell(row=_rr, column=_c)
                                _cell.number_format = '+0.0%;-0.0%'
                                if _val is not None:
                                    _cell.font = _wk4_up_font if _val >= 0 else _wk4_down_font
                        for r in _rows:
                            _row_vals = [r["카테고리"]] + [
                                None if v is None or pd.isna(v) else (round(v / 100, 4) if _is_pct else round(v))
                                for v in r["값"]
                            ]
                            _row_vals.append(None if r["전주비"] is None else round(r["전주비"] / 100, 4))
                            _row_vals.append(None if r["전년비"] is None else round(r["전년비"] / 100, 4))
                            _wv = r["작년값"]
                            _row_vals.append(
                                None if _wv is None or pd.isna(_wv) else (round(_wv / 100, 4) if _is_pct else round(_wv))
                            )
                            _wk4_cat_excel_ws.append(_row_vals)
                            _rr = _wk4_cat_excel_ws.max_row
                            if _is_pct:
                                for _c in range(2, 2 + len(_wk4_labels)):
                                    _wk4_cat_excel_ws.cell(row=_rr, column=_c).number_format = "0.0%"
                                _wk4_cat_excel_ws.cell(row=_rr, column=len(_hdr_row)).number_format = "0.0%"
                            for _c, _val in [(len(_hdr_row) - 1, r["전주비"]), (len(_hdr_row), r["전년비"])]:
                                _cell = _wk4_cat_excel_ws.cell(row=_rr, column=_c)
                                _cell.number_format = '+0.0%;-0.0%'
                                if _val is not None:
                                    _cell.font = _wk4_up_font if _val >= 0 else _wk4_down_font
                        _wk4_cat_excel_ws.append([])  # 섹션 사이 빈 줄
            if _wk4_cat_excel_ws is not None:
                for _c in range(1, 8):
                    _wk4_cat_excel_ws.column_dimensions[openpyxl.utils.get_column_letter(_c)].width = 13
                _wk4_cat_excel_ws.freeze_panes = "A2"
        else:
            st.info("카테고리 데이터가 없어서 카테고리별 흐름은 건너뛰었어요.")

        _wk4_buf = io.BytesIO()
        _wk4_wb.save(_wk4_buf)
        st.download_button(
            "⬇️ 엑셀 다운로드 (최근 4주 일평균)",
            data=_wk4_buf.getvalue(),
            file_name=f"최근4주일평균_{_wk4_abs_last.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # 규칙 기반 자동 인사이트 — 이미 계산된 _wk4_all_rows(거래액/Total 행)를 재사용해서
        # "최근 주차 실적" 한 줄 요약을 만든다.
        try:
            _wk4_gmv_total = next(
                (r for r in _wk4_all_rows if r["지표"] == "거래액" and r["구분"] == "Total"), None
            )
            if _wk4_gmv_total:
                _latest_wk_v = _wk4_gmv_total["값"][-1]
                _wow_v, _yoy_v = _wk4_gmv_total.get("전주비"), _wk4_gmv_total.get("전년비")
                _wk4_body = f"최신 주차({_wk4_labels[-1]}) 거래액 일평균 {_latest_wk_v:,.0f}"
                if _wow_v is not None:
                    _wk4_body += f" · 전주비 {format_delta_html(_wow_v)}"
                if _yoy_v is not None:
                    _wk4_body += f" · 전년비(동요일) {format_delta_html(_yoy_v)}"
                render_insight_panel([{"title": "① 최근 주차 실적은?", "body": _wk4_body}])
        except Exception:
            pass


# ============================================================
# 페이지 7/8/9: 전체/회원/신규 실적 (주차별) — 25년 vs 26년 vs 25년(FF제외)
# 세 페이지가 트래픽 세그먼트만 다르고 나머지 구조는 완전히 동일해서 공유 함수로 처리.
# ============================================================
def _render_weekly_segment_page(page_title, traffic_segment):
    """page_title: 화면에 보일 이름(예: '전체 실적'). traffic_segment: 트래픽 섹션에 쓸
    회원구분('전체'/'회원'/'신규'). 거래액 섹션은 기존에 정한 대로 항상 '전체' 세그먼트를 쓴다
    (거래액은 세그먼트 구분보다 전체 기준으로 보는 게 맞다고 판단했던 부분 그대로 유지)."""
    st.markdown("---")
    st.markdown(f"### 📅 {page_title} · 주차별 25년 vs 26년")
    st.caption(
        "각 연도 1월 1일부터 월~일 단위로 주차를 매겨서(그 해의 1주차부터), "
        "올해/작년 같은 주차끼리 나란히 비교해요. 자사 정상=e-영업1, 자사 이월=e-영업2, "
        f"입점=e-영업3+e-영업4. 트래픽·거래액 둘 다 '{traffic_segment}' 세그먼트 기준이에요."
    )

    if df_traffic.empty:
        st.info("데이터가 없습니다. 사이드바에서 ep_traffic.csv를 업로드해주세요.")
        return

    # 최신(마지막) 주차의 정확한 날짜 범위를 26년/25년 둘 다 보여준다 — 부분주(진행중인
    # 주)일 때 "정확히 며칠까지 들어있는지" 헷갈리기 쉬워서 (방금 발견한 부분주 비교
    # 버그도 이런 확인 과정에서 나온 것).
    _wk_ref_total = df_traffic[(df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")]
    if not _wk_ref_total.empty:
        _wk_ref_26 = _wk_ref_total[_wk_ref_total["날짜"].dt.year == 2026]
        if not _wk_ref_26.empty:
            _wk_last_date_26 = _wk_ref_26["날짜"].max()
            _wk_mon0_26 = pd.Timestamp("2026-01-01") - pd.Timedelta(days=pd.Timestamp("2026-01-01").weekday())
            _wk_last_week_num = ((_wk_last_date_26 - _wk_mon0_26).days // 7) + 1
            _wk_week_start_26 = _wk_mon0_26 + pd.Timedelta(weeks=_wk_last_week_num - 1)
            _wk_actual_days_26 = _wk_ref_26[
                (_wk_ref_26["날짜"] >= _wk_week_start_26) & (_wk_ref_26["날짜"] <= _wk_week_start_26 + pd.Timedelta(days=6))
            ]["날짜"].sort_values()
            if not _wk_actual_days_26.empty:
                _wk_mon0_25 = pd.Timestamp("2025-01-01") - pd.Timedelta(days=pd.Timestamp("2025-01-01").weekday())
                _wk_week_start_25 = _wk_mon0_25 + pd.Timedelta(weeks=_wk_last_week_num - 1)
                _wk_weekday_positions = sorted(set(d.weekday() for d in _wk_actual_days_26))
                _wk_matching_25 = [_wk_week_start_25 + pd.Timedelta(days=wd) for wd in _wk_weekday_positions]

                def _md(d):
                    return f"{d.month}/{d.day}"

                _wk_26_rng = f"{_md(_wk_actual_days_26.min())}-{_md(_wk_actual_days_26.max())}"
                _wk_25_rng = f"{_md(min(_wk_matching_25))}-{_md(max(_wk_matching_25))}"
                _wk_partial_note = "" if len(_wk_actual_days_26) >= 7 else f" (진행 중 — {len(_wk_actual_days_26)}일치)"
                st.caption(
                    f"📅 최신 주차({_wk_last_week_num}주차) 기준: 26년 {_wk_26_rng}{_wk_partial_note}"
                    f"  vs  25년(동요일) {_wk_25_rng}"
                )

    def _render_weekly_yearly_chart(title, s_by_label, wk_range, color_scheme=None, y_domain=None, week_labels=None):
        """s_by_label: {라벨: Series(주차 정수 인덱스)} — 라벨 순서대로 그린다.
        wk_range: (시작주차, 끝주차) — 이 범위만 잘라서 표시.
        color_scheme: {"25년":색, "26년":색, "25년(FF제외)":색} — 차트마다 다른 색 팔레트.
        y_domain: (최소,최대) 주어지면 y축을 이 값으로 고정한다(같은 섹션의 4개 차트가
        같은 y축 범위를 공유해서 서로 높이를 비교할 수 있게 하기 위함). 없으면 차트별로
        자동 확대(zero=False).
        week_labels: 주차 정수(1-based) -> 'N월 N주차' 라벨 리스트. 주면 x축에 숫자 대신
        이 라벨을 쓴다(상단 주차 범위 슬라이더와 동일한 라벨 체계)."""
        import altair as alt
        _colors = color_scheme or {"25년": "#2563eb", "26년": "#f97316", "25년(FF제외)": "#9ca3af"}
        frames = []
        for label, s in s_by_label.items():
            if s is not None and not s.empty:
                s = s[(s.index >= wk_range[0]) & (s.index <= wk_range[1])]
                if not s.empty:
                    frames.append(pd.DataFrame({"주차": s.index, "값": s.values, "구분": label}))
        if not frames:
            st.info("선택한 주차 범위에 데이터가 없습니다.")
            return
        long_df = pd.concat(frames, ignore_index=True)

        # "26년" 라인 위에 마우스 올렸을 때 전년비(%)도 같이 보이게 — 같은 주차의 "25년"
        # 값과 비교해서 계산한다. 26년이 아닌 행(25년/25년(FF제외))에는 안 채운다.
        long_df["전년비"] = pd.NA
        if "25년" in s_by_label and "26년" in s_by_label:
            _s25 = s_by_label["25년"]
            _s26_mask = long_df["구분"] == "26년"
            for idx in long_df[_s26_mask].index:
                _wk = long_df.at[idx, "주차"]
                if _s25 is not None and _wk in _s25.index and pd.notna(_s25.loc[_wk]) and _s25.loc[_wk] != 0:
                    long_df.at[idx, "전년비"] = pct_delta_safe(long_df.at[idx, "값"], _s25.loc[_wk])
        long_df["전년비_표시"] = long_df["전년비"].apply(
            lambda v: f"{'' if v >= 0 else '△'}{abs(v):.1f}%" if pd.notna(v) else "-"
        )

        _domain = [d for d in s_by_label.keys() if d in long_df["구분"].unique()]
        _range = [_colors.get(d, "#000000") for d in _domain]
        # "25년(FF제외)"만 점선으로 그려서 실제값(25/26년)과 구분되게 함
        _dash_range = [[1, 0] if d != "25년(FF제외)" else [5, 3] for d in _domain]
        _y_scale = alt.Scale(domain=list(y_domain)) if y_domain is not None else alt.Scale(zero=False)

        if week_labels:
            long_df["주차_라벨"] = long_df["주차"].apply(
                lambda w: week_labels[w - 1] if 0 <= w - 1 < len(week_labels) else str(w)
            )
            _wk_uniq = sorted(long_df["주차"].unique())
            _wk_sort = [week_labels[w - 1] if 0 <= w - 1 < len(week_labels) else str(w) for w in _wk_uniq]
            _x_field, _x_sort = "주차_라벨:O", _wk_sort
            _x_tooltip_field, _x_tooltip_title = "주차_라벨:N", "주차"
        else:
            _x_field, _x_sort = "주차:O", None
            _x_tooltip_field, _x_tooltip_title = "주차:O", "주차"

        chart = (
            alt.Chart(long_df)
            .mark_line(strokeWidth=3, point=alt.OverlayMarkDef(size=32, filled=True))
            .encode(
                x=alt.X(_x_field, title=None, sort=_x_sort, axis=alt.Axis(labelAngle=-90, labelFontSize=8, labelPadding=2)),
                y=alt.Y("값:Q", title=None, axis=alt.Axis(format="~s"), scale=_y_scale),
                color=alt.Color(
                    "구분:N", scale=alt.Scale(domain=_domain, range=_range),
                    legend=alt.Legend(orient="bottom", title=None),
                ),
                strokeDash=alt.StrokeDash(
                    "구분:N", scale=alt.Scale(domain=_domain, range=_dash_range), legend=None,
                ),
                tooltip=[
                    alt.Tooltip(_x_tooltip_field, title=_x_tooltip_title),
                    alt.Tooltip("구분:N", title="구분"),
                    alt.Tooltip("값:Q", title="값", format=",.0f"),
                    alt.Tooltip("전년비_표시:N", title="전년비"),
                ],
            )
            .properties(height=280)
            .interactive()  # 확대(fullscreen) 보기에서 휠 줌/드래그 팬 가능하게
        )
        st.markdown(f"**{title}**")
        st.altair_chart(chart, use_container_width=True)

    _has_ff_data = not df_category.empty and (df_category["브랜드"] == "FF").any()
    _tr_ff_adj = exclude_ff_from_traffic(df_traffic, df_category) if _has_ff_data else None

    # 차트(BPU)별로 다른 색 계열 — 25년=연한색, 26년=진한색, FF제외=같은 계열의 중간톤
    _charts_def = [
        ("전체", "single", "Total", {"25년": "#c7c7c7", "26년": "#7f7f7f", "25년(FF제외)": "#ffbb78"}),
        ("자사 정상", "single", "e-영업1", {"25년": "#aec7e8", "26년": "#1f77b4", "25년(FF제외)": "#ffbb78"}),
        ("자사 이월", "single", "e-영업2", {"25년": "#98df8a", "26년": "#2ca02c", "25년(FF제외)": "#ffbb78"}),
        ("입점", "multi", ["e-영업3", "e-영업4"], {"25년": "#c5b0d5", "26년": "#9467bd", "25년(FF제외)": "#ffbb78"}),
    ]

    # 주차 범위(wk_range)와 라벨(_wk7_labels)은 상단 고정 영역에서 이미 만들어져서 넘어온다.

    def _render_section(section_title, metric, segment_value, show_ff):
        """metric(트래픽/거래액), segment_value(세그먼트) 기준으로 4개 BPU 차트를 그린다.
        '전체'는 규모가 훨씬 커서 같이 묶으면 나머지가 다 눌려 보이므로 독자적인 y축을 쓰고,
        '자사 정상/자사 이월/입점' 3개는 서로 규모가 비슷해서 공통 y축을 공유해 흐름을
        비교할 수 있게 한다."""
        st.markdown(f"#### {section_title}")
        _base = df_traffic[df_traffic["회원구분"] == segment_value]
        _base_ff = None
        if show_ff and _tr_ff_adj is not None:
            _base_ff = _tr_ff_adj[_tr_ff_adj["회원구분"] == segment_value]

        # 1단계: 4개 차트 데이터를 먼저 만들면서, 화면에 실제로 표시될(주차 범위로 자른)
        # 값들의 최소/최대를 '전체' 따로, '정상/이월/입점' 따로 모은다.
        _chart_series = []
        _total_vals, _shared_vals = [], []
        for label, kind, bpu_sel, color_scheme in _charts_def:
            if kind == "single":
                _d = _base[_base["BPU"] == bpu_sel][["날짜", metric]]
            else:
                _d = _base[_base["BPU"].isin(bpu_sel)].groupby("날짜", as_index=False)[metric].sum()

            s2025 = _weekly_of_year(_d, metric, 2025)
            s2026 = _weekly_of_year(_d, metric, 2026)

            # 26년 마지막 주차가 아직 진행 중(부분)이면, 25년 같은 주차도 동일한 요일
            # 위치까지만 반영하도록 보정한다 (안 그러면 '26년 며칠치'를 '25년 7일 전체'와
            # 비교하는 불공정한 비교가 됨 — 2026-08-13에 실제로 이 문제가 있는 걸 확인함).
            _raw_target = _d.set_index("날짜")[metric].sort_index()
            s2025 = _correct_partial_week_yoy(s2026, s2025, _raw_target, _raw_target, 2026, 2025)

            s2025_ff = None
            if show_ff and kind == "single" and _base_ff is not None:
                _d_ff = _base_ff[_base_ff["BPU"] == bpu_sel][["날짜", metric]]
                s2025_ff = _weekly_of_year(_d_ff, metric, 2025)
                _raw_ff_ref = _d_ff.set_index("날짜")[metric].sort_index()
                s2025_ff = _correct_partial_week_yoy(s2026, s2025_ff, _raw_target, _raw_ff_ref, 2026, 2025)

            _series_map = {"25년": s2025, "26년": s2026}
            if s2025_ff is not None and not s2025_ff.empty:
                _series_map["25년(FF제외)"] = s2025_ff

            _chart_series.append((label, color_scheme, _series_map))
            _vals_bucket = _total_vals if label == "전체" else _shared_vals
            for s in _series_map.values():
                if s is not None and not s.empty:
                    _s_clip = s[(s.index >= wk_range[0]) & (s.index <= wk_range[1])]
                    if not _s_clip.empty:
                        _vals_bucket.append(float(_s_clip.min()))
                        _vals_bucket.append(float(_s_clip.max()))

        def _domain_from(vals):
            if not vals:
                return None
            _min, _max = min(vals), max(vals)
            _pad = (_max - _min) * 0.05 if _max > _min else (abs(_max) * 0.05 or 1)
            return (_min - _pad, _max + _pad)

        _total_domain = _domain_from(_total_vals)
        _shared_domain = _domain_from(_shared_vals)

        # '전체' 대비 비중(%) — 26년, 선택된 주차 범위의 합계 기준으로 계산해서 타이틀에 붙인다
        _total_26_sum = None
        for _label0, _cs0, _sm0 in _chart_series:
            if _label0 == "전체":
                _s26_total = _sm0.get("26년")
                if _s26_total is not None and not _s26_total.empty:
                    _s26_total_clip = _s26_total[(_s26_total.index >= wk_range[0]) & (_s26_total.index <= wk_range[1])]
                    if not _s26_total_clip.empty:
                        _total_26_sum = float(_s26_total_clip.sum())
                break

        # 2단계: '전체'는 자기 domain, 나머지 3개는 공유 domain으로 그린다
        _cols = st.columns(4)
        for i, (label, color_scheme, _series_map) in enumerate(_chart_series):
            with _cols[i]:
                _y_domain = _total_domain if label == "전체" else _shared_domain
                # "전체 실적" 페이지의 "전체"(Total) 차트처럼, label과 section_title이
                # 겹쳐서 "전체 전체 트래픽"같이 중복되는 경우 label을 또 붙이지 않는다.
                _title = section_title if section_title.startswith(label) else f"{label} {section_title}"
                if label != "전체" and _total_26_sum:
                    _s26 = _series_map.get("26년")
                    if _s26 is not None and not _s26.empty:
                        _s26_clip = _s26[(_s26.index >= wk_range[0]) & (_s26.index <= wk_range[1])]
                        if not _s26_clip.empty:
                            _share = float(_s26_clip.sum()) / _total_26_sum * 100
                            _title += f" (전체 중 {_share:.0f}%)"
                _render_weekly_yearly_chart(
                    _title, _series_map, wk_range, color_scheme,
                    y_domain=_y_domain, week_labels=_wk7_labels,
                )

    _render_section(f"{traffic_segment} 트래픽", "트래픽", traffic_segment, show_ff=True)

    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
    st.markdown("---")
    _render_section(f"{traffic_segment} 거래액", "거래액", traffic_segment, show_ff=True)

    # 규칙 기반 자동 인사이트 — 최신 주차 기준, "주별" 단위로 compute_bpu_comparison_rows를
    # 새로 호출해서(위 차트들과는 별개 계산이지만 같은 함수라 기준이 항상 일치) 만든다.
    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
    st.markdown("---")
    try:
        # 회원/신규 페이지는 df_traffic을 그 세그먼트로 필터링하고 '전체'로 재라벨링해서
        # 넘긴다 — compute_bpu_comparison_rows의 핵심 지표(거래액/트래픽/CR/객단가)들이
        # member='전체'로 고정돼 있어서, 그냥 넘기면 세그먼트와 무관하게 항상 '전체' 값만
        # 계산되는 버그가 있었음(회원/신규 페이지 인사이트가 서로 똑같이 나오던 원인).
        if traffic_segment == "전체":
            _wk_seg_df = df_traffic
        else:
            _wk_seg_df = df_traffic[df_traffic["회원구분"] == traffic_segment].copy()
            _wk_seg_df["회원구분"] = "전체"
        _wk_bpu_rows, _wk_bpu_cfg, _ = compute_bpu_comparison_rows(_wk_seg_df, "주별", df_traffic["날짜"].max())
        _wk_rb_sections = generate_rule_based_insights(_wk_bpu_rows, _wk_bpu_cfg)
        render_insight_panel(_wk_rb_sections)
    except Exception:
        pass  # 인사이트는 보조 기능이라, 계산 중 문제가 있어도 위 차트는 그대로 보여준다


if side["page"].startswith("6."):
    _render_weekly_segment_page("전체 실적", "전체")

if side["page"].startswith("7."):
    _render_weekly_segment_page("회원 실적", "회원")

if side["page"].startswith("8."):
    _render_weekly_segment_page("신규 실적", "신규")


# ============================================================
# 페이지 10: 종합 요약 — 여기저기 흩어진 핵심 지표를 한 화면에 모음
# (A: 핵심 지표 한눈에 + C: 카테고리×브랜드 피벗 테이블)
# ============================================================
if side["page"].startswith("3."):
    if df_traffic.empty:
        st.info("데이터가 없습니다. 사이드바에서 데이터를 업로드해주세요.")
    else:
        _sum_ref = sum_selected_period_date
        _sum_segment = sum_segment
        # unit은 위 스티키 헤더에서 이미 사이드바 전역값 그대로 쓰고 있어서 별도 지정 불필요

        # ── 섹션 1: 핵심 지표 (전체 BPU) ──
        st.markdown("#### 📊 핵심 지표 (전체)")
        _sum_bpu_rows, _sum_cfg, _ = compute_bpu_comparison_rows(df_traffic, unit, _sum_ref)
        _sum_total_metrics = {r["metric_label"]: r for r in _sum_bpu_rows if r["bpu"] == "Total" and r["stats"]}
        _sum_metric_order = ["EP UV", "거래액(순결제)", "구매객수", "구매전환율(%)", "객단가"]
        _sum_kpi_cols = st.columns(len(_sum_metric_order))
        for _i, _mlabel in enumerate(_sum_metric_order):
            with _sum_kpi_cols[_i]:
                _r = _sum_total_metrics.get(_mlabel)
                if _r is None:
                    st.markdown(f"<div style='color:#9ca3af;font-size:0.8rem;'>{_mlabel}: -</div>", unsafe_allow_html=True)
                    continue
                _stats = _r["stats"]
                _is_pct = _r["is_pct"]
                _val_str = f"{_stats['current']:.1f}%" if _is_pct else f"{_stats['current']:,.0f}"
                st.markdown(
                    f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:12px 14px;min-height:118px;'>"
                    f"<div style='color:#6b7280;font-size:0.76rem;'>{_mlabel}</div>"
                    f"<div style='font-size:1.15rem;font-weight:700;color:#111827;'>{_val_str}</div>"
                    f"<div style='font-size:0.72rem;margin-top:4px;'>"
                    f"{_sum_cfg['prev_label']} {format_delta_html(_stats.get('prev_delta'))}<br/>"
                    f"{_sum_cfg['yoy_label']} {format_delta_html(_stats.get('yoy_delta'))}"
                    f"</div></div>", unsafe_allow_html=True,
                )

        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

        # ── 섹션 2: BPU별 요약 실적 (전체 지표 x 전체 BPU 표, HTML 테이블로 증감률 색상 적용) ──
        st.markdown("#### 🏢 사업부(BPU)별 요약 실적")
        _sum_bpu_order = ["Total", "e-영업1", "e-영업2", "e-영업3", "e-영업4", "자사", "입점"]
        _sum_bpu_display = {"Total": "전체"}
        _sum_header_html = "<th>지표</th>" + "".join(
            f"<th>{_sum_bpu_display.get(b, b)}</th>" for b in _sum_bpu_order
        )
        _sum_body_rows = []
        for _mlabel in _sum_metric_order:
            _cells = [f"<td class='m'>{_mlabel}</td>"]
            for _bpu in _sum_bpu_order:
                _match = [r for r in _sum_bpu_rows if r["metric_label"] == _mlabel and r["bpu"] == _bpu]
                _r = _match[0] if _match else None
                if _r is None or _r["stats"] is None:
                    _cells.append("<td class='v'>-</td>")
                    continue
                _stats = _r["stats"]
                _is_pct = _r["is_pct"]
                _val_str = f"{_stats['current']:.1f}%" if _is_pct else f"{_stats['current']:,.0f}"
                _yoy_d = _stats.get("yoy_delta")
                _delta_html = format_delta_html(_yoy_d) if _yoy_d is not None else ""
                _cells.append(
                    f"<td class='v'>{_val_str}<br/><span style='font-size:0.78em;font-weight:400;'>{_delta_html}</span></td>"
                )
            _sum_body_rows.append(f"<tr>{''.join(_cells)}</tr>")
        st.markdown(
            f"<table class='summary-table'><thead><tr>{_sum_header_html}</tr></thead>"
            f"<tbody>{''.join(_sum_body_rows)}</tbody></table>",
            unsafe_allow_html=True,
        )
        st.caption(f"괄호 없이 아래 작은 글씨가 {_sum_cfg['yoy_label']}. 5개 지표 x 7개 BPU 구분 전체를 한 번에 봐요 — 상세 필터는 1번 페이지에서.")

        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        st.markdown("---")

        # ── 섹션 3: 카테고리 하이라이트 (거래액 전년비 최대상승/하락, BPU별로 구분) ──
        st.markdown(f"#### 📈 카테고리 하이라이트 (거래액 전년비, BPU별 · {_sum_segment})")
        _sum_category_movers = []
        if df_category.empty:
            st.info("카테고리 데이터가 없습니다.")
        else:
            _hl_bpu_cols = st.columns(4)
            for _i, _hbpu in enumerate(["e-영업1", "e-영업2", "e-영업3", "e-영업4"]):
                with _hl_bpu_cols[_i]:
                    st.markdown(f"<div style='font-weight:600;font-size:0.85rem;margin-bottom:4px;'>{_hbpu}</div>", unsafe_allow_html=True)
                    _cat_base_b = df_category[
                        (df_category["BPU"] == _hbpu) & (df_category["카테고리"] != "전체") & (df_category["브랜드"] == "전체")
                        & (df_category.get("회원구분", _sum_segment) == _sum_segment if "회원구분" in df_category.columns else True)
                    ]
                    if _cat_base_b.empty:
                        st.caption("데이터 없음")
                        continue
                    _cat_daily_b = _cat_base_b.groupby(["날짜", "카테고리"], as_index=False)["거래액"].sum()
                    _movers_b = []
                    for _cat_name, _g in _cat_daily_b.groupby("카테고리"):
                        _s = _g.set_index("날짜")["거래액"].sort_index()
                        _series = _s.resample(UNIT_CONFIG[unit]["rule"]).agg("mean")
                        if unit == "주별":
                            _series.index = _series.index - pd.Timedelta(days=6)
                        _series = _series[_series.index <= _sum_ref]
                        _raw = _s[_s.index <= raw_cutoff_date(_sum_ref, unit)]
                        _stats = compute_kpi_deltas(_series, unit, raw_daily=_raw)
                        if _stats is None or not _stats["current"] or _stats.get("yoy_delta") is None:
                            continue
                        _movers_b.append({
                            "카테고리": _cat_name, "거래액": _stats["current"], "전년비": _stats["yoy_delta"],
                            "작년거래액": _stats.get("yoy_value"),
                        })
                    if len(_movers_b) < 2:
                        st.caption("전년비 계산 가능한 카테고리 부족")
                        continue
                    _movers_b.sort(key=lambda r: r["전년비"], reverse=True)
                    # 상위 2개(상승) + 하위 2개(하락) — 핏플랍 종료 영향으로 슈즈가 당분간
                    # 계속 '최대 하락' 1위일 수 있어서, 그 다음(2위) 하락 카테고리도 봐야
                    # 실제 이상 신호를 놓치지 않는다.
                    _n_each = min(2, len(_movers_b) // 2) if len(_movers_b) >= 4 else 1
                    _tops = _movers_b[:_n_each]
                    _bottoms = _movers_b[-_n_each:][::-1] if _n_each > 0 else []  # 1위 하락부터 순서대로
                    _cards_html = ""
                    for _rank, _r in enumerate(_tops, start=1):
                        _tag = "최대 상승" if _rank == 1 else f"상승 {_rank}위"
                        _cards_html += (
                            f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:8px 10px;margin-bottom:6px;'>"
                            f"<div style='font-size:0.68rem;color:#16a34a;font-weight:600;'>{_tag}</div>"
                            f"<div style='font-size:0.85rem;font-weight:700;'>{_r['카테고리']}</div>"
                            f"<div style='font-size:0.72rem;'>{_r['거래액']:,.0f} · {format_delta_html(_r['전년비'])}</div>"
                            f"</div>"
                        )
                    for _rank, _r in enumerate(_bottoms, start=1):
                        _tag = "최대 하락" if _rank == 1 else f"하락 {_rank}위"
                        _cards_html += (
                            f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:8px 10px;margin-bottom:6px;'>"
                            f"<div style='font-size:0.68rem;color:#dc2626;font-weight:600;'>{_tag}</div>"
                            f"<div style='font-size:0.85rem;font-weight:700;'>{_r['카테고리']}</div>"
                            f"<div style='font-size:0.72rem;'>{_r['거래액']:,.0f} · {format_delta_html(_r['전년비'])}</div>"
                            f"</div>"
                        )
                    st.markdown(_cards_html, unsafe_allow_html=True)
                    _sum_category_movers.append((_hbpu, _tops, _bottoms, _movers_b))

        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        st.markdown("---")

        # 규칙 기반 자동 인사이트 — 이미 계산된 _sum_bpu_rows/_sum_cfg/카테고리 하이라이트를
        # 그대로 재사용해서 화면 숫자와 100% 일치하는 요약을 만든다(LLM 호출 없음).
        _sum_rb_sections = generate_rule_based_insights(_sum_bpu_rows, _sum_cfg, category_movers=_sum_category_movers)
        render_insight_panel(_sum_rb_sections)

        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        st.markdown("---")

        # ── 섹션 4: 월별 실적 비교 (26년 | 전년비 | 25년) — Total/정상/이월/입점 구분별로
        # 각각 표를 만든다(render_monthly_comparison_table 재사용, 2번 페이지와 공통). ──
        st.markdown(f"#### 📆 월별 실적 비교 (26년 vs 25년 · {_sum_segment})")
        if df_traffic.empty:
            st.info("데이터가 없습니다.")
        else:
            _mc_bpu_defs = [("Total", ["e-영업1", "e-영업2", "e-영업3", "e-영업4"]), ("정상", ["e-영업1"]), ("이월", ["e-영업2"]), ("입점", ["e-영업3", "e-영업4"])]
            for _mc_bpu_label, _mc_bpu_list in _mc_bpu_defs:
                _mc_base = df_traffic[(df_traffic["BPU"].isin(_mc_bpu_list)) & (df_traffic["회원구분"] == _sum_segment)]
                if _mc_base.empty:  # 세그먼트별 행이 없는 소스면 전체로 폴백
                    _mc_base = df_traffic[df_traffic["BPU"].isin(_mc_bpu_list)]
                if len(_mc_bpu_list) > 1:  # Total처럼 여러 BPU를 합쳐야 하면 날짜별로 먼저 합산
                    _mc_base = _mc_base.groupby("날짜", as_index=False)[["거래액", "트래픽", "구매객수"]].sum()
                render_monthly_comparison_table(_mc_base, _mc_bpu_label)
                st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        st.markdown("---")

        # ── 섹션 5: 쿠폰 비용 스냅샷 ──
        st.markdown("#### 🎟️ 쿠폰 비용 (이번 달)")
        if df_coupon_daily.empty:
            st.info("쿠폰 데이터가 없습니다.")
        else:
            _cp_month_start = _sum_ref.replace(day=1)
            _cp_this_month = df_coupon_daily[
                (df_coupon_daily["날짜"] >= _cp_month_start) & (df_coupon_daily["날짜"] <= _sum_ref)
            ]
            _cp_total = _cp_this_month["쿠폰할인"].sum()
            _cp_gmv_month = df_traffic[
                (df_traffic["BPU"] == "Total") & (df_traffic["회원구분"] == "전체")
                & (df_traffic["날짜"] >= _cp_month_start) & (df_traffic["날짜"] <= _sum_ref)
            ]["거래액"].sum()
            _cp_rate = (_cp_total / _cp_gmv_month * 100) if _cp_gmv_month else None
            _cp_cols = st.columns(3)
            with _cp_cols[0]:
                st.markdown(
                    f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 12px;'>"
                    f"<div style='font-size:0.76rem;color:#6b7280;'>쿠폰할인(이번달 누계)</div>"
                    f"<div style='font-size:1.05rem;font-weight:700;'>{_cp_total:,.0f}</div></div>",
                    unsafe_allow_html=True,
                )
            with _cp_cols[1]:
                st.markdown(
                    f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 12px;'>"
                    f"<div style='font-size:0.76rem;color:#6b7280;'>거래액(이번달 누계)</div>"
                    f"<div style='font-size:1.05rem;font-weight:700;'>{_cp_gmv_month:,.0f}</div></div>",
                    unsafe_allow_html=True,
                )
            with _cp_cols[2]:
                _rate_str = f"{_cp_rate:.2f}%" if _cp_rate is not None else "-"
                st.markdown(
                    f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 12px;'>"
                    f"<div style='font-size:0.76rem;color:#6b7280;'>비용률</div>"
                    f"<div style='font-size:1.05rem;font-weight:700;color:#7c3aed;'>{_rate_str}</div></div>",
                    unsafe_allow_html=True,
                )
            st.caption("자세한 쿠폰 분석은 9번 페이지에서 확인하세요.")


# ============================================================
# 페이지 11: 마감 예상 실적 — 진행 중인 달을 일할계산으로 마감까지 추정해서
# Total/자사/정상/이월/입점 x 1~12월 표로 보여준다.
# ============================================================
if side["page"].startswith("10."):
    st.markdown("### 📈 마감 예상 실적")
    st.caption(
        "진행 중인(아직 안 끝난) 달은 지금까지 실적을 경과일수로 나눠 이번 달 전체 일수만큼 "
        "곱한 '일할계산' 방식으로 마감 시점 예상치를 추정해요. 지나간 달은 실제 확정값 그대로예요. "
        "비율 지표(CR/객단가)는 분자/분모를 각각 예상한 뒤 나눠서 계산해요(단순 평균 아님)."
    )

    if df_traffic.empty:
        st.info("데이터가 없습니다. 사이드바에서 ep_traffic.csv를 업로드해주세요.")
    else:
        _fc_abs_last = df_traffic["날짜"].max()
        _fc_cur_month_start = _fc_abs_last.replace(day=1)
        _fc_cur_month_end = (_fc_cur_month_start + pd.offsets.MonthBegin(1)) - pd.Timedelta(days=1)
        _fc_is_partial_month = (_fc_abs_last.year == forecast_year) and (_fc_cur_month_end > _fc_abs_last) and (_fc_cur_month_start <= _fc_abs_last)
        _fc_cur_month_num = _fc_abs_last.month if _fc_abs_last.year == forecast_year else None

        if _fc_is_partial_month:
            _fc_days_elapsed = df_traffic[
                (df_traffic["날짜"] >= _fc_cur_month_start) & (df_traffic["날짜"] <= _fc_abs_last)
            ]["날짜"].nunique()
            st.caption(
                f"📅 마감예상 대상: {_fc_cur_month_num}월 (지금까지 {_fc_days_elapsed}일 실적 → "
                f"{_fc_cur_month_end.day}일 기준으로 환산)"
            )

        _fc_excel_bytes = build_forecast_excel(df_traffic, df_coupon_daily, df_ep, forecast_year, _fc_cur_month_num)
        st.download_button(
            "⬇️ 엑셀 다운로드 (마감예상 전체)",
            data=_fc_excel_bytes,
            file_name=f"마감예상실적_{forecast_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        def _style_forecast_table(df, is_pct=False, is_currency_like=False):
            """마감예상 달 컬럼은 배경을 살짝 강조하고, 0(아직 안 온 달)은 빈칸(-)으로 표시."""
            _month_cols = [c for c in df.columns if c != "합계"]

            def _fmt(v):
                if v is None or pd.isna(v):
                    return "-"
                if v == 0:
                    return "-"
                if is_pct:
                    return f"{v:.1f}%"
                return f"{v:,.0f}"

            def _style_func(data):
                out = pd.DataFrame("", index=data.index, columns=data.columns)
                if _fc_is_partial_month and f"{_fc_cur_month_num}월" in out.columns:
                    out[f"{_fc_cur_month_num}월"] = "background-color: #fef3c7; font-weight: 600;"
                return out

            return df.style.format(_fmt).apply(_style_func, axis=None)

        _fc_metric_defs = [
            ("거래액", "거래액", None, False, False),
            ("트래픽", "트래픽", None, False, False),
            ("구매객수", "구매객수", None, False, False),
            ("객단가", "거래액", "구매객수", True, False),
            ("구매전환율(CR)", "구매객수", "트래픽", True, True),
        ]

        for _label, _num_col, _den_col, _is_ratio, _is_pct in _fc_metric_defs:
            st.markdown(f"#### {_label}")
            _tbl = build_forecast_table(
                df_traffic, _label, _num_col, _den_col, forecast_year,
                is_ratio=_is_ratio, ratio_scale=100 if _is_pct else 1.0,
            )
            _styled = _style_forecast_table(_tbl, is_pct=_is_pct)
            st.dataframe(_styled, use_container_width=True)
            st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)

        # --- 전년비 (마감예상 달 vs 작년 같은 달 실제값) ---
        if _fc_is_partial_month:
            st.markdown("---")
            st.markdown(f"#### 📊 {_fc_cur_month_num}월 마감예상 vs 작년 {_fc_cur_month_num}월")
            _fc_yoy_rows = []
            for _label, _num_col, _den_col, _is_ratio, _is_pct in _fc_metric_defs:
                _cur_tbl = build_forecast_table(
                    df_traffic, _label, _num_col, _den_col, forecast_year,
                    is_ratio=_is_ratio, ratio_scale=100 if _is_pct else 1.0,
                )
                _prev_tbl = build_forecast_table(
                    df_traffic, _label, _num_col, _den_col, forecast_year - 1,
                    is_ratio=_is_ratio, ratio_scale=100 if _is_pct else 1.0,
                )
                _col_name = f"{_fc_cur_month_num}월"
                for _row_name in FORECAST_BPU_ROWS.keys():
                    _cur_v = _cur_tbl.loc[_row_name, _col_name] if _col_name in _cur_tbl.columns else None
                    _prev_v = _prev_tbl.loc[_row_name, _col_name] if _col_name in _prev_tbl.columns else None
                    _yoy = pct_delta_safe(_cur_v, _prev_v) if (_cur_v is not None and _prev_v) else None
                    _fc_yoy_rows.append({
                        "지표": _label, "구분": _row_name,
                        "마감예상": _cur_v, "작년실적": _prev_v, "전년비": _yoy,
                    })
            _fc_yoy_df = pd.DataFrame(_fc_yoy_rows)

            def _fmt_yoy_val(v):
                return "-" if v is None or pd.isna(v) or v == 0 else f"{v:,.1f}"

            def _fmt_yoy_pct(v):
                return format_delta_text(v) if v is not None else "-"

            _fc_yoy_styled = _fc_yoy_df.style.format(
                {"마감예상": _fmt_yoy_val, "작년실적": _fmt_yoy_val, "전년비": _fmt_yoy_pct}
            ).map(
                lambda v: ("color:#16a34a;" if v >= 0 else "color:#dc2626;") if isinstance(v, (int, float)) and pd.notna(v) else "",
                subset=["전년비"],
            )
            st.dataframe(_fc_yoy_styled, use_container_width=True, height=400)

            # 규칙 기반 자동 인사이트 — 거래액/Total 행을 재사용해서 forecast_stats로 전달.
            try:
                _fc_total_gmv = next(
                    (r for r in _fc_yoy_rows if r["지표"] == "거래액" and r["구분"] == "Total"), None
                )
                if _fc_total_gmv and _fc_total_gmv.get("전년비") is not None:
                    _fc_rb = generate_rule_based_insights(
                        [], {},
                        forecast_stats={"yoy": _fc_total_gmv["전년비"], "month": f"{_fc_cur_month_num}월"},
                    )
                    render_insight_panel(_fc_rb)
            except Exception:
                pass

        # --- 비용/비용률 (쿠폰 데이터, df_coupon_daily) ---
        if not df_coupon_daily.empty:
            st.markdown("#### 비용 (쿠폰할인)")
            _tbl_cost = build_forecast_table(df_coupon_daily, "비용", "쿠폰할인", None, forecast_year)
            st.dataframe(_style_forecast_table(_tbl_cost), use_container_width=True)
            st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)

            # 비용률 = 비용(쿠폰할인) ÷ 거래액 — 분자(쿠폰)와 분모(거래액)가 서로 다른
            # 데이터소스(df_coupon_daily / df_traffic)라 build_forecast_table 하나로는 못
            # 만들고, 각각 예상한 뒤 행 단위로 나눠서 직접 합친다(단순 평균 아님 원칙 유지).
            st.markdown("#### 비용률")
            _rate_rows = []
            for _row_name, _bpu_list in FORECAST_BPU_ROWS.items():
                _cost_nums, _ = compute_monthly_forecast_series(df_coupon_daily, "쿠폰할인", None, forecast_year, _bpu_list)
                _gmv_nums, _ = compute_monthly_forecast_series(df_traffic, "거래액", None, forecast_year, _bpu_list)
                _row = {"구분": _row_name}
                for _m in range(1, 13):
                    _c, _g = _cost_nums[_m - 1], _gmv_nums[_m - 1]
                    _row[f"{_m}월"] = (_c / _g * 100) if _g else (0.0 if _c == 0 else None)
                _tot_c, _tot_g = sum(_cost_nums), sum(_gmv_nums)
                _row["합계"] = (_tot_c / _tot_g * 100) if _tot_g else 0.0
                _rate_rows.append(_row)
            _tbl_rate = pd.DataFrame(_rate_rows).set_index("구분")
            st.dataframe(_style_forecast_table(_tbl_rate, is_pct=True), use_container_width=True)
            st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)
        else:
            st.info("쿠폰 데이터가 없어서 비용/비용률은 건너뛰었어요 (사이드바에서 ep_coupon_daily.csv를 업로드하면 나와요).")

        # --- 전시상품수 (EP채널 데이터, df_ep) ---
        # df_ep는 원부매칭여부/최저가여부라는 추가 필터 축이 있어서, '전체(Total/Total)'
        # 기준으로 먼저 좁혀야 df_traffic과 같은 방식(BPU만 보는)으로 다룰 수 있다.
        _ep_total_scope = df_ep[(df_ep[COL_MATCH] == "Total") & (df_ep[COL_LOWEST] == "Total")]
        if not _ep_total_scope.empty:
            st.markdown("#### 전시상품수")
            _ep_scope_renamed = _ep_total_scope.rename(columns={COL_DATE: "날짜", COL_BPU: "BPU"})
            _tbl_disp = build_forecast_table(_ep_scope_renamed, "전시상품수", "평균 EP 전시 상품수", None, forecast_year)
            st.dataframe(_style_forecast_table(_tbl_disp), use_container_width=True)
            st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)
        else:
            st.info("EP채널 데이터가 없어서 전시상품수는 건너뛰었어요 (사이드바에서 ep_data_long.csv를 업로드하면 나와요).")
