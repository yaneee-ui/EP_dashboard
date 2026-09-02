"""EP 대시보드의 계산/렌더링 헬퍼 함수 모음.

app.py가 너무 커져서(5600줄+) 페이지 스크립트와 안 섞이는 순수 함수들만
분리했다 — 데이터 로딩(df_traffic 등)이나 사이드바 상태처럼 스크립트
실행 시점에만 존재하는 전역은 여기서 참조하지 않고, 전부 파라미터로
넘겨받는 함수들만 모았다.
"""
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import io

from utils import (
    COL_DATE, COL_BPU, COL_MATCH, COL_LOWEST, UNIT_CONFIG,
    compute_kpi_deltas, raw_cutoff_date, effective_month_of_week, week_of_month,
    format_delta_html, pct_delta_safe,
)
from ai_insights import generate_ranking_insights, render_ranking_insight_box


def _ref_str(val, is_pct=False):
    """비교 대상 실제 값을 괄호로 표시."""
    if val is None or pd.isna(val):
        return ""
    if is_pct:
        return f" <span style='color:#9ca3af'>({val:.1f}%)</span>"
    return f" <span style='color:#9ca3af'>({val:,.0f})</span>"


def _status_entry(df, date_col):
    if df is None or df.empty or date_col not in df.columns:
        return None, None, None
    _dmin = df[date_col].min()
    _dmax = df[date_col].max()
    _n = df[date_col].nunique()
    return _dmin.strftime("%Y-%m-%d"), _dmax.strftime("%Y-%m-%d"), _n


BPU_GROUPS = {
    "자사": ["e-영업1", "e-영업2"],
    "입점": ["e-영업3", "e-영업4"],
}


def _reset_date_range(_state_key, _value):
    """'최근으로' 버튼용 콜백. on_click으로 넘겨서 위젯이 다시 그려지기 전에
    session_state를 먼저 갱신한다 (버튼 클릭 처리 중 위젯이 이미 인스턴스화된
    뒤에 st.session_state[key]=... 를 직접 대입하면 StreamlitAPIException이 남)."""
    st.session_state[_state_key] = _value


def render_week_range_filter(series, key_prefix, col_slider, col_reset, default_weeks=12):
    """'주별' 조회 시 차트 영역에 'N월 N주차' 라벨 기반 주차 범위 슬라이더를 그린다.
    (7/8/9번 페이지에서 쓰던 것과 동일한 라벨 방식을 1/2번 페이지 차트에도 적용)
    '일별'의 날짜 range picker와 같은 자리에, 같은 3분할 레이아웃(슬라이더|최근으로|전년비교선)을
    유지하기 위해 컬럼을 호출부에서 미리 만들어 넘겨받는다(컬럼 안에 컬럼을 또 만들면 레이아웃이
    불균형해지는 문제가 있어서).
    series: 이미 '주별'로 리샘플된(Monday 라벨) 시리즈. 반환값: 선택 범위로 잘라진 시리즈."""
    if series.empty:
        return series
    labels = [f"{effective_month_of_week(d).month}월 {week_of_month(d)}주차" for d in series.index]
    n = len(labels)
    default_start_idx = max(0, n - default_weeks)
    _range_key = f"{key_prefix}_wkrange"
    with col_slider:
        st.markdown("<div style='font-size:0.78rem;color:#6b7280;margin-bottom:1px;'>주차 범위</div>", unsafe_allow_html=True)
        sel_labels = st.select_slider(
            "주차 범위", options=labels, value=(labels[default_start_idx], labels[-1]),
            label_visibility="collapsed", key=_range_key,
        )
    with col_reset:
        st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
        st.button(
            "🔄 최근으로", key=f"{key_prefix}_wkrange_reset", use_container_width=True,
            on_click=_reset_date_range, args=(_range_key, (labels[default_start_idx], labels[-1])),
        )
    start_idx = labels.index(sel_labels[0])
    end_idx = labels.index(sel_labels[1])
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    return series.iloc[start_idx:end_idx + 1]


def aggregate_traffic(df, bpus, member="전체"):
    """여러 BPU의 트래픽 데이터를 합산. CR/객단가는 재계산."""
    sub = df[(df["BPU"].isin(bpus)) & (df["회원구분"] == member)]
    if sub.empty:
        return sub
    agg = sub.groupby("날짜").agg({
        "트래픽": "sum", "거래액": "sum", "구매객수": "sum",
    }).reset_index()
    agg["CR"] = (agg["구매객수"] / agg["트래픽"] * 100).where(agg["트래픽"] > 0, 0)
    agg["객단가"] = (agg["거래액"] / agg["구매객수"]).where(agg["구매객수"] > 0, 0)
    agg["BPU"] = "+".join(bpus)
    agg["회원구분"] = member
    return agg


FF_BRAND_CODE = "FF"


def exclude_ff_brand(df):
    """핏플랍(FF, 2025-10월 종료) 브랜드 실적을 카테고리 집계에서 제외한다.

    - '브랜드=전체' 집계행(FF가 속한 카테고리 + '카테고리=전체' 전체 집계행)에서
      FF의 트래픽/거래액/구매객수를 빼고 CR/객단가를 재계산 (비율 지표는 그냥
      빼면 안 되고 항상 분자/분모를 먼저 뺀 뒤 재계산해야 함 - 데이터 정합성 원칙).
    - 개별 브랜드 목록(랭킹 등)에서는 FF 행 자체를 제외.
    - 다른 카테고리/브랜드 행은 전혀 건드리지 않음 (FF는 슈즈에만 있었으므로).
    """
    if df.empty or "브랜드" not in df.columns:
        return df

    ff_rows = df[(df["브랜드"] == FF_BRAND_CODE) & (df["카테고리"] != "전체")]
    df2 = df[df["브랜드"] != FF_BRAND_CODE].copy()  # FF 자체 행(전체/개별 카테고리 다 포함) 제외
    if ff_rows.empty:
        return df2

    group_keys = ["날짜", "BPU"] + (["회원구분"] if "회원구분" in df.columns else [])
    ff_agg = ff_rows.groupby(group_keys, as_index=False)[["트래픽", "거래액", "구매객수"]].sum()
    ff_agg = ff_agg.rename(columns={"트래픽": "_ff_t", "거래액": "_ff_g", "구매객수": "_ff_b"})

    ff_categories = set(ff_rows["카테고리"].unique()) | {"전체"}
    mask = (df2["브랜드"] == "전체") & (df2["카테고리"].isin(ff_categories))
    if not mask.any():
        return df2

    target = df2.loc[mask].merge(ff_agg, on=group_keys, how="left")
    target[["_ff_t", "_ff_g", "_ff_b"]] = target[["_ff_t", "_ff_g", "_ff_b"]].fillna(0)
    target["트래픽"] = target["트래픽"] - target["_ff_t"]
    target["거래액"] = target["거래액"] - target["_ff_g"]
    target["구매객수"] = target["구매객수"] - target["_ff_b"]
    target["CR"] = (target["구매객수"] / target["트래픽"] * 100).where(target["트래픽"] > 0)
    target["객단가"] = (target["거래액"] / target["구매객수"]).where(target["구매객수"] > 0)

    df2.loc[mask, ["트래픽", "거래액", "구매객수", "CR", "객단가"]] = target[
        ["트래픽", "거래액", "구매객수", "CR", "객단가"]
    ].values
    return df2


def exclude_ff_from_traffic(df_traffic, df_category):
    """EP실적 원본(df_traffic)엔 브랜드 정보가 없어서 핏플랍을 직접 제외할 수 없으므로,
    카테고리 원본(df_category)에서 FF의 일자별 트래픽/거래액/구매객수를 가져와 그만큼
    df_traffic에서 빼고 CR/객단가를 재계산한다.

    - 'e-영업1'/'e-영업2' 리터럴 행(FF가 실제로 존재하는 BPU)에서 직접 차감.
    - 'Total' 리터럴 행은 e-영업1~4를 미리 합쳐둔 별도 행이라, e-영업1+e-영업2 합계만큼
      추가로 차감해야 함 (자사/입점 뷰는 앱에서 e-영업1~4를 그때그때 합산하는 구조라
      e-영업1/e-영업2를 고치면 자동으로 반영되지만, Total은 그렇지 않음).
    - 세그먼트는 카테고리 원본에 남아있는 전체/회원/신규만 보정 가능하다(비회원/기존은
      용량 문제로 카테고리 원본에서 뺐던 세그먼트라 원천적으로 보정 불가 — 그대로 둠)."""
    if df_traffic.empty or df_category.empty:
        return df_traffic

    ff = df_category[(df_category["브랜드"] == "FF") & (df_category["카테고리"] != "전체")]
    if ff.empty:
        return df_traffic

    ff_by_bpu = ff.groupby(["날짜", "BPU", "회원구분"], as_index=False)[["트래픽", "거래액", "구매객수"]].sum()
    ff_total = ff_by_bpu.groupby(["날짜", "회원구분"], as_index=False)[["트래픽", "거래액", "구매객수"]].sum()
    ff_total["BPU"] = "Total"
    ff_all = pd.concat([ff_by_bpu, ff_total], ignore_index=True)
    ff_all = ff_all.rename(columns={"트래픽": "_ff_t", "거래액": "_ff_g", "구매객수": "_ff_b"})

    group_keys = ["날짜", "BPU", "회원구분"]
    mask = df_traffic["BPU"].isin(["Total", "e-영업1", "e-영업2"])
    if not mask.any():
        return df_traffic

    target = df_traffic.loc[mask].merge(ff_all, on=group_keys, how="left")
    target[["_ff_t", "_ff_g", "_ff_b"]] = target[["_ff_t", "_ff_g", "_ff_b"]].fillna(0)
    target["트래픽"] = target["트래픽"] - target["_ff_t"]
    target["거래액"] = target["거래액"] - target["_ff_g"]
    target["구매객수"] = target["구매객수"] - target["_ff_b"]
    target["CR"] = (target["구매객수"] / target["트래픽"] * 100).where(target["트래픽"] > 0)
    target["객단가"] = (target["거래액"] / target["구매객수"]).where(target["구매객수"] > 0)

    df2 = df_traffic.copy()
    df2.loc[mask, ["트래픽", "거래액", "구매객수", "CR", "객단가"]] = target[
        ["트래픽", "거래액", "구매객수", "CR", "객단가"]
    ].values
    return df2


def _weekly_of_year(daily_df, metric, year):
    """daily_df(날짜 컬럼 포함)를 해당 연도만 잘라서 월~일 주간 평균으로 리샘플하고,
    그 해 첫 주=1주차로 시작하는 정수 인덱스를 붙인다. (7번 페이지 전용, 스티키 헤더의
    주차 필터와 페이지 본문 둘 다에서 같은 정의를 써야 해서 모듈 레벨에 둠)"""
    d = daily_df[daily_df["날짜"].dt.year == year]
    if d.empty:
        return pd.Series(dtype="float64")
    s = d.set_index("날짜")[metric].sort_index()
    weekly = s.resample("W-SUN").mean()
    weekly.index = weekly.index - pd.Timedelta(days=6)  # 그 주의 월요일로 라벨
    weekly = weekly.reset_index(drop=True)
    weekly.index = weekly.index + 1  # 1주차부터 시작
    return weekly


def _week_labels_for_year(daily_df, metric, year):
    """_weekly_of_year와 정확히 같은 순서로 'N월 N주차' 라벨을 만든다 (1:1 대응 보장).
    그 해 1/1이 월요일이 아니면 첫 주는 실제로 전년도 12월 마지막 주에 걸치는데,
    이건 utils.week_of_month의 '월 경계는 이전 달 마지막 주로 본다' 규칙과 동일하게 처리
    (그래야 예: 1월 1주차가 두 번 나오는 라벨 중복이 안 생김). 그 경우엔 연도를 앞에 붙여
    구분한다(예: '24년 12월 5주차')."""
    d = daily_df[daily_df["날짜"].dt.year == year]
    if d.empty:
        return []
    s = d.set_index("날짜")[metric].sort_index()
    weekly_raw = s.resample("W-SUN").mean()
    labels = []
    for bucket_end in weekly_raw.index:
        mon = bucket_end - pd.Timedelta(days=6)
        eff = effective_month_of_week(mon)
        if eff.year != year:
            labels.append(f"{eff.year % 100}년 {eff.month}월 {week_of_month(mon)}주차")
        else:
            labels.append(f"{eff.month}월 {week_of_month(mon)}주차")
    return labels


def _correct_partial_week_yoy(s_target_year, s_ref_year, raw_daily_target, raw_daily_ref, target_year, ref_year):
    """s_target_year(예: 26년)의 마지막 주차가 아직 진행 중(부분)이면, s_ref_year(예: 25년)의
    같은 주차 값도 '같은 요일 위치까지만' 반영하도록 재계산해서 돌려준다.

    문제였던 것: _weekly_of_year가 각 연도를 독립적으로 리샘플(평균)하다 보니, 올해
    최신 주차가 아직 2일치(예: 월·화)만 있어도 작년 같은 주차는 이미 다 지나서 7일
    전체 평균이 그대로 나온다 — '올해 2일'을 '작년 7일'과 비교하는 불공정한 비교였음.
    이 함수는 올해 마지막 주차에 실제로 존재하는 요일 위치(월=0~일=6)를 확인해서,
    작년 같은 주차에서도 그 요일들에 해당하는 값만 뽑아 평균 내서 덮어쓴다.
    (완성된 주차라면 그대로 반환 — 보정이 필요 없음)"""
    if s_target_year.empty or raw_daily_target.empty or s_ref_year.empty:
        return s_ref_year
    last_week_num = s_target_year.index[-1]
    if last_week_num not in s_ref_year.index:
        return s_ref_year

    _jan1_target = pd.Timestamp(f"{target_year}-01-01")
    _mon0_target = _jan1_target - pd.Timedelta(days=_jan1_target.weekday())
    _week_start_target = _mon0_target + pd.Timedelta(weeks=last_week_num - 1)
    _week_end_target = _week_start_target + pd.Timedelta(days=6)

    _actual_days_target = raw_daily_target[
        (raw_daily_target.index >= _week_start_target) & (raw_daily_target.index <= _week_end_target)
    ]
    if _actual_days_target.empty or len(_actual_days_target) >= 7:
        return s_ref_year  # 데이터가 없거나 이미 7일 다 있으면(완성된 주) 보정 필요 없음

    _weekday_positions = sorted(set(d.weekday() for d in _actual_days_target.index))

    _jan1_ref = pd.Timestamp(f"{ref_year}-01-01")
    _mon0_ref = _jan1_ref - pd.Timedelta(days=_jan1_ref.weekday())
    _week_start_ref = _mon0_ref + pd.Timedelta(weeks=last_week_num - 1)
    _matching_ref_dates = [_week_start_ref + pd.Timedelta(days=wd) for wd in _weekday_positions]

    _matched_vals = raw_daily_ref.reindex(_matching_ref_dates).dropna()
    if _matched_vals.empty:
        return s_ref_year

    s_ref_year = s_ref_year.copy()
    s_ref_year.loc[last_week_num] = _matched_vals.mean()
    return s_ref_year


def aggregate_ep(df, bpus, match_status, lowest_status):
    """여러 BPU의 EP채널 데이터를 합산. 비율 지표는 재계산."""
    sub = df[(df[COL_BPU].isin(bpus)) & (df[COL_MATCH] == match_status) & (df[COL_LOWEST] == lowest_status)]
    if sub.empty:
        return sub
    # 합산 가능한 컬럼(수량) vs 재계산이 필요한 컬럼(비율) 분리
    sum_cols = [c for c in sub.columns if c not in [COL_DATE, COL_BPU, COL_MATCH, COL_LOWEST,
                "원부매칭율(%)", "최저가율(%)", "구매전환율(%)", "첫구매거래액(%)", "신규가입율", "첫구매 전환율(%)"]]
    agg = sub.groupby(COL_DATE)[sum_cols].sum().reset_index()
    # 비율 재계산
    if "평균 원부매칭 상품수" in agg.columns and "평균 EP 전시 상품수" in agg.columns:
        agg["원부매칭율(%)"] = (agg["평균 원부매칭 상품수"] / agg["평균 EP 전시 상품수"] * 100).where(agg["평균 EP 전시 상품수"] > 0, 0)
    if "평균 최저가 상품수" in agg.columns and "평균 EP 전시 상품수" in agg.columns:
        agg["최저가율(%)"] = (agg["평균 최저가 상품수"] / agg["평균 EP 전시 상품수"] * 100).where(agg["평균 EP 전시 상품수"] > 0, 0)
    if "평균 EP 고객수(총결제)" in agg.columns and "평균 EP UV" in agg.columns:
        agg["구매전환율(%)"] = (agg["평균 EP 고객수(총결제)"] / agg["평균 EP UV"] * 100).where(agg["평균 EP UV"] > 0, 0)
    if "평균 EP 첫구매 거래액(총결제)" in agg.columns and "평균 EP 거래액(총결제)" in agg.columns:
        agg["첫구매거래액(%)"] = (agg["평균 EP 첫구매 거래액(총결제)"] / agg["평균 EP 거래액(총결제)"] * 100).where(agg["평균 EP 거래액(총결제)"] > 0, 0)
    if "평균 EP 신규가입수" in agg.columns and "평균 EP UV" in agg.columns:
        agg["신규가입율"] = (agg["평균 EP 신규가입수"] / agg["평균 EP UV"] * 100).where(agg["평균 EP UV"] > 0, 0)
    if "평균 EP 첫구매 고객수(총결제)" in agg.columns and "평균 EP UV" in agg.columns:
        agg["첫구매 전환율(%)"] = (agg["평균 EP 첫구매 고객수(총결제)"] / agg["평균 EP UV"] * 100).where(agg["평균 EP UV"] > 0, 0)
    agg[COL_BPU] = "+".join(bpus)
    agg[COL_MATCH] = match_status
    agg[COL_LOWEST] = lowest_status
    return agg


def build_weekly_report_excel(unit, selected_period_date, df_traffic, df_category, cat_segment, ff_exclude):
    """주간보고용 엑셀(bytes) 생성.
    - 시트1 'BPU별': compute_bpu_comparison_rows() 그대로 사용 → 1번 페이지 비교표와 100% 동일 로직/숫자
    - 시트2 '카테고리별(거래액)': compute_category_yoy_rows() 그대로 사용(BPU=e-영업1~4 각각)
      → 2번 페이지 '카테고리별 요약'표와 동일한 세그먼트 필터·핏플랍 제외 규칙 적용
    별도로 값을 재계산하지 않고 두 페이지가 쓰는 함수를 그대로 호출하므로, 화면에 보이는 표와
    엑셀 숫자가 어긋날 일이 없다.

    (예전엔 여기서 df_traffic/df_category 중 더 짧은 쪽 마지막 날짜로 기준시점을 강제로
    맞추는 로직이 있었는데, 그건 "혹시 두 파일 마지막 날짜가 다를 수도 있다"는 가정 하의
    방어 코드였고 실제로 그게 원인이었던 적은 없었음. 오히려 이 로직 때문에 6번 페이지가
    1·2번 페이지랑 다른 기준일자를 쓰게 돼서 숫자가 어긋나는 새 문제가 생겨서 제거함 —
    이제 selected_period_date를 그대로 써서 1·2번 페이지와 완전히 동일한 기준을 쓴다.)"""
    cur_year = pd.Timestamp(selected_period_date).year
    prev_year = cur_year - 1
    col_prev, col_cur = f"{prev_year}년", f"{cur_year}년"

    # --- 시트1: BPU별 (1번 페이지와 동일 함수) — 이미지 기준: 전체/e-영업1~4만 ---
    _bpu_rows, cfg, BPU_COLS = compute_bpu_comparison_rows(df_traffic, unit, selected_period_date)
    _bpu_keep = {"Total", "e-영업1", "e-영업2", "e-영업3", "e-영업4"}
    _bpu_label = {"Total": "전체"}
    left_rows = []
    # 객단가는 거래액/구매객수 값이 나온 뒤에 그 두 값으로 다시 계산해야 해서(비율 지표는
    # 날짜별 단순평균 금지 — 이 표에 있는 거래액/구매객수 숫자와 정확히 일치하도록),
    # BPU별로 거래액·구매객수 값을 따로 모아둔다.
    _gmv_by_bpu, _cnt_by_bpu = {}, {}
    for r in _bpu_rows:
        if r["bpu"] not in _bpu_keep or r["stats"] is None:
            continue
        if r["metric_label"] == "거래액(순결제)":
            _gmv_by_bpu[r["bpu"]] = r["stats"]
        elif r["metric_label"] == "구매객수":
            _cnt_by_bpu[r["bpu"]] = r["stats"]

    for r in _bpu_rows:
        if r["bpu"] not in _bpu_keep:
            continue
        stats = r["stats"]
        if stats is None:
            continue
        is_pct = r["is_pct"]
        _cur_val, _yoy_val = stats["current"], stats.get("yoy_value")
        if r["metric_label"] == "객단가":
            _gmv, _cnt = _gmv_by_bpu.get(r["bpu"]), _cnt_by_bpu.get(r["bpu"])
            if _gmv and _cnt:
                _cur_val = _gmv["current"] / _cnt["current"] if _cnt["current"] else None
                _yoy_val = (
                    _gmv.get("yoy_value") / _cnt["yoy_value"]
                    if _cnt.get("yoy_value") else None
                )
        _yoy_delta = pct_delta_safe(_cur_val, _yoy_val) if (_cur_val is not None and _yoy_val) else stats.get("yoy_delta")
        # CR(구매전환율)은 지금 4.8처럼 '이미 100배 된 숫자'로 저장돼 있는데, 엑셀에서
        # 그냥 숫자 옆에 %를 글자로 붙이는 대신 — 0.048처럼 소수로 저장하고 셀 서식을
        # 퍼센트(0.0%)로 지정하면 엑셀이 알아서 "4.8%"로 보여준다(진짜 엑셀 percent
        # 타입이라 다른 셀에서 참조 계산해도 정상 작동). 아래에서 cell.number_format으로
        # 적용한다.
        _cur_store = (_cur_val / 100) if (is_pct and _cur_val is not None) else _cur_val
        _yoy_store = (_yoy_val / 100) if (is_pct and _yoy_val is not None) else _yoy_val
        left_rows.append({
            "지표": r["metric_label"], "구분": _bpu_label.get(r["bpu"], r["bpu"]),
            col_prev: round(_yoy_store, 3) if is_pct else (round(_yoy_store) if _yoy_store is not None else None),
            col_cur: round(_cur_store, 3) if is_pct else (round(_cur_store) if _cur_store is not None else None),
            "전년비(%)": round(_yoy_delta, 1) if _yoy_delta is not None else None,
        })
    left_df = pd.DataFrame(left_rows)

    # --- 시트2: 카테고리별 (거래액, e-영업1~4 각각) — 2번 페이지와 동일 함수 ---
    right_rows = []
    for bv in ["e-영업1", "e-영업2", "e-영업3", "e-영업4"]:
        for r in compute_category_yoy_rows(df_category, bv, cat_segment, ff_exclude, unit, selected_period_date):
            right_rows.append({
                "BPU": bv, "카테고리": r["카테고리"],
                col_prev: round(r["yoy_value"]) if r.get("yoy_value") is not None else None,
                col_cur: round(r["current"]),
                "전년비(%)": round(r["yoy_delta"], 1) if r.get("yoy_delta") is not None else None,
            })
    right_df = pd.DataFrame(right_rows)

    # --- 시트3: 카테고리별 (트래픽, e-영업1~4 각각) — 시트2와 동일 로직, 지표만 트래픽 ---
    right_rows_traffic = []
    for bv in ["e-영업1", "e-영업2", "e-영업3", "e-영업4"]:
        for r in compute_category_yoy_rows(df_category, bv, cat_segment, ff_exclude, unit, selected_period_date, metric_col="트래픽"):
            right_rows_traffic.append({
                "BPU": bv, "카테고리": r["카테고리"],
                col_prev: round(r["yoy_value"]) if r.get("yoy_value") is not None else None,
                col_cur: round(r["current"]),
                "전년비(%)": round(r["yoy_delta"], 1) if r.get("yoy_delta") is not None else None,
            })
    right_df_traffic = pd.DataFrame(right_rows_traffic)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        left_df.to_excel(writer, index=False, sheet_name="BPU별")
        right_df.to_excel(writer, index=False, sheet_name="카테고리별(거래액)")
        right_df_traffic.to_excel(writer, index=False, sheet_name="카테고리별(트래픽)")

        # CR(구매전환율) 행의 값 셀에 엑셀 퍼센트 서식을 입혀서, 저장해둔 소수(0.048)가
        # 화면엔 "4.8%"로 보이게 한다.
        _ws_bpu = writer.sheets["BPU별"]
        _pct_col_letters = [
            _ws_bpu.cell(row=1, column=ci + 1).column_letter
            for ci, cname in enumerate(left_df.columns) if cname in (col_prev, col_cur)
        ]
        for _ridx, _lbl in enumerate(left_df["지표"], start=2):  # 엑셀 행은 1=헤더, 2부터 데이터
            if _lbl == "구매전환율(%)":
                for _cl in _pct_col_letters:
                    _ws_bpu[f"{_cl}{_ridx}"].number_format = "0.0%"

        # 전년비(%) 컬럼: 증가=초록, 감소=[빨강]△(마이너스 기호 대신). 저장된 값은 이미
        # "-11.0"처럼 퍼센트 숫자라(소수 아님) 서식에서 %를 그냥 곱하면(0.0%) 100배 더
        # 커져버리니, 리터럴 문자 "%"를 따옴표로 감싸서 곱하기 없이 그대로 붙인다.
        _YOY_DELTA_FMT = '[Green]0.0"%";[Red]"△"0.0"%"'
        for _sheet_name, _df in [("BPU별", left_df), ("카테고리별(거래액)", right_df), ("카테고리별(트래픽)", right_df_traffic)]:
            _ws = writer.sheets[_sheet_name]
            if "전년비(%)" not in _df.columns:
                continue
            _delta_col_idx = list(_df.columns).index("전년비(%)") + 1
            _delta_col_letter = _ws.cell(row=1, column=_delta_col_idx).column_letter
            for _ridx in range(2, len(_df) + 2):
                _ws[f"{_delta_col_letter}{_ridx}"].number_format = _YOY_DELTA_FMT

        for _ws in writer.sheets.values():
            for _col in _ws.columns:
                _w = max((len(str(c.value)) for c in _col if c.value is not None), default=8)
                _ws.column_dimensions[_col[0].column_letter].width = min(_w + 3, 40)
    return buf.getvalue(), left_df, right_df, right_df_traffic


def build_forecast_excel(df_traffic, df_coupon_daily, df_ep, forecast_year, fc_cur_month_num):
    """마감 예상 실적을 사용자가 보여준 양식(구분 | 1~12월 | 전년비)대로 엑셀 한 시트에
    지표별 섹션을 쌓아서 만든다. 진행 중인(마감예상) 달 컬럼은 노란 배경으로 강조하고,
    전년비는 증가=초록/감소=빨강으로 색칠한다."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb_buf = io.BytesIO()
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "마감예상"

    HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
    FORECAST_FILL = PatternFill("solid", fgColor="FFF2CC")
    SECTION_FONT = Font(bold=True)
    UP_FONT = Font(color="16A34A")
    DOWN_FONT = Font(color="DC2626")

    _ep_scope = df_ep[(df_ep[COL_MATCH] == "Total") & (df_ep[COL_LOWEST] == "Total")].rename(
        columns={COL_DATE: "날짜", COL_BPU: "BPU"}
    )
    _sections = [
        ("거래액", df_traffic, "거래액", None, False, False),
        ("비용", df_coupon_daily, "쿠폰할인", None, False, False),
        ("트래픽", df_traffic, "트래픽", None, False, False),
        ("구매객수", df_traffic, "구매객수", None, False, False),
        ("객단가", df_traffic, "거래액", "구매객수", True, False),
        ("구매전환율(CR)", df_traffic, "구매객수", "트래픽", True, True),
        ("전시상품수", _ep_scope, "평균 EP 전시 상품수", None, False, False),
    ]

    row_idx = 1
    for label, src_df, num_col, den_col, is_ratio, is_pct in _sections:
        ws.cell(row=row_idx, column=1, value=label).font = SECTION_FONT
        row_idx += 1
        header = ["구분"] + [f"{m}월" for m in range(1, 13)] + ["합계"]
        if fc_cur_month_num:
            header.append(f"전년비({fc_cur_month_num}월)")
            header.append(f"작년({fc_cur_month_num}월)")
        for c, h in enumerate(header, start=1):
            cell = ws.cell(row=row_idx, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
        header_row = row_idx
        row_idx += 1

        if src_df is None or src_df.empty:
            ws.cell(row=row_idx, column=1, value="(데이터 없음)")
            row_idx += 2
            continue

        _cur_tbl = build_forecast_table(src_df, label, num_col, den_col, forecast_year, is_ratio=is_ratio, ratio_scale=100 if is_pct else 1.0)
        _prev_tbl = None
        if fc_cur_month_num:
            _prev_tbl = build_forecast_table(src_df, label, num_col, den_col, forecast_year - 1, is_ratio=is_ratio, ratio_scale=100 if is_pct else 1.0)

        for row_name in FORECAST_BPU_ROWS.keys():
            ws.cell(row=row_idx, column=1, value=row_name)
            for m in range(1, 13):
                col = m + 1
                v = _cur_tbl.loc[row_name, f"{m}월"]
                if v is None or pd.isna(v) or v == 0:
                    cell = ws.cell(row=row_idx, column=col, value=None)
                elif is_pct:
                    cell = ws.cell(row=row_idx, column=col, value=round(float(v) / 100, 4))
                    cell.number_format = "0.0%"
                else:
                    cell = ws.cell(row=row_idx, column=col, value=round(float(v)))
                if fc_cur_month_num and m == fc_cur_month_num:
                    cell.fill = FORECAST_FILL
            _tot_v = _cur_tbl.loc[row_name, "합계"]
            if _tot_v is None or pd.isna(_tot_v):
                ws.cell(row=row_idx, column=14, value=None)
            elif is_pct:
                _tot_cell = ws.cell(row=row_idx, column=14, value=round(float(_tot_v) / 100, 4))
                _tot_cell.number_format = "0.0%"
            else:
                ws.cell(row=row_idx, column=14, value=round(float(_tot_v)))
            if fc_cur_month_num and _prev_tbl is not None:
                _cv = _cur_tbl.loc[row_name, f"{fc_cur_month_num}월"]
                _pv = _prev_tbl.loc[row_name, f"{fc_cur_month_num}월"]
                _yoy = pct_delta_safe(_cv, _pv) if (_cv is not None and _pv) else None
                _cell = ws.cell(row=row_idx, column=15, value=None if _yoy is None else round(_yoy / 100, 4))
                if _yoy is not None:
                    _cell.font = UP_FONT if _yoy >= 0 else DOWN_FONT
                    _cell.number_format = '+0.0%;-0.0%'
                # 작년 실제값(전년비 계산에 쓴 기준 숫자) — 비교하기 편하게 바로 옆에 표시
                if _pv is None or pd.isna(_pv):
                    ws.cell(row=row_idx, column=16, value=None)
                elif is_pct:
                    _pv_cell = ws.cell(row=row_idx, column=16, value=round(float(_pv) / 100, 4))
                    _pv_cell.number_format = "0.0%"
                else:
                    ws.cell(row=row_idx, column=16, value=round(float(_pv)))
            row_idx += 1
        row_idx += 1  # 섹션 사이 빈 줄

    ws.column_dimensions["A"].width = 12
    for c in range(2, 17):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = "B2"

    wb.save(wb_buf)
    return wb_buf.getvalue()


def render_excel_download(df_export, filename, label="⬇️ 엑셀 다운로드"):
    """DataFrame을 엑셀 파일로 변환해 다운로드 버튼으로 제공."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="누적 데이터")
    st.download_button(
        label,
        data=buf.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )


def _truncate_by_range(series, cum_start, cum_end, cum_unit):
    """누적 데이터의 시작일/종료일 필터를 적용.
    월별/월마감은 리샘플 라벨이 '그 달의 말일'이라, 종료일이 월중(예: 7/27)이면
    라벨(7/31)이 종료일보다 커서 그 달 전체가 잘못 걸러지는 문제가 있었음.
    → 월 단위(연-월)로 비교해 종료월이 포함되면 살리도록 수정.
    """
    if series.empty:
        return series
    start_ts = pd.Timestamp(cum_start)
    end_ts = pd.Timestamp(cum_end)
    if cum_unit in ("월별", "월마감"):
        mask = (series.index.to_period("M") >= start_ts.to_period("M")) & \
               (series.index.to_period("M") <= end_ts.to_period("M"))
    else:
        mask = (series.index >= start_ts) & (series.index <= end_ts)
    return series[mask]


def compute_bpu_comparison_rows(df_traffic, unit="일별", selected_period_date=None):
    """render_bpu_comparison_table과 동일한 로직(집계 방식·월마감 규칙·부분기간 보정)으로
    지표×BPU 조합별 compute_kpi_deltas 결과를 계산해 구조화된 리스트로 반환한다.
    HTML 렌더링(render_bpu_comparison_table)과 엑셀 내보내기가 이 함수를 공유해서
    두 군데 숫자가 어긋나지 않게 한다."""
    BPU_COLS = ["Total", "e-영업1", "e-영업2", "e-영업3", "e-영업4", "자사", "입점"]
    METRICS = [
        ("트래픽", "전체", "EP UV", False),
        ("거래액", "전체", "거래액(순결제)", False),
        ("구매객수", "전체", "구매객수", False),
        ("CR", "전체", "구매전환율(%)", True),
        ("객단가", "전체", "객단가", False),
        ("트래픽", "회원", "회원UV", False),
        ("거래액", "회원", "회원거래액", False),
        ("트래픽", "신규", "신규UV", False),
        ("거래액", "신규", "신규거래액", False),
    ]
    SUMMABLE_METRICS = {"트래픽", "거래액", "구매객수"}
    cfg = UNIT_CONFIG[unit]

    def _series_for(bpu_key, metric, member="전체"):
        if bpu_key in BPU_GROUPS:
            sub = aggregate_traffic(df_traffic, BPU_GROUPS[bpu_key], member)
        else:
            sub = df_traffic[(df_traffic["BPU"] == bpu_key) & (df_traffic["회원구분"] == member)]
        if sub.empty:
            return pd.Series(dtype="float64"), pd.Series(dtype="float64")
        s = sub.set_index("날짜")[metric].sort_index()
        _agg = "sum" if (unit == "월마감" and metric in SUMMABLE_METRICS) else "mean"
        series = s.resample(cfg["rule"]).agg(_agg)
        if unit == "주별":
            series.index = series.index - pd.Timedelta(days=6)
        elif unit == "월마감":
            if not series.empty and s.index.max() < series.index[-1]:
                series = series.iloc[:-1]
        _s_raw = s
        if selected_period_date is not None and not series.empty:
            series = series[series.index <= selected_period_date]
            _s_raw = s[s.index <= raw_cutoff_date(selected_period_date, unit)]
        return series, _s_raw

    rows = []
    for metric_key, member, metric_label, is_pct in METRICS:
        for bpu_key in BPU_COLS:
            series, _s_raw = _series_for(bpu_key, metric_key, member)
            stats = compute_kpi_deltas(series, unit, raw_daily=_s_raw)
            rows.append({"metric_label": metric_label, "is_pct": is_pct, "bpu": bpu_key, "stats": stats})
    return rows, cfg, BPU_COLS


def render_bpu_comparison_table(df_traffic, unit="일별", selected_period_date=None):
    """사업부별(Total/e-영업1~4/자사/입점) 실적 비교표.
    상단의 '조회 단위'(일별/주별/월별/월마감)와 '기준 시점'을 그대로 따른다.
    → 위 KPI 카드/요약표와 동일한 집계 방식이므로 Total 열은 KPI 카드 값과 일치한다."""
    rows, cfg, BPU_COLS = compute_bpu_comparison_rows(df_traffic, unit, selected_period_date)

    by_metric = {}
    for r in rows:
        by_metric.setdefault(r["metric_label"], {"is_pct": r["is_pct"], "cells": {}})
        by_metric[r["metric_label"]]["cells"][r["bpu"]] = r["stats"]

    for metric_label, info in by_metric.items():
        is_pct = info["is_pct"]
        header_html = "<th>구분</th>" + "".join(f"<th>{b}</th>" for b in BPU_COLS)
        row_val, row_prev, row_avg, row_yoy = [], [], [], []
        for bpu_key in BPU_COLS:
            stats = info["cells"].get(bpu_key)
            if stats is None:
                row_val.append("<td>-</td>")
                row_prev.append("<td>-</td>")
                row_avg.append("<td>-</td>")
                row_yoy.append("<td>-</td>")
                continue
            val_str = f"{stats['current']:.1f}%" if is_pct else f"{stats['current']:,.0f}"
            row_val.append(f"<td class='v'>{val_str}</td>")
            row_prev.append(f"<td class='d'>{format_delta_html(stats['prev_delta'])}{_ref_str(stats.get('prev_value'), is_pct)}</td>")
            row_avg.append(f"<td class='d'>{format_delta_html(stats['avg_delta'])}{_ref_str(stats.get('avg_value'), is_pct)}</td>")
            row_yoy.append(f"<td class='d'>{format_delta_html(stats['yoy_delta'])}{_ref_str(stats.get('yoy_value'), is_pct)}</td>")

        cfg = UNIT_CONFIG[unit]
        table_html = (
            "<table class='summary-table' style='margin-bottom:14px;'>"
            f"<thead><tr><th colspan='{len(BPU_COLS)+1}' style='background:#eef2ff;color:#374151;font-weight:700;'>{metric_label}</th></tr>"
            f"<tr>{header_html}</tr></thead>"
            "<tbody>"
            f"<tr><td class='m'>값</td>{''.join(row_val)}</tr>"
            f"<tr><td class='m'>{cfg['prev_label']}</td>{''.join(row_prev)}</tr>"
            f"<tr><td class='m'>{cfg['avg_label']}</td>{''.join(row_avg)}</tr>"
            f"<tr><td class='m'>{cfg['yoy_label']}</td>{''.join(row_yoy)}</tr>"
            "</tbody></table>"
        )
        st.markdown(table_html, unsafe_allow_html=True)


def compute_category_yoy_rows(df_category, bpu_value, cat_segment, ff_exclude, unit, selected_period_date, metric_col="거래액"):
    """2번 페이지 '카테고리별 요약'표와 완전히 동일한 규칙(세그먼트 필터·핏플랍 제외·
    compute_kpi_deltas)으로 특정 BPU 하나의 카테고리별 metric_col(기본 거래액) stats를 계산한다.
    '전체' 카테고리 행도 포함해서 맨 위에 오도록 반환."""
    d = df_category[df_category["BPU"] == bpu_value]
    if d.empty:
        return []
    if "회원구분" in d.columns:
        d = d[d["회원구분"] == cat_segment]
    if ff_exclude:
        d = exclude_ff_brand(d)
    if metric_col not in d.columns:
        raise ValueError(
            f"ep_category.csv에 '{metric_col}' 컬럼이 없어요. 컨버터에서 만든 파일이 맞는지, "
            "혹은 오래된 버전의 컨버터로 만든 파일은 아닌지 확인해주세요."
        )
    d = d[d["브랜드"] == "전체"][["날짜", "카테고리", metric_col]]
    if d.empty:
        return []

    cfg = UNIT_CONFIG[unit]
    _agg = "sum" if unit == "월마감" else "mean"
    rows = []
    for cat_name, g in d.groupby("카테고리"):
        s = g.set_index("날짜")[metric_col].sort_index()
        series = s.resample(cfg["rule"]).agg(_agg)
        if unit == "주별":
            series.index = series.index - pd.Timedelta(days=6)
        elif unit == "월마감" and not series.empty and s.index.max() < series.index[-1]:
            series = series.iloc[:-1]
        if not series.empty:
            series = series[series.index <= selected_period_date]
        _s_raw = s[s.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s
        stats = compute_kpi_deltas(series, unit, raw_daily=_s_raw)
        if stats is None or not stats["current"]:
            continue
        rows.append({"카테고리": cat_name, **stats})

    _head = [r for r in rows if r["카테고리"] == "전체"]
    _rest = sorted([r for r in rows if r["카테고리"] != "전체"], key=lambda r: r["current"], reverse=True)
    return _head + _rest


DASHBOARD_EVENTS = [
    (pd.Timestamp("2026-08-04"), "최저가쿠폰 초기화(18시)"),
    (pd.Timestamp("2026-08-05"), "다나와 기준 쿠폰 배치"),
]


def render_line_chart(chart_df, height=350, unit="일별", yoy_actual_dates=None):
    """줌/팬이 비활성화된 라인 차트 (마커 + 호버 툴팁 포함).
    st.line_chart는 마우스 휠 확대/축소가 기본 활성화돼 스크롤 시 화면이 튀므로,
    Altair로 직접 그려 인터랙션을 끈다. (첫 컬럼=금년 진한 파랑, 둘째=전년 하늘색)
    unit이 '월별'/'월마감'이면 x축을 월 단위(연-월)로 표시한다.
    yoy_actual_dates: chart_df.index와 같은 길이의 날짜 배열/시리즈. '전년' 계열은 화면상
    올해 날짜 위치에 겹쳐 그려지지만(비교하기 좋으라고) 실제 값은 작년 것이므로, 이걸 주면
    '(전년)'이 포함된 계열의 툴팁 날짜만 실제 작년 날짜로 보여준다(안 주면 x축 날짜 그대로).
    일별(비월별) 조회일 땐, DASHBOARD_EVENTS 중 화면에 보이는 날짜 범위에 해당하는 이벤트를
    빨간 점선 세로줄 + 라벨로 같이 표시한다."""
    import altair as alt

    if chart_df is None or chart_df.empty:
        st.info("표시할 데이터가 없습니다.")
        return

    cols = list(chart_df.columns)
    colors = ["#2563eb", "#7dd3fc"][: len(cols)]

    _df = chart_df.copy()
    _df.index.name = "날짜"

    # 금년(cols[0]) 라인 호버 시 전년비(%)도 같이 보이게 — 같은 날짜의 전년 비교 컬럼
    # (cols[1], "(전년)"이 붙은 컬럼)과 비교해서 미리 계산해둔다. melt 대상 컬럼(cols)에
    # 섞이지 않도록, 계산은 여기서 다 끝내고 결과(딕셔너리)만 남긴다.
    _yoy_col = cols[1] if len(cols) > 1 else None
    _yoy_map = {}
    if _yoy_col is not None:
        # 전년 데이터가 통째로 없으면(브랜드/카테고리 조합에 작년 실적이 아예 없는 경우)
        # 이 컬럼이 전부 None이 되면서 dtype이 object로 잡혀서 .abs() 호출 시
        # "TypeError: bad operand type for abs(): 'NoneType'"가 났음 — 숫자로 강제
        # 변환(coerce)해서 방지한다(변환 안 되는 값은 NaN이 되고, NaN끼리 연산하면
        # 결과도 NaN이라 이후 dropna 등에서 자연스럽게 걸러짐).
        _yoy_numeric = pd.to_numeric(_df[_yoy_col], errors="coerce")
        _cur_numeric = pd.to_numeric(_df[cols[0]], errors="coerce")
        _yoy_pct = (_cur_numeric - _yoy_numeric) / _yoy_numeric.abs() * 100
        _yoy_map = _yoy_pct.to_dict()

    long_df = _df[cols].reset_index().melt("날짜", var_name="구분", value_name="값")
    long_df = long_df.dropna(subset=["값"])

    if _yoy_map:
        long_df["전년비"] = long_df["날짜"].map(_yoy_map)
        long_df["전년비_표시"] = long_df.apply(
            lambda r: (f"{'' if r['전년비'] >= 0 else '△'}{abs(r['전년비']):.1f}%"
                       if r["구분"] == cols[0] and pd.notna(r["전년비"]) else "-"),
            axis=1,
        )
    else:
        long_df["전년비_표시"] = "-"

    # 범례 문구를 간소화한다 — 파랑(cols[0])은 항상 "N년"(데이터의 최신 연도, 동적 계산이라
    # 해가 바뀌어도 하드코딩 없이 자동으로 맞음), 하늘(cols[1])은 조회단위별로 미리 정한
    # 짧은 문구로. 원래 컬럼명("거래액"/"전년동요일비(전년)" 등)은 이제 legend/tooltip에
    # 안 보이고, 대신 아래 매핑된 라벨만 쓴다.
    _YOY_LEGEND_LABEL = {
        "일별": "전년 동요일", "주별": "전년 동일주차",
        "월별": "전년 동월(동요일 기준)", "월마감": "전년 동월",
    }
    _cur_year_label = f"{int(pd.DatetimeIndex(chart_df.index).year.max()) % 100}년"
    _yoy_display_label = _YOY_LEGEND_LABEL.get(unit, "전년")
    _label_map = {cols[0]: _cur_year_label}
    if len(cols) > 1:
        _label_map[cols[1]] = _yoy_display_label
    long_df["구분"] = long_df["구분"].map(lambda c: _label_map.get(c, c))
    cols = [_label_map.get(c, c) for c in cols]  # 아래 색상 도메인도 같이 맞춘다

    # 전년 비교선의 툴팁 날짜를 실제 작년 날짜로 교체 (x축 위치는 올해 날짜 그대로 유지)
    long_df["_tooltip_date"] = long_df["날짜"]
    if yoy_actual_dates is not None:
        _actual_map = dict(zip(pd.DatetimeIndex(chart_df.index), pd.DatetimeIndex(yoy_actual_dates)))
        _is_yoy_row = long_df["구분"] == _yoy_display_label
        long_df.loc[_is_yoy_row, "_tooltip_date"] = long_df.loc[_is_yoy_row, "날짜"].map(_actual_map)

    _is_monthly = unit in ("월별", "월마감")
    _event_rows = []
    if _is_monthly:
        x_enc = alt.X(
            "날짜:T", title=None, timeUnit="yearmonth",
            axis=alt.Axis(format="%Y-%m", labelAngle=0),
        )
        _date_fmt = "%Y-%m"
    else:
        # temporal(T) 타입 x축은 Vega-Lite가 자체적으로 '보기 좋은' 틱 간격을 자동으로 골라서,
        # 며칠 안 되는 좁은 구간에서도 하루에 여러 개의 틱(포맷은 %m/%d라 같은 날짜로 중복 표시)이
        # 찍히는 문제가 있었음. 실제 데이터 포인트 날짜만 문자열 라벨(ordinal)로 써서 방지하고,
        # 점이 많으면(20개 초과) 라벨을 적당히 솎아서 겹치지 않게 한다.
        long_df["_date_label"] = long_df["날짜"].dt.strftime("%m/%d")
        _uniq_dates = sorted(long_df["날짜"].unique())
        _label_order = [pd.Timestamp(d).strftime("%m/%d") for d in _uniq_dates]
        _n = len(_label_order)
        if _n > 20:
            _step = -(-_n // 20)  # ceil(n/20)
            _tick_vals = _label_order[::_step]
            if _label_order[-1] not in _tick_vals:
                _tick_vals.append(_label_order[-1])
        else:
            _tick_vals = _label_order
        x_enc = alt.X(
            "_date_label:O", title=None, sort=_label_order,
            axis=alt.Axis(labelAngle=0, values=_tick_vals),
        )
        _date_fmt = "%Y-%m-%d"

        # 화면에 보이는 날짜 범위 안에 있는 이벤트만 세로 점선+라벨로 표시
        _event_rows = []
        for ev_date, ev_label in DASHBOARD_EVENTS:
            if ev_date.normalize() in set(pd.DatetimeIndex(_uniq_dates).normalize()):
                _event_rows.append({"_date_label": ev_date.strftime("%m/%d"), "이벤트": ev_label})

    chart = (
        alt.Chart(long_df)
        .mark_line(strokeWidth=3, point=alt.OverlayMarkDef(size=45, filled=True, opacity=1))
        .encode(
            x=x_enc,
            y=alt.Y("값:Q", title=None, axis=alt.Axis(format="~s"), scale=alt.Scale(zero=False)),
            color=alt.Color(
                "구분:N",
                scale=alt.Scale(domain=cols, range=colors),
                legend=alt.Legend(orient="bottom", title=None),
            ),
            tooltip=[
                alt.Tooltip("_tooltip_date:T", title="날짜", format=_date_fmt),
                alt.Tooltip("구분:N", title="구분"),
                alt.Tooltip("값:Q", title="값", format=",.0f"),
                alt.Tooltip("전년비_표시:N", title="전년비"),
            ],
        )
        .properties(height=height)
    )

    if _event_rows:
        _ev_df = pd.DataFrame(_event_rows)
        # 세로선 없이, 금년(cols[0]) 라인의 그 날짜 실제 값 위치에만 마커를 찍는다
        _main_vals = long_df[long_df["구분"] == cols[0]][["_date_label", "값"]]
        _ev_df = _ev_df.merge(_main_vals, on="_date_label", how="left").dropna(subset=["값"])
        if not _ev_df.empty:
            _ev_marker = alt.Chart(_ev_df).mark_point(
                shape="triangle-down", size=110, color="#dc2626", filled=True, opacity=0.95,
            ).encode(
                x=alt.X("_date_label:O", sort=_label_order),
                y=alt.Y("값:Q"),
                tooltip=[alt.Tooltip("_date_label:N", title="날짜"), alt.Tooltip("이벤트:N", title="이벤트")],
            )
            chart = alt.layer(chart, _ev_marker).resolve_scale(x="shared", y="shared")

    # .interactive()를 호출하지 않으므로 휠 확대/축소·드래그 팬이 비활성화됨
    st.altair_chart(chart, use_container_width=True)

    if _event_rows:
        _ev_caption = "  ·  ".join(f"📌 {r['_date_label']} {r['이벤트']}" for r in _event_rows)
        st.caption(_ev_caption)


FORECAST_BPU_ROWS = {"Total": None, "자사": BPU_GROUPS["자사"], "정상": ["e-영업1"], "이월": ["e-영업2"], "입점": BPU_GROUPS["입점"]}


def compute_monthly_forecast_series(df, num_col, den_col, year, bpu_list, segment="전체"):
    """1~12월 각각의 (분자합계, 분모합계)를 계산한다. 진행 중인(마지막) 달은 일할계산으로
    마감예상 처리 — 지금까지의 합계를 경과일수로 나눠 이번 달 전체 일수만큼 곱해서 추정.
    완성된(지나간) 달은 그대로 실제 합계, 아직 시작 안 한 달은 (0, 0).
    den_col이 None이면 절대값 지표(거래액/트래픽 등)라 분모 계산은 건너뛴다."""
    if bpu_list is None:
        # "Total"을 뜻하는 bpu_list=None인 경우: 데이터에 BPU="Total" 행이 실제로 있으면
        # 그걸 쓰고(df_traffic처럼), 없으면(쿠폰 데이터처럼 Total 집계행 자체가 없는 소스)
        # e-영업1~4를 다 더해서 Total을 만든다.
        if "Total" in df["BPU"].unique().tolist():
            sub = df[df["BPU"] == "Total"]
        else:
            sub = df[df["BPU"].isin(["e-영업1", "e-영업2", "e-영업3", "e-영업4"])]
    else:
        sub = df[df["BPU"].isin(bpu_list)]
    if "회원구분" in sub.columns and segment in sub["회원구분"].unique().tolist():
        sub = sub[sub["회원구분"] == segment]
    _abs_last = df["날짜"].max()
    nums, dens = [], []
    for m in range(1, 13):
        m_start = pd.Timestamp(year, m, 1)
        m_end = (m_start + pd.offsets.MonthBegin(1)) - pd.Timedelta(days=1)
        if m_start > _abs_last:
            nums.append(0.0)
            dens.append(0.0 if den_col else None)
            continue
        month_data = sub[(sub["날짜"] >= m_start) & (sub["날짜"] <= min(m_end, _abs_last))]
        days_elapsed = month_data["날짜"].nunique()
        if days_elapsed == 0:
            nums.append(0.0)
            dens.append(0.0 if den_col else None)
            continue
        days_in_month = m_end.day
        num_sum = month_data[num_col].sum()
        den_sum = month_data[den_col].sum() if den_col else None
        _is_partial = m_end > _abs_last  # 이번 달 마지막날이 아직 안 지났으면 진행 중
        if _is_partial:
            num_sum = num_sum / days_elapsed * days_in_month
            if den_sum is not None:
                den_sum = den_sum / days_elapsed * days_in_month
        nums.append(num_sum)
        dens.append(den_sum)
    return nums, dens


def build_forecast_table(df_traffic, metric_label, num_col, den_col, year, segment="전체", is_ratio=False, ratio_scale=1.0):
    """지표 하나(예: 거래액)에 대해 Total/자사/정상/이월/입점 5개 행 x 1~12월+합계 컬럼의
    DataFrame을 만든다. is_ratio=True면 분자/분모를 각각 예상한 뒤 나눠서 비율을 재계산한다
    (비율 지표는 절대 단순평균 금지 원칙 — 예상 CR = 예상 구매객수/예상 트래픽 이런 식으로)."""
    rows = []
    for row_name, bpu_list in FORECAST_BPU_ROWS.items():
        nums, dens = compute_monthly_forecast_series(df_traffic, num_col, den_col, year, bpu_list, segment)
        if is_ratio:
            vals = [
                (n / d * ratio_scale) if (d and d != 0) else (0.0 if n == 0 else None)
                for n, d in zip(nums, dens)
            ]
        else:
            vals = nums
        row = {"구분": row_name}
        for m, v in enumerate(vals, start=1):
            row[f"{m}월"] = v
        if is_ratio:
            _tot_num = sum(nums)
            _tot_den = sum(d for d in dens if d)
            row["합계"] = (_tot_num / _tot_den * ratio_scale) if _tot_den else 0.0
        else:
            row["합계"] = sum(nums)
        rows.append(row)
    df_out = pd.DataFrame(rows).set_index("구분")
    df_out.attrs["metric_label"] = metric_label
    return df_out


_DIGIT_HAS_BATCHIM = {"0": True, "1": False, "2": False, "3": True, "4": False, "5": False, "6": True, "7": True, "8": True, "9": True}


def _has_batchim(word):
    """단어 마지막 글자의 받침 유무를 판단한다 (한글이면 유니코드 계산, 숫자로 끝나면
    그 숫자의 한국어 발음 기준 — 'e-영업1'처럼 BPU 이름이 숫자로 끝나는 경우 대응)."""
    if not word:
        return False
    last = word[-1]
    if last.isdigit():
        return _DIGIT_HAS_BATCHIM.get(last, False)
    if "가" <= last <= "힣":
        return (ord(last) - ord("가")) % 28 != 0
    return False


def _emphasize(text):
    """인사이트 결론/액션 문구를 노란 배경으로 강조 표시."""
    return f"<span style='background:#fef3c7;padding:1px 5px;border-radius:4px;font-weight:600;'>{text}</span>"


def _josa_ga(word):
    """단어 뒤에 '이' 또는 '가'를 자동으로 붙인다."""
    return "이" if _has_batchim(word) else "가"


def _josa_eun(word):
    """단어 뒤에 '은' 또는 '는'을 자동으로 붙인다."""
    return "은" if _has_batchim(word) else "는"


def generate_rule_based_insights(bpu_rows, bpu_cfg, category_movers=None, coupon_stats=None, forecast_stats=None):
    """규칙 기반 자동 인사이트. 이미 계산된 stats(compute_bpu_comparison_rows 결과 등)만
    받아서 문장으로 조립한다 — LLM을 호출하지 않으므로 화면에 보이는 숫자와 항상 100%
    일치하고, 할루시네이션(엉뚱한 숫자를 말하는) 위험이 없다.
    각 섹션은 {"title": str, "body": str} 형태로 반환하고, 필요한 입력이 없으면(예:
    category_movers=None) 해당 섹션은 건너뛴다."""
    sections = []
    _prev_label = bpu_cfg.get("prev_label", "전기간비")
    _yoy_label = bpu_cfg.get("yoy_label", "전년비")

    def _find(metric_label, bpu="Total"):
        for r in bpu_rows:
            if r["metric_label"] == metric_label and r["bpu"] == bpu and r["stats"]:
                return r["stats"]
        return None

    def _fmt(v, is_pct=False):
        if v is None or pd.isna(v):
            return "-"
        return f"{v:.1f}%" if is_pct else f"{v:,.0f}"

    def _fmt_delta(v):
        return format_delta_html(v) if v is not None else "-"

    # 1) 기본 지표 해석
    _gmv = _find("거래액(순결제)")
    _uv = _find("EP UV")
    _cr = _find("구매전환율(%)")
    _aov = _find("객단가")
    if _gmv:
        _body = (
            f"거래액 {_fmt(_gmv['current'])} ({_prev_label} {_fmt_delta(_gmv.get('prev_delta'))}, "
            f"{_yoy_label} {_fmt_delta(_gmv.get('yoy_delta'))}) · "
            f"트래픽 {_fmt(_uv['current']) if _uv else '-'} · "
            f"CR {_fmt(_cr['current'], True) if _cr else '-'} · "
            f"객단가 {_fmt(_aov['current']) if _aov else '-'}"
        )
        sections.append({"title": "① 이번 기간 기본 지표는?", "body": _body})

    # 2) 자사 vs 입점 성장 비교
    _gmv_jasa = _find("거래액(순결제)", "자사")
    _gmv_ipjeom = _find("거래액(순결제)", "입점")
    if _gmv_jasa and _gmv_ipjeom:
        _jy, _iy = _gmv_jasa.get("yoy_delta"), _gmv_ipjeom.get("yoy_delta")
        if _jy is not None and _iy is not None:
            _leader = "자사" if _jy > _iy else "입점"
            _body = (
                f"자사 거래액 {_yoy_label} {_fmt_delta(_jy)}, 입점 거래액 {_yoy_label} {_fmt_delta(_iy)}.<br>"
                f"{_emphasize(f'{_leader}{_josa_ga(_leader)} 더 빠르게 성장하고 있어요.')}"
            )
            sections.append({"title": "② 자사·입점 중 어디가 더 크고 있나?", "body": _body})

    # 3) BPU별 병목/기회 — 거래액뿐 아니라 트래픽·CR도 같이 봐서, 부진이 '유입 감소'
    # 때문인지 '전환율 저하' 때문인지 원인까지 구분해서 보여준다.
    _bpu_names = ["e-영업1", "e-영업2", "e-영업3", "e-영업4"]
    _bpu_yoys = [(b, _find("거래액(순결제)", b)) for b in _bpu_names]
    _bpu_yoys = [(b, s.get("yoy_delta")) for b, s in _bpu_yoys if s and s.get("yoy_delta") is not None]
    if len(_bpu_yoys) >= 2:
        _bpu_yoys.sort(key=lambda x: x[1], reverse=True)
        _best_b, _best_v = _bpu_yoys[0]
        _worst_b, _worst_v = _bpu_yoys[-1]
        _worst_uv = _find("EP UV", _worst_b)
        _worst_cr = _find("구매전환율(%)", _worst_b)
        _cause_bits = []
        if _worst_uv and _worst_uv.get("yoy_delta") is not None:
            _cause_bits.append(f"트래픽 {_yoy_label} {_fmt_delta(_worst_uv['yoy_delta'])}")
        if _worst_cr and _worst_cr.get("yoy_delta") is not None:
            _cause_bits.append(f"CR {_yoy_label} {_fmt_delta(_worst_cr['yoy_delta'])}")
        _cause_txt = f" (원인 분해 — {' · '.join(_cause_bits)})" if _cause_bits else ""
        _body = (
            f"거래액 {_yoy_label} 기준 <b>{_best_b}</b>{_josa_ga(_best_b)} 가장 좋아요({_fmt_delta(_best_v)}).<br>"
            f"반대로 <b>{_worst_b}</b>{_josa_eun(_worst_b)} {_fmt_delta(_worst_v)}로 가장 부진해요{_cause_txt}"
            f" — {_emphasize('우선 점검이 필요해요.')}"
        )
        sections.append({"title": "③ 가장 큰 병목·기회는 어디?", "body": _body})

    # 4) 카테고리 하이라이트 — 비율(전년비 %) 기준 상승/하락에 더해, 절대액(거래액 실제
    # 감소분) 기준으로 영향도가 가장 큰 카테고리도 같이 본다. 비율만 보면 원래 작았던
    # 카테고리가 확 꺾여도(예: -90%) 튀어 보이는데, 실제 사업에 미치는 영향은 절대액이
    # 큰 카테고리(예: -15%지만 규모가 훨씬 큰 경우)가 더 클 수 있어서 둘 다 봐야 한다.
    if category_movers:
        _lines = []
        _all_cats_flat = []
        for _item in category_movers:
            _bpu_name, _tops, _bottoms = _item[0], _item[1], _item[2]
            _all_movers = _item[3] if len(_item) > 3 else []
            if _tops:
                _lines.append(f"{_bpu_name} 최대 상승(비율): <b>{_tops[0]['카테고리']}</b> ({_fmt_delta(_tops[0]['전년비'])})")
            if _bottoms:
                _lines.append(f"{_bpu_name} 최대 하락(비율): <b>{_bottoms[0]['카테고리']}</b> ({_fmt_delta(_bottoms[0]['전년비'])})")
            for _c in _all_movers:
                if _c.get("작년거래액") is not None:
                    _all_cats_flat.append({**_c, "BPU": _bpu_name, "절대변화": _c["거래액"] - _c["작년거래액"]})
        if _lines:
            sections.append({"title": "④ 카테고리 중 뭐가 성장을 끌고/깎아먹나? (비율 기준)", "body": "<br>".join(_lines)})

        if _all_cats_flat:
            _worst_abs = min(_all_cats_flat, key=lambda r: r["절대변화"])
            _best_abs = max(_all_cats_flat, key=lambda r: r["절대변화"])
            _abs_body = (
                f"절대액 기준 가장 큰 감소: <b>{_worst_abs['BPU']} · {_worst_abs['카테고리']}</b> "
                f"{_worst_abs['절대변화']:,.0f} ({_fmt_delta(_worst_abs['전년비'])})<br>"
                f"{_emphasize('비율은 작아 보여도 실제 매출 임팩트는 이쪽이 더 클 수 있어요.')}<br>"
                f"가장 큰 증가: <b>{_best_abs['BPU']} · {_best_abs['카테고리']}</b> +{_best_abs['절대변화']:,.0f}"
            )
            sections.append({"title": "④-2 절대 매출액 기준으로는 어디 영향이 가장 큰가?", "body": _abs_body})

    # 5) 쿠폰 비용 효율 (옵션)
    if coupon_stats and coupon_stats.get("비용률") is not None:
        _rate = coupon_stats["비용률"]
        _rate_prev = coupon_stats.get("비용률_전기간")
        if _rate_prev is not None:
            _dir = "상승" if _rate > _rate_prev else "하락"
            _body = f"비용률 {_rate:.2f}% (지난 기간 {_rate_prev:.2f}% 대비 {_dir}). 거래액 대비 쿠폰 지출 비중을 계속 지켜봐야 해요."
        else:
            _body = f"비용률 {_rate:.2f}%."
        sections.append({"title": "⑤ 쿠폰 비용 효율은 괜찮나?", "body": _body})

    # 6) 마감예상 vs 작년 (옵션)
    if forecast_stats and forecast_stats.get("yoy") is not None:
        _fy = forecast_stats["yoy"]
        _fm = forecast_stats.get("month", "이번 달")
        _tone = "양호해요" if _fy >= 0 else "작년보다 부진할 것으로 보여요"
        _body = f"{_fm} 마감예상 거래액이 작년 동월 대비 {_fmt_delta(_fy)} — {_tone}."
        sections.append({"title": "⑥ 이번 달 마감예상이 작년 대비 어떤가?", "body": _body})

    return sections


def generate_category_page_insights(cat_payload, cfg, cat_movers=None):
    """2번(카테고리 실적 요약) 페이지 전용 규칙 기반 인사이트. cat_payload는 KPI 카드용
    {"name","value","prev_delta","yoy_delta"} 리스트, cat_movers는 카테고리별
    {"카테고리","거래액","yoy"} 리스트(이미 표에 쓰는 것 그대로 재사용)."""
    sections = []
    _prev_label, _yoy_label = cfg.get("prev_label", "전기간비"), cfg.get("yoy_label", "전년비")

    def _fmt_delta(v):
        return format_delta_html(v) if v is not None else "-"

    if cat_payload:
        _bits = [f"{p['name']} {p['value']} ({_yoy_label} {_fmt_delta(p.get('yoy_delta'))})" for p in cat_payload[:3]]
        sections.append({"title": "① 이 카테고리·브랜드 조합의 기본 지표는?", "body": " · ".join(_bits)})

    if cat_movers:
        _valid = [r for r in cat_movers if r.get("yoy") is not None]
        if len(_valid) >= 2:
            _ranked = sorted(_valid, key=lambda r: r["yoy"], reverse=True)
            _top, _bottom = _ranked[0], _ranked[-1]
            _body = (
                f"최대 상승: <b>{_top['카테고리']}</b> {_top['거래액']:,.0f} ({_fmt_delta(_top['yoy'])})<br>"
                f"최대 하락: <b>{_bottom['카테고리']}</b> {_bottom['거래액']:,.0f} ({_fmt_delta(_bottom['yoy'])})"
            )
            sections.append({"title": "② 어느 카테고리가 성장을 끌고/깎아먹나? (비율 기준)", "body": _body})

        # 절대액 기준 — 비율만 보면 원래 작았던 카테고리가 크게 흔들려도(예: -90%) 튀어
        # 보이는데, 실제 매출 임팩트는 절대액이 큰 카테고리가 더 클 수 있어서 같이 본다.
        _valid_abs = [r for r in cat_movers if r.get("yoy_v") is not None]
        if len(_valid_abs) >= 2:
            _with_abs = [{**r, "절대변화": r["거래액"] - r["yoy_v"]} for r in _valid_abs]
            _worst_abs = min(_with_abs, key=lambda r: r["절대변화"])
            _best_abs = max(_with_abs, key=lambda r: r["절대변화"])
            _abs_body = (
                f"가장 큰 감소: <b>{_worst_abs['카테고리']}</b> {_worst_abs['절대변화']:,.0f} "
                f"({_fmt_delta(_worst_abs['yoy'])})<br>"
                f"{_emphasize('비율은 작아 보여도 실제 매출 임팩트는 이쪽이 더 클 수 있어요.')}<br>"
                f"가장 큰 증가: <b>{_best_abs['카테고리']}</b> "
                f"+{_best_abs['절대변화']:,.0f} ({_fmt_delta(_best_abs['yoy'])})"
            )
            sections.append({"title": "②-2 절대 매출액 기준으로는 어디 영향이 가장 큰가?", "body": _abs_body})
    return sections


def render_monthly_comparison_table(base_df, title, caption_extra=""):
    """26년|전년비|25년 월별 실적 비교표. base_df는 이미 원하는 BPU/카테고리/세그먼트로
    필터링된 df_traffic 형식(날짜/거래액/트래픽/구매객수 컬럼)이어야 한다. 일할계산 없이
    있는 만큼의 실제값만 쓰고, 당월(진행 중인 달)의 전년비만 '동요일 매칭'(26년에 실제
    존재하는 날짜들을 364일씩 당겨서 25년의 그 날짜들만 비교)으로 공정하게 계산한다."""
    st.markdown(f"**{title}**")
    if base_df is None or base_df.empty:
        st.info("데이터가 없습니다.")
        return

    _mc_abs_last = base_df["날짜"].max()
    _mc_cur_month = _mc_abs_last.month

    def _monthly_actual_series(year, num_col, den_col, is_ratio, scale=1.0):
        vals = []
        for m in range(1, 13):
            m_start = pd.Timestamp(year, m, 1)
            m_end = (m_start + pd.offsets.MonthBegin(1)) - pd.Timedelta(days=1)
            if m_start > _mc_abs_last:
                vals.append(None)
                continue
            md = base_df[(base_df["날짜"] >= m_start) & (base_df["날짜"] <= min(m_end, _mc_abs_last))]
            if md.empty:
                vals.append(None)
                continue
            if is_ratio:
                _den_sum = md[den_col].sum()
                vals.append((md[num_col].sum() / _den_sum * scale) if _den_sum else None)
            else:
                vals.append(md[num_col].sum())
        return vals

    def _cur_month_yoy_value_matched(num_col, den_col, is_ratio, scale=1.0):
        cur_start = pd.Timestamp(_mc_abs_last.year, _mc_cur_month, 1)
        cur_dates = base_df[(base_df["날짜"] >= cur_start) & (base_df["날짜"] <= _mc_abs_last)]["날짜"].unique()
        if len(cur_dates) == 0:
            return None, []
        matched = [pd.Timestamp(d) - pd.Timedelta(days=364) for d in cur_dates]
        md = base_df[base_df["날짜"].isin(matched)]
        if md.empty:
            return None, matched
        if is_ratio:
            _den_sum = md[den_col].sum()
            return ((md[num_col].sum() / _den_sum * scale) if _den_sum else None), matched
        return md[num_col].sum(), matched

    _mc_metric_defs = [
        ("거래액", "거래액", None, False, 1.0),
        ("트래픽", "트래픽", None, False, 1.0),
        ("구매객수", "구매객수", None, False, 1.0),
        ("CR", "구매객수", "트래픽", True, 100.0),
        ("객단가", "거래액", "구매객수", True, 1.0),
    ]

    def _mc_fmt(v, is_pct, is_billion=False):
        if v is None or pd.isna(v):
            return "-"
        if is_pct:
            return f"{v:.1f}%"
        if is_billion:
            return f"{v / 100_000_000:,.2f}억"
        return f"{v:,.0f}"

    _mc_last_day_label = f"~{_mc_abs_last.month}/{_mc_abs_last.day}"
    # 당월 강조는 배경색 대신 테두리로 — 좌우는 얇게(2px), 위/아래는 굵게(4px)해서
    # 컬럼 전체(헤더 2줄 + 데이터 행 전부)를 감싼다.
    _MC_CUR_HL = "border-left:2px solid #f59e0b;border-right:2px solid #f59e0b;"
    _MC_CUR_HL_TOP = _MC_CUR_HL + "border-top:4px solid #f59e0b;"
    _MC_CUR_HL_BOTTOM = _MC_CUR_HL + "border-bottom:4px solid #f59e0b;"

    _, _mc_matched_dates = _cur_month_yoy_value_matched("거래액", None, False)
    if _mc_matched_dates:
        _mc_25_start, _mc_25_end = min(_mc_matched_dates), max(_mc_matched_dates)
        _mc_25_day_label = f"{_mc_25_start.month}/{_mc_25_start.day}~{_mc_25_end.month}/{_mc_25_end.day}"
    else:
        _mc_25_day_label = _mc_last_day_label

    def _mc_th(m, day_label=None):
        _day_label = day_label if day_label is not None else _mc_last_day_label
        _hl = _MC_CUR_HL_TOP if m == _mc_cur_month else ""
        _lbl = f"{m}월{f'({_day_label})' if m == _mc_cur_month else ''}"
        return f"<th style='white-space:nowrap;{_hl}'>{_lbl}</th>"

    _mc_month_headers_26 = "".join(_mc_th(m) for m in range(1, _mc_cur_month + 1))
    _mc_month_headers_yoy = "".join(_mc_th(m) for m in range(1, _mc_cur_month + 1))
    _mc_month_headers_25 = "".join(_mc_th(m, _mc_25_day_label) for m in range(1, _mc_cur_month + 1))

    def _mc_td(v, m, is_pct, is_billion=False, is_delta=False, is_last_row=False):
        if m == _mc_cur_month:
            _hl = _MC_CUR_HL_BOTTOM if is_last_row else _MC_CUR_HL
        else:
            _hl = ""
        _content = format_delta_html(v) if is_delta else _mc_fmt(v, is_pct, is_billion)
        return f"<td style='text-align:right;white-space:nowrap;{_hl}'>{_content}</td>"

    _mc_rows_html = ""
    for _mc_row_idx, (_mc_label, _mc_num, _mc_den, _mc_is_ratio, _mc_scale) in enumerate(_mc_metric_defs):
        _mc_is_pct = _mc_label == "CR"
        _mc_is_billion = _mc_label == "거래액"
        _mc_is_last_row = _mc_row_idx == len(_mc_metric_defs) - 1
        _v26 = _monthly_actual_series(2026, _mc_num, _mc_den, _mc_is_ratio, _mc_scale)
        _v25 = _monthly_actual_series(2025, _mc_num, _mc_den, _mc_is_ratio, _mc_scale)
        _v25[_mc_cur_month - 1], _ = _cur_month_yoy_value_matched(_mc_num, _mc_den, _mc_is_ratio, _mc_scale)
        _cells_26 = "".join(_mc_td(_v26[m - 1], m, _mc_is_pct, _mc_is_billion, is_last_row=_mc_is_last_row) for m in range(1, _mc_cur_month + 1))
        _cells_25 = "".join(_mc_td(_v25[m - 1], m, _mc_is_pct, _mc_is_billion, is_last_row=_mc_is_last_row) for m in range(1, _mc_cur_month + 1))
        _cells_yoy = ""
        for m in range(1, _mc_cur_month + 1):
            _yoy = pct_delta_safe(_v26[m - 1], _v25[m - 1]) if (_v26[m - 1] is not None and _v25[m - 1]) else None
            _cells_yoy += _mc_td(_yoy, m, _mc_is_pct, is_delta=True, is_last_row=_mc_is_last_row)
        _mc_rows_html += (
            f"<tr class='mc-row' data-i='{_mc_row_idx}'><td class='m' style='white-space:nowrap;'>{_mc_label}</td>"
            f"{_cells_26}{_cells_yoy}{_cells_25}</tr>"
        )

    _mc_n_rows = len(_mc_metric_defs)
    _mc_n_cols = 1 + _mc_cur_month * 3
    _mc_frame_h = 46 + _mc_n_rows * 24

    # 클릭 강조가 되려면 JS가 필요해서 components.html(iframe)로 렌더링한다 — iframe은
    # 부모 문서의 전역 CSS를 상속받지 않으므로, summary-table/delta 클래스 스타일을
    # 이 문서 안에 직접 넣어준다(styles.py의 정의와 동일하게 맞춤).
    _mc_doc = f"""
<html><head><style>
  body {{ margin:0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; }}
  .summary-table {{ width:100%; border-collapse:collapse; font-size:0.72rem; background:#fff;
    border:1px solid #eaecef; border-radius:8px; overflow:hidden; table-layout:auto; }}
  .summary-table thead th {{ background:#f7f8fa; color:#6b7280; font-weight:600; text-align:left;
    padding:4px 5px; border-bottom:1px solid #eaecef; font-size:0.66rem; }}
  .summary-table tbody td {{ padding:4px 5px; border-bottom:1px solid #f1f2f4; color:#111827; }}
  .summary-table tbody tr:last-child td {{ border-bottom:none; }}
  .summary-table td.m {{ font-weight:500; }}
  .delta.up {{ color:#16a34a; font-weight:600; }}
  .delta.down {{ color:#dc2626; font-weight:600; }}
  .delta.neutral {{ color:#9ca3af; font-weight:600; }}
  .mc-row {{ cursor:pointer; transition:background .15s; }}
  .mc-row:hover {{ background:#f8fafc; }}
  .mc-row.sel {{ background:#eff6ff; }}
  .mc-row.sel td.m {{ color:#2563eb; font-weight:700; }}
</style></head><body>
  <div style="overflow-x:auto;"><table class="summary-table">
    <thead>
      <tr><th rowspan="2" style="white-space:nowrap;">구분</th><th colspan="{_mc_cur_month}" style="text-align:center;background:#eef2ff;white-space:nowrap;">26년</th>
      <th colspan="{_mc_cur_month}" style="text-align:center;background:#fef3c7;white-space:nowrap;">전년비</th>
      <th colspan="{_mc_cur_month}" style="text-align:center;background:#f3f4f6;white-space:nowrap;">25년</th></tr>
      <tr>{_mc_month_headers_26}{_mc_month_headers_yoy}{_mc_month_headers_25}</tr>
    </thead>
    <tbody>{_mc_rows_html}</tbody>
  </table></div>
<script>
(function() {{
  var rows = Array.prototype.slice.call(document.querySelectorAll('.mc-row'));
  rows.forEach(function(r) {{
    r.addEventListener('click', function() {{ r.classList.toggle('sel'); }});
  }});
}})();
</script>
</body></html>
"""
    components.html(_mc_doc, height=_mc_frame_h, scrolling=True)
    _cap = "일할계산(마감예상) 없이, 진행 중인 달은 있는 날짜까지의 실제값만 보여줘요."
    if caption_extra:
        _cap += f" {caption_extra}"
    st.caption(_cap)


def render_insight_panel(sections, key_prefix=""):
    """generate_rule_based_insights() 결과를 카드 형태로 렌더링."""
    if not sections:
        return
    st.markdown(
        "<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:16px 20px;'>"
        "<div style='font-weight:700;font-size:0.95rem;margin-bottom:4px;'>🔍 자동 인사이트</div>"
        "<div style='font-size:0.76rem;color:#9ca3af;margin-bottom:12px;'>현재 조회조건 기준으로 자동 정리돼요 (계산된 값 그대로 조립 — AI 호출 없음).</div>"
        + "".join(
            f"<div style='margin-bottom:10px;'>"
            f"<div style='font-weight:600;font-size:0.85rem;color:#374151;margin-bottom:2px;'>{s['title']}</div>"
            f"<div style='font-size:0.82rem;color:#4b5563;line-height:1.5;'>{s['body']}</div>"
            f"</div>"
            for s in sections
        )
        + "</div>",
        unsafe_allow_html=True,
    )


def render_donut_chart(labels, values, colors=None, center_title="", center_value="", size=300,
                       deltas=None, delta_label="", prev_delta_label="", center_sub=""):
    """클릭 가능한 SVG 도넛 차트.
    - 조각이나 범례를 클릭하면 해당 항목이 강조되고 나머지는 흐려진다(다시 클릭하면 해제).
    - deltas가 주어지면 범례에 항목별 전년 거래액/전년비 증감을 함께 표시한다.
    - JS 동작이 필요하므로 components.html(iframe)로 렌더링한다."""
    import math
    import html as _html

    total = sum(v for v in values if v and v > 0)
    if total <= 0:
        st.info("표시할 데이터가 없습니다.")
        return

    palette = colors or [
        "#2563eb", "#3b82f6", "#60a5fa", "#93c5fd", "#7dd3fc",
        "#38bdf8", "#0ea5e9", "#0284c7", "#a5b4fc", "#c7d2fe", "#e2e8f0",
    ]
    NEG_COLOR = "#ef4444"  # 마이너스(반품 등) 항목은 도넛 조각엔 안 넣지만, 범례엔 이 색으로 표시
    cx = cy = size / 2
    r_outer = size / 2 - 6
    r_inner = r_outer * 0.62

    def _pt(r, ang):
        rad = math.radians(ang - 90)
        return cx + r * math.cos(rad), cy + r * math.sin(rad)

    paths = []
    start = 0.0
    visible_idx = []
    for i, (lab, val) in enumerate(zip(labels, values)):
        if not val or val <= 0:
            continue
        visible_idx.append(i)
        frac = val / total
        end = start + frac * 360
        mid = (start + end) / 2
        # 선택 시 바깥으로 살짝 튀어나오는 오프셋
        _mrad = math.radians(mid - 90)
        pop = f"translate({math.cos(_mrad) * 9:.2f}px, {math.sin(_mrad) * 9:.2f}px)"
        tip = _html.escape(f"{lab}: {val:,.0f} ({frac * 100:.1f}%)")
        color = palette[i % len(palette)]

        if frac >= 0.9999:
            paths.append(
                f"<circle class='seg' data-i='{i}' data-pop='{pop}' cx='{cx}' cy='{cy}' "
                f"r='{(r_outer + r_inner) / 2:.2f}' fill='none' stroke='{color}' "
                f"stroke-width='{r_outer - r_inner:.2f}'><title>{tip}</title></circle>"
            )
            start = end
            continue

        large = 1 if (end - start) > 180 else 0
        x1, y1 = _pt(r_outer, start)
        x2, y2 = _pt(r_outer, end)
        x3, y3 = _pt(r_inner, end)
        x4, y4 = _pt(r_inner, start)
        d = (
            f"M {x1:.2f} {y1:.2f} A {r_outer:.2f} {r_outer:.2f} 0 {large} 1 {x2:.2f} {y2:.2f} "
            f"L {x3:.2f} {y3:.2f} A {r_inner:.2f} {r_inner:.2f} 0 {large} 0 {x4:.2f} {y4:.2f} Z"
        )
        paths.append(
            f"<path class='seg' data-i='{i}' data-pop='{pop}' d='{d}' fill='{color}' "
            f"stroke='#fff' stroke-width='1.5'><title>{tip}</title></path>"
        )
        start = end

    center_html = ""
    if center_title or center_value or center_sub:
        _cy_base = cy - 14 if center_sub else cy - 6
        center_html = (
            f"<text x='{cx}' y='{_cy_base}' text-anchor='middle' font-size='11' fill='#6b7280' "
            f"style='pointer-events:none'>{_html.escape(center_title)}</text>"
            f"<text x='{cx}' y='{_cy_base + 20}' text-anchor='middle' font-size='15' font-weight='700' "
            f"fill='#111827' style='pointer-events:none'>{_html.escape(center_value)}</text>"
        )
        if center_sub:
            center_html += (
                f"<text x='{cx}' y='{_cy_base + 38}' text-anchor='middle' font-size='10.5' "
                f"fill='#6b7280' style='pointer-events:none'>{_html.escape(center_sub)}</text>"
            )

    _has_delta = deltas is not None
    header_html = ""
    if _has_delta:
        _prev_col_html = (
            f"<span style='width:96px;text-align:right;'>{_html.escape((prev_delta_label or '전기간').replace('비',''))}</span>"
            f"<span style='width:70px;text-align:right;'>{_html.escape(prev_delta_label or '전기간비')}</span>"
            if prev_delta_label else ""
        )
        header_html = (
            "<div class='lg-head'>"
            "<span style='width:17px;flex-shrink:0;'></span>"
            "<span style='flex:1;'>항목</span>"
            "<span style='width:96px;text-align:right;'>올해</span>"
            "<span style='width:44px;text-align:right;'>비중</span>"
            "<span style='width:96px;text-align:right;'>작년</span>"
            f"<span style='width:70px;text-align:right;'>{_html.escape(delta_label or '전년비')}</span>"
            f"{_prev_col_html}"
            "</div>"
        )

    def _delta_span(v):
        if v is None:
            return "<span style='color:#9ca3af'>-</span>"
        color = "#16a34a" if v >= 0 else "#dc2626"
        text = f"{abs(v):.1f}%" if v >= 0 else f"△ {abs(v):.1f}%"
        return f"<span style='color:{color};font-weight:600'>{text}</span>"

    legend_items = []
    for i, (lab, val) in enumerate(zip(labels, values)):
        if not val:
            continue
        is_neg = val < 0
        pct = val / total * 100
        _dot_color = NEG_COLOR if is_neg else palette[i % len(palette)]
        row = (
            f"<div class='lg' data-i='{i}'>"
            f"<span class='dot' style='background:{_dot_color};'></span>"
            f"<span style='flex:1;color:#374151;'>{_html.escape(str(lab))}</span>"
        )
        if _has_delta:
            _d = deltas[i] if i < len(deltas) else None
            _prev_val = _d.get("prev") if isinstance(_d, dict) else None
            _yoy_val = _d.get("yoy") if isinstance(_d, dict) else None
            _prev_delta_val = _d.get("prev_delta") if isinstance(_d, dict) else None
            _prev_delta_value_val = _d.get("prev_delta_value") if isinstance(_d, dict) else None
            _prev_str = f"{_prev_val:,.0f}" if _prev_val is not None else "-"
            _prev_delta_value_str = f"{_prev_delta_value_val:,.0f}" if _prev_delta_value_val is not None else "-"
            _prev_delta_col = (
                f"<span style='color:#9ca3af;width:96px;text-align:right;'>{_prev_delta_value_str}</span>"
                f"<span style='width:70px;text-align:right;'>{_delta_span(_prev_delta_val)}</span>"
                if prev_delta_label else ""
            )
            row += (
                f"<span style='color:#374151;width:96px;text-align:right;'>{val:,.0f}</span>"
                f"<span style='color:#9ca3af;width:44px;text-align:right;'>{pct:.1f}%</span>"
                f"<span style='color:#9ca3af;width:96px;text-align:right;'>{_prev_str}</span>"
                f"<span style='width:70px;text-align:right;'>{_delta_span(_yoy_val)}</span>"
                f"{_prev_delta_col}"
            )
        else:
            row += (
                f"<span style='color:#6b7280;margin-left:10px;'>{val:,.0f}</span>"
                f"<span style='color:#9ca3af;margin-left:8px;width:48px;text-align:right;'>{pct:.1f}%</span>"
            )
        row += "</div>"
        legend_items.append(row)

    n_rows = len(legend_items)
    frame_h = int(max(size + 46, 70 + n_rows * 25 + (24 if _has_delta else 0)))
    legend_min = (606 if prev_delta_label else 440) if _has_delta else 260

    doc = f"""
<!DOCTYPE html><html><head><meta charset='utf-8'><style>
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font-family:'Source Sans Pro','Apple SD Gothic Neo','Malgun Gothic',sans-serif; }}
  .wrap {{ display:flex; align-items:center; gap:24px; flex-wrap:wrap;
          background:#fff; border:1px solid #e5e7eb; border-radius:10px; padding:16px 20px; }}
  svg {{ flex-shrink:0; }}
  .seg {{ cursor:pointer; transition:opacity .15s ease, transform .15s ease; }}
  .seg.dim {{ opacity:.22; }}
  .legend {{ flex:1; min-width:{legend_min}px; }}
  .lg-head {{ display:flex; align-items:center; font-size:11px; color:#9ca3af;
              border-bottom:1px solid #f1f2f4; padding-bottom:4px; margin-bottom:6px; }}
  .lg {{ display:flex; align-items:center; margin-bottom:4px; font-size:12.5px;
         cursor:pointer; border-radius:5px; padding:2px 5px; transition:opacity .15s, background .15s; }}
  .lg:hover {{ background:#f8fafc; }}
  .lg.sel {{ background:#eef2ff; font-weight:600; }}
  .lg.dim {{ opacity:.4; }}
  .dot {{ display:inline-block; width:10px; height:10px; border-radius:2px;
          margin-right:7px; flex-shrink:0; }}
  .hint {{ font-size:10.5px; color:#9ca3af; margin-top:8px; }}
</style></head><body>
<div class='wrap'>
  <svg width='{size}' height='{size}' viewBox='0 0 {size} {size}'>{''.join(paths)}{center_html}</svg>
  <div class='legend'>{header_html}{''.join(legend_items)}
    <div class='hint'>💡 조각이나 항목을 클릭하면 해당 항목만 강조됩니다 (다시 클릭하면 해제)</div>
  </div>
</div>
<script>
(function() {{
  var sel = null;
  var segs = Array.prototype.slice.call(document.querySelectorAll('.seg'));
  var lgs  = Array.prototype.slice.call(document.querySelectorAll('.lg'));
  function apply() {{
    segs.forEach(function(s) {{
      var i = parseInt(s.getAttribute('data-i'), 10);
      if (sel === null) {{ s.classList.remove('dim'); s.style.transform = ''; }}
      else if (i === sel) {{ s.classList.remove('dim'); s.style.transform = s.getAttribute('data-pop'); }}
      else {{ s.classList.add('dim'); s.style.transform = ''; }}
    }});
    lgs.forEach(function(l) {{
      var i = parseInt(l.getAttribute('data-i'), 10);
      l.classList.toggle('sel', sel === i);
      l.classList.toggle('dim', sel !== null && i !== sel);
    }});
  }}
  function toggle(i) {{ sel = (sel === i) ? null : i; apply(); }}
  segs.forEach(function(s) {{
    s.addEventListener('click', function() {{ toggle(parseInt(s.getAttribute('data-i'), 10)); }});
  }});
  lgs.forEach(function(l) {{
    l.addEventListener('click', function() {{ toggle(parseInt(l.getAttribute('data-i'), 10)); }});
  }});
}})();
</script></body></html>
"""
    components.html(doc, height=frame_h, scrolling=False)


def compute_official_total(df_scope, unit, selected_period_date, metric_col="거래액"):
    """df_scope(보통 카테고리='전체'&브랜드='전체' 등으로 필터된 단일 그룹)의
    metric_col(기본 거래액) 값을 조회단위로 리샘플하여 (현재값, 전년동기값) 튜플로 반환.
    도넛 중앙에 표시할 '진짜 전체값'을 개별 항목 합산과 별개로 정확히 구하기 위함.
    compute_kpi_deltas를 그대로 재사용해서, 진행 중인(부분) 달/주에도 KPI 카드·요약표와
    똑같은 기준(같은 날짜 수만큼 동요일 매칭)으로 전년비가 계산되도록 한다."""
    if df_scope.empty:
        return None, None
    s_full = df_scope.set_index("날짜")[metric_col].sort_index()
    _agg = "sum" if unit == "월마감" else "mean"
    series_full = s_full.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
    if unit == "주별":
        series_full.index = series_full.index - pd.Timedelta(days=6)
    elif unit == "월마감" and not series_full.empty and s_full.index.max() < series_full.index[-1]:
        series_full = series_full.iloc[:-1]
    series = series_full[series_full.index <= selected_period_date] if not series_full.empty else series_full
    if series.empty:
        return None, None
    _s_raw = s_full[s_full.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s_full
    stats = compute_kpi_deltas(series, unit, raw_daily=_s_raw)
    if stats is None:
        return None, None
    return stats["current"], stats.get("yoy_value")


def render_revenue_ranking(sub_df, group_col, unit, selected_period_date, title, subtitle, label_map=None, hide_zero=False, ai_key=None, ai_context=None, donut=False, official_total=None, metric_col="거래액", metric_label="거래액", bar_color_cur="#2563eb", bar_color_prev="#7dd3fc", donut_colors=None):
    """
    official_total: (현재값, 작년값) 튜플이 주어지면, 도넛 중앙의 '총 {지표}'를
    개별 항목 합산이 아니라 이 값으로 표시한다. (카테고리/브랜드가 여러 개 겹치는 거래는
    개별 항목 합산이 실제 전체보다 커질 수 있어, KPI 카드와 항상 일치시키기 위함)

    group_col(카테고리 또는 브랜드) 기준 값 랭킹을 올해/작년 이중 막대로 렌더링.
    label_map이 주어지면 표시 라벨을 매핑해서 보여준다 (예: 브랜드코드 -> 브랜드명).
    hide_zero=True면 올해/작년 값이 둘 다 0(또는 0에 가까움)인 항목은 목록에서 제외.
    ai_key가 주어지면 'AI 인사이트' 버튼과 결과 박스를 함께 표시한다.
    """
    rows = []
    for name in sorted(sub_df[group_col].unique()):
        s_full = sub_df[sub_df[group_col] == name].set_index("날짜")[metric_col].sort_index()
        _agg = "sum" if unit == "월마감" else "mean"
        series_full = s_full.resample(UNIT_CONFIG[unit]["rule"]).agg(_agg)
        if unit == "주별":
            series_full.index = series_full.index - pd.Timedelta(days=6)
        elif unit == "월마감" and not series_full.empty and s_full.index.max() < series_full.index[-1]:
            series_full = series_full.iloc[:-1]
        series = series_full[series_full.index <= selected_period_date] if not series_full.empty else series_full
        if series.empty:
            continue
        _s_raw = s_full[s_full.index <= raw_cutoff_date(selected_period_date, unit)] if selected_period_date is not None else s_full
        # compute_kpi_deltas 재사용 — KPI 카드·요약표와 동일 기준(부분월이면 동요일 매칭)으로 전년비 계산
        _stats = compute_kpi_deltas(series, unit, raw_daily=_s_raw)
        if _stats is None:
            continue
        cur_val = _stats["current"]
        prev_val = _stats.get("yoy_value")

        if hide_zero:
            _cur_zero = pd.isna(cur_val) or abs(cur_val) < 0.5
            _prev_zero = prev_val is None or pd.isna(prev_val) or abs(prev_val) < 0.5
            if _cur_zero and _prev_zero:
                continue

        rows.append({
            group_col: name, "값": cur_val, "전년값": prev_val,
            "전기간값": _stats.get("prev_value"), "전기간비": _stats.get("prev_delta"),
        })

    if ai_key:
        _rk_c1, _rk_c2 = st.columns([4, 1])
        with _rk_c1:
            st.markdown(f"**{title}**  ·  <span style='color:#6b7280;font-size:0.85rem'>{subtitle}</span>", unsafe_allow_html=True)
        with _rk_c2:
            _rk_clicked = st.button("🤖 AI 인사이트", key=f"ai_btn_{ai_key}", use_container_width=True)
    else:
        st.markdown(f"**{title}**  ·  <span style='color:#6b7280;font-size:0.85rem'>{subtitle}</span>", unsafe_allow_html=True)
        _rk_clicked = False

    if not rows:
        st.info("해당 조건에 데이터가 없습니다.")
        return

    share_df = pd.DataFrame(rows).sort_values("값", ascending=False).reset_index(drop=True)
    _total_gmv = share_df["값"].sum()
    share_df["비중"] = (share_df["값"] / _total_gmv * 100) if _total_gmv > 0 else 0
    _max_gmv = max(
        share_df["값"].max() if not share_df.empty else 1,
        share_df["전년값"].max(skipna=True) if share_df["전년값"].notna().any() else 0,
        1,
    )
    _yoy_label_share = UNIT_CONFIG[unit]["yoy_label"]
    _prev_label_share = UNIT_CONFIG[unit]["prev_label"]

    # --- AI 인사이트 (요청 시) ---
    if ai_key:
        _ai_rank_result = None
        _rk_ctx_key = f"ai_rank_ctx_{ai_key}"
        _rk_cur_ctx = ai_context or subtitle
        if _rk_clicked:
            _rank_payload = []
            for _, r in share_df.iterrows():
                _nm = r[group_col]
                _nm_disp = label_map.get(_nm, _nm) if label_map else _nm
                _yv = None
                if pd.notna(r["전년값"]) and r["전년값"] != 0:
                    _yv = pct_delta_safe(r["값"], r["전년값"])
                _rank_payload.append({
                    "name": str(_nm_disp),
                    "current": float(r["값"]) if pd.notna(r["값"]) else 0.0,
                    "prev": float(r["전년값"]) if pd.notna(r["전년값"]) else 0.0,
                    "share": float(r["비중"]),
                    "yoy": round(float(_yv), 1) if _yv is not None else None,
                })
            with st.spinner("AI 인사이트 생성 중..."):
                _ai_rank_result = generate_ranking_insights(
                    _rank_payload, ai_context or subtitle, f"rank_{ai_key}"
                )
                st.session_state[f"ai_rank_result_{ai_key}"] = _ai_rank_result
                st.session_state[_rk_ctx_key] = _rk_cur_ctx
        elif (
            f"ai_rank_result_{ai_key}" in st.session_state
            and st.session_state.get(_rk_ctx_key) == _rk_cur_ctx
        ):
            _ai_rank_result = st.session_state[f"ai_rank_result_{ai_key}"]
        else:
            # 매체/세그먼트/조회단위 등 조회 조건이 바뀌어서 이전 인사이트가 더 이상
            # 맞지 않으므로 캐시를 지운다 (버튼을 다시 눌러야 새 조건으로 재생성됨).
            st.session_state.pop(f"ai_rank_result_{ai_key}", None)
            st.session_state.pop(_rk_ctx_key, None)
        render_ranking_insight_box(_ai_rank_result)

    bar_rows_html = []
    for _, r in share_df.iterrows():
        _pct_width = max(0, (r["값"] / _max_gmv * 100)) if _max_gmv > 0 else 0
        _has_prev = pd.notna(r["전년값"])
        _prev_pct_width = max(0, (r["전년값"] / _max_gmv * 100)) if _has_prev and _max_gmv > 0 else 0
        _yoy_delta = pct_delta_safe(r["값"], r["전년값"]) if _has_prev and r["전년값"] != 0 else None
        _prev_val_str = f"{r['전년값']:,.0f}" if _has_prev else "-"

        _raw_label = r[group_col]
        _display_label = label_map.get(_raw_label, _raw_label) if label_map else _raw_label
        _label_width = 190 if label_map else 80

        _prev_delta_val = r.get("전기간비")
        _prev_delta_str = f" · {_prev_label_share} {format_delta_html(_prev_delta_val)}" if pd.notna(_prev_delta_val) else ""

        bar_rows_html.append(
            "<div style='margin-bottom:12px;'>"
            "<div style='display:flex;align-items:center;margin-bottom:3px;'>"
            f"<div style='width:{_label_width}px;flex-shrink:0;font-size:0.8rem;color:#374151;font-weight:600;'>{_display_label}</div>"
            "<div style='flex:1;background:#f1f2f4;border-radius:4px;height:20px;margin:0 10px;position:relative;'>"
            f"<div style='width:{_pct_width:.1f}%;background:{bar_color_cur};height:100%;border-radius:4px;'></div>"
            "</div>"
            f"<div style='width:190px;flex-shrink:0;text-align:right;font-size:0.82rem;color:#374151;'>"
            f"{r['값']:,.0f} <span style='color:#9ca3af'>({r['비중']:.1f}%)</span>{_prev_delta_str}</div>"
            "</div>"
            "<div style='display:flex;align-items:center;'>"
            f"<div style='width:{_label_width}px;flex-shrink:0;'></div>"
            "<div style='flex:1;background:#f1f2f4;border-radius:4px;height:14px;margin:0 10px;position:relative;'>"
            f"<div style='width:{_prev_pct_width:.1f}%;background:{bar_color_prev};height:100%;border-radius:4px;'></div>"
            "</div>"
            f"<div style='width:190px;flex-shrink:0;text-align:right;font-size:0.76rem;color:#9ca3af;'>"
            f"{_prev_val_str}{f' · {_yoy_label_share} ' + format_delta_html(_yoy_delta) if _yoy_delta is not None else ''}</div>"
            "</div>"
            "</div>"
        )
    # --- 도넛 차트 모드: 구성비를 한눈에 + 전년비 상세는 접이식 ---
    if donut:
        _top_n = 10
        _dn_pos = share_df[share_df["값"] > 0].copy()
        _dn_neg = share_df[share_df["값"] < 0].copy()
        _labels, _values, _deltas = [], [], []
        for _, r in _dn_pos.head(_top_n).iterrows():
            _nm = r[group_col]
            _labels.append(str(label_map.get(_nm, _nm) if label_map else _nm))
            _values.append(float(r["값"]))
            _pv = float(r["전년값"]) if pd.notna(r["전년값"]) else None
            _yv = pct_delta_safe(r["값"], r["전년값"]) if (_pv is not None and r["전년값"] != 0) else None
            _deltas.append({"prev": _pv, "yoy": _yv, "prev_delta": r.get("전기간비"), "prev_delta_value": r.get("전기간값")})

        _rest_df = _dn_pos.iloc[_top_n:] if len(_dn_pos) > _top_n else None
        if _rest_df is not None and len(_rest_df) > 0:
            _rest_cur = float(_rest_df["값"].sum())
            if _rest_cur > 0:
                _rest_prev = float(_rest_df["전년값"].sum(skipna=True)) if _rest_df["전년값"].notna().any() else None
                _rest_yoy = pct_delta_safe(_rest_cur, _rest_prev) if (_rest_prev and _rest_prev != 0) else None
                _labels.append(f"기타 ({len(_rest_df)}개)")
                _values.append(_rest_cur)
                _deltas.append({"prev": _rest_prev, "yoy": _rest_yoy, "prev_delta": None, "prev_delta_value": None})

        # 값이 마이너스인 항목(반품 등으로 순값이 음수)도 도넛에 그대로 포함시킨다.
        # top_n/기타 묶음과는 무관하게 항상 개별 조각으로 넣어서 묻히지 않게 함
        # (render_donut_chart 쪽에서 절대값 크기로 조각을 그리고 빨간 점선으로 구분해서 보여줌).
        for _, r in _dn_neg.sort_values("값").iterrows():
            _nm = r[group_col]
            _labels.append(str(label_map.get(_nm, _nm) if label_map else _nm))
            _values.append(float(r["값"]))
            _pv = float(r["전년값"]) if pd.notna(r["전년값"]) else None
            _yv = pct_delta_safe(r["값"], r["전년값"]) if (_pv is not None and r["전년값"] != 0) else None
            _deltas.append({"prev": _pv, "yoy": _yv, "prev_delta": r.get("전기간비"), "prev_delta_value": r.get("전기간값")})

        # 전체 합계 기준 전년비 — official_total(카테고리=전체 등 진짜 전체값)이 있으면 그걸 우선 사용.
        # (개별 항목을 단순 합산하면, 여러 카테고리에 걸친 거래가 중복 집계되어
        #  KPI 카드의 진짜 전체값보다 커질 수 있음 — 그래서 중앙 표시는 항상 official_total과 일치시킴)
        if official_total is not None:
            _tot_cur, _tot_prev = official_total
            _tot_cur = float(_tot_cur) if _tot_cur is not None else float(share_df["값"].sum())
        else:
            _tot_cur = float(share_df["값"].sum())
            _tot_prev = float(share_df["전년값"].sum(skipna=True)) if share_df["전년값"].notna().any() else None
        if _tot_prev and _tot_prev != 0:
            _tot_yoy = pct_delta_safe(_tot_cur, _tot_prev)
            _center_sub = f"{_yoy_label_share} {_tot_yoy:+.1f}%"
        else:
            _center_sub = ""

        render_donut_chart(
            _labels, _values,
            colors=donut_colors,
            center_title=f"총 {metric_label}",
            center_value=f"{_tot_cur:,.0f}",
            center_sub=_center_sub,
            deltas=_deltas,
            delta_label=_yoy_label_share,
            prev_delta_label=_prev_label_share,
        )
        if official_total is not None:
            st.caption(
                f"ℹ️ 중앙 '총 {metric_label}'은 KPI카드와 동일한 전체 집계값이에요. "
                "여러 카테고리/브랜드에 걸친 거래가 있으면, 아래 항목별 값을 다 더한 합계는 "
                "이 값과 정확히 일치하지 않을 수 있어요(항목 간 중복 집계 가능)."
            )
        if len(_dn_neg) > 0:
            st.caption(
                "🔴 빨간 점 항목은 값이 마이너스예요(반품 등으로 환불이 매출보다 많은 경우). "
                "도넛 조각으로는 표시되지 않고, 오른쪽 목록에 실제 마이너스 값 그대로 표시돼요."
            )

        with st.expander(f"📊 전체 항목 · 전년 대비 막대로 보기 ({_yoy_label_share})", expanded=False):
            st.markdown(
                "<div style='display:flex;gap:14px;margin-bottom:10px;font-size:0.76rem;color:#6b7280;'>"
                f"<span><span style='display:inline-block;width:10px;height:10px;background:{bar_color_cur};border-radius:2px;margin-right:4px;'></span>올해</span>"
                f"<span><span style='display:inline-block;width:10px;height:10px;background:{bar_color_prev};border-radius:2px;margin-right:4px;'></span>작년(동시점)</span>"
                "</div>" + "".join(bar_rows_html),
                unsafe_allow_html=True,
            )
        return

    st.markdown(
        "<div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:16px 20px;'>"
        "<div style='display:flex;gap:14px;margin-bottom:10px;font-size:0.76rem;color:#6b7280;'>"
        f"<span><span style='display:inline-block;width:10px;height:10px;background:{bar_color_cur};border-radius:2px;margin-right:4px;'></span>올해</span>"
        f"<span><span style='display:inline-block;width:10px;height:10px;background:{bar_color_prev};border-radius:2px;margin-right:4px;'></span>작년(동시점)</span>"
        "</div>"
        + "".join(bar_rows_html) +
        "</div>",
        unsafe_allow_html=True,
    )
